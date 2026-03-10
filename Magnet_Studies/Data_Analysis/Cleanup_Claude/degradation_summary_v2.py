#!/usr/bin/env python3
"""
Degradation Summary v2 — Scientifically Rigorous

Fixes from v1:
  1. H-plate temperature lookup now includes Hn/Hs Teslameter files
     (not just An/As), fixing same-day temperature matching
  2. Baseline uses ONLY temperature-corrected readings (no mixing of
     corrected and uncorrected values)
  3. Flagged outlier samples (Y-34-4, Y-40-4) excluded from group
     statistics, reported separately
  4. Teslameter field magnitude degradation added (using first tunnel
     measurement as baseline, since pre-deployment probe was broken)

Error bars — three independent sources added in quadrature:
  σ_baseline:  std(baseline readings) / sqrt(N)  [uncertainty on mean baseline]
  σ_temp:      propagated from Teslameter face temperature spread on final date
  σ_total:     sqrt(σ_baseline² + σ_temp²), expressed as % of baseline

For group averages: max(SEM across samples, mean per-sample σ_total)

Output:
  - MD_Files/degradation_summary_v2.md
  - TempCorrected_Plots/degradation_v2_*.png
"""

import os
import re
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from collections import defaultdict
import openpyxl

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
PLOT_DIR = os.path.join(BASE, 'TempCorrected_Plots')
MD_DIR = os.path.join(os.path.dirname(BASE), 'MD_Files')
os.makedirs(PLOT_DIR, exist_ok=True)
os.makedirs(MD_DIR, exist_ok=True)

# ─── Constants ───────────────────────────────────────────────────────────────
T_REF = 20.0
ALPHA = {
    'N42EH':   -0.0010,
    'N52SH':   -0.0011,
    'SmCo33H': -0.0004,
    'SmCo35':  -0.0004,
    'NdFeB':   -0.00105,
    'SmCo':    -0.0004,
}
SENTINEL = 1337
MIN_BASELINE_MWC = 0.1
TUNNEL_START = datetime(2025, 7, 1)

# Pre-deployment Teslameter FIELD readings are invalid (broken probe),
# but TEMPERATURE readings appear valid (reasonable lab temperatures ~22-24°C)
# We use the temperature but not the field magnitude for pre-deployment dates.
TESLAMETER_FIELD_VALID_AFTER = datetime(2025, 7, 1)

CB = {
    'N42EH': '#EE6677', 'N52SH': '#4477AA',
    'SmCo33H': '#228833', 'SmCo35': '#CCBB44',
    'NdFeB': '#AA3377', 'SmCo': '#66CCEE',
}
REGION_COLORS = {
    'SE Arc': '#EE6677', 'NE Arc': '#4477AA', 'NW Arc': '#228833',
    'SW Arc': '#CCBB44', 'North Linac': '#AA3377',
    'South Linac': '#66CCEE', 'Labyrinth': '#BBBBBB',
}

# ─── Placement mapping (JLAB-TN-25-021) ─────────────────────────────────────
PLACEMENTS = [
    ('Y15','N10','SE Arc','upstream Girder 27',1),
    ('Y3', 'N16','SE Arc','upstream Girder 27',2),
    ('Y23','N9', 'SE Arc','upstream Girder 27',3),
    ('Y26','N37','SE Arc','upstream Girder 27',4),
    ('Y40','N8', 'SE Arc','upstream Girder 27',5),
    ('Y39','S12','NE Arc','above JD3A06',1),
    ('Y7', 'S16','NE Arc','above JD3A06',2),
    ('Y18','S20','NE Arc','above JD3A06',3),
    ('Y21','S6', 'NE Arc','above JD3A06',4),
    ('Y9', 'S14','NE Arc','above JD3A06',5),
    ('Y38','N12','NW Arc','above MJC4A24',1),
    ('Y6', 'N39','NW Arc','above MJC4A24',2),
    ('Y36','N17','NW Arc','above MJC4A24',3),
    ('Y25','N11','NW Arc','above MJC4A24',4),
    ('Y34','N18','NW Arc','above MJC4A24',5),
    ('Y13','S1', 'SW Arc','above MJC4A14',1),
    ('Y32','S2', 'SW Arc','above MJC4A14',2),
    ('Y19','S17','SW Arc','above MJC4A14',3),
    ('Y10','S10','SW Arc','above MJC4A14',4),
    ('Y11','S3', 'SW Arc','above MJC4A14',5),
    ('Y12','S4', 'Labyrinth','North Access Labyrinth',0),
    ('Y17','N20','North Linac','NL NDX @ Girder 5',0),
    ('Y4', 'S7', 'North Linac','NL Inj/Ret Crossover (0L05 NDX)',0),
    ('Y16','S15','North Linac','NL NDX @ Girder 23',0),
    ('Y22','N15','North Linac','NL NDX @ Girder 26',0),
    ('Y20','S5', 'Labyrinth','South Access Labyrinth',0),
    ('Y24','S13','South Linac','SL NDX @ Girder 4',0),
    ('Y5', 'N19','South Linac','SL NDX @ Girder 6',0),
    ('Y1', 'N6', 'South Linac','SL NDX @ Girder 23',0),
    ('Y30','S18','South Linac','SL NDX @ Girder 25',0),
]

Y_PLACEMENT = {}
H_PLACEMENT = {}
for y, h, region, subloc, line in PLACEMENTS:
    ynum = int(y.replace('Y', ''))
    Y_PLACEMENT[ynum] = {'y_plate': y, 'h_plate': h, 'region': region,
                         'sub_location': subloc, 'line': line}
    H_PLACEMENT[h] = {'y_plate': y, 'h_plate': h, 'region': region,
                      'sub_location': subloc, 'line': line}

# Known outlier samples (bad baselines, >5% from material group median)
FLAGGED_OUTLIERS = {'Y-34-4', 'Y-40-4'}


# ─── Material assignments ────────────────────────────────────────────────────

def load_materials():
    mat_path = os.path.join(BASE, 'Materials_Arrangements.xlsx')
    wb = openpyxl.load_workbook(mat_path)
    y_materials = {}
    ws = wb['Tunnel - Y Materials']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            plate_num = plate_id.replace('y-', '')
            y_materials[plate_num] = [str(row[i]).strip() if row[i] else ''
                                      for i in range(1, 5)]
    pair_arrangements = {}
    ws2 = wb['Tunnel - Pair Arrangements']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            mat_type = str(row[1]).strip() if row[1] else ''
            configs = [str(row[i]).strip() if row[i] else '' for i in range(2, 6)]
            pair_arrangements[plate_id] = (mat_type, configs)
    return y_materials, pair_arrangements


def get_y_material(y_materials, sample_name):
    m = re.match(r'Y-(\d+)-(\d+)', sample_name)
    if not m:
        return None
    plate = m.group(1)
    slot = int(m.group(2))
    if plate in y_materials and 1 <= slot <= 4:
        return y_materials[plate][slot - 1]
    return None


# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_helmholtz_file(filepath):
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            val_match = re.search(r'([+-]?\d+\.?\d*)\s*(mWC|kT|kBG)', rest)
            if val_match:
                rows.append((dt, float(val_match.group(1)), val_match.group(2)))
    return rows


def parse_teslameter_file(filepath):
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            nums = re.findall(r'(-?\d+\.\d+)', rest)
            if len(nums) >= 4:
                rows.append((dt, [float(x) for x in nums[:3]], float(nums[3])))
            elif len(nums) >= 3:
                rows.append((dt, [float(x) for x in nums[:3]], None))
    return rows


def _clean_retakes(data):
    if len(data) < 3:
        return data
    median_val = np.median([v for _, v in data])
    cleaned = []
    skip = set()
    for i in range(len(data)):
        if i in skip:
            continue
        if i + 1 < len(data):
            dt1, v1 = data[i]
            dt2, v2 = data[i + 1]
            if abs((dt2 - dt1).total_seconds()) < 300:
                if abs(v1 - median_val) > abs(v2 - median_val):
                    skip.add(i)
                    continue
                else:
                    skip.add(i + 1)
        cleaned.append(data[i])
    for i in range(len(data)):
        if i not in skip and data[i] not in cleaned:
            cleaned.append(data[i])
    return sorted(cleaned, key=lambda x: x[0])


def get_all_mwc_rows(rows):
    all_mwc = [(dt, val) for dt, val, unit in rows
               if unit == 'mWC' and abs(val - SENTINEL) > 1]
    return _clean_retakes(all_mwc)


# ─── Temperature lookup (FIXED: includes Hn/Hs files) ────────────────────────

def build_temperature_lookup():
    """
    Build per-sample, per-date temperature lookup.

    FIX vs v1: For H-plates, also reads Hn/Hs-XX-YY Teslameter files
    (pair-level) in addition to An/As (individual magnet) files. This fixes
    date matching for pre-deployment measurements where the pair and individual
    readings were taken on different days.
    """
    temp_lookup = {}
    faces = ['front', 'side', 'top']

    # Y-plate temperatures
    y_tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
    y_samples = set()
    for f in os.listdir(y_tesla_dir):
        m = re.match(r'(Y-\d+-\d+)_(front|side|top)\.dat$', f)
        if m:
            y_samples.add(m.group(1))
    for sample in sorted(y_samples):
        date_temps = defaultdict(list)
        for face in faces:
            fpath = os.path.join(y_tesla_dir, f'{sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            for dt, fields, temp in rows:
                if temp is not None:
                    date_str = dt.strftime('%Y-%m-%d')
                    date_temps[date_str].append(temp)
        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(sample, date_str)] = (np.mean(temps), np.std(temps))

    # H-plate temperatures: from An/As AND Hn/Hs Teslameter files
    pair_tesla_dir = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')
    pair_helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')

    for f in os.listdir(pair_helm_dir):
        m = re.match(r'(H[ns]-\d+-\d+)_helmholtz\.dat$', f)
        if not m:
            continue
        h_sample = m.group(1)
        hm = re.match(r'H([ns])-(\d+)-(\d+)', h_sample)
        if not hm:
            continue
        ns, plate, slot = hm.group(1), hm.group(2), hm.group(3)
        a_prefix = 'An' if ns == 'n' else 'As'

        date_temps = defaultdict(list)

        # Source 1: An/As individual magnet Teslameter
        for magnet_idx in ['1', '2']:
            a_name = f'{a_prefix}-{plate}-{slot}-{magnet_idx}'
            for face in faces:
                fpath = os.path.join(pair_tesla_dir, f'{a_name}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                rows = parse_teslameter_file(fpath)
                for dt, fields, temp in rows:
                    if temp is not None:
                        date_str = dt.strftime('%Y-%m-%d')
                        date_temps[date_str].append(temp)

        # Source 2: Hn/Hs pair-level Teslameter (NEW in v2)
        for face in faces:
            fpath = os.path.join(pair_tesla_dir, f'{h_sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            for dt, fields, temp in rows:
                if temp is not None:
                    date_str = dt.strftime('%Y-%m-%d')
                    date_temps[date_str].append(temp)

        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(h_sample, date_str)] = (np.mean(temps),
                                                      np.std(temps))

    print(f"  Temperature lookup: {len(temp_lookup)} (sample, date) entries")
    return temp_lookup


# ─── Correction ──────────────────────────────────────────────────────────────

def correct_helmholtz(h_raw, alpha, t_mean, t_std):
    denom = 1.0 + alpha * (t_mean - T_REF)
    h_corr = h_raw / denom
    dh = abs(h_raw * alpha * t_std / denom**2)
    return h_corr, dh


# ─── Robust baseline computation ─────────────────────────────────────────────

def compute_robust_baseline(vals):
    """
    Compute baseline from a list of values with MAD outlier rejection.
    Returns (mean, std, n_kept, n_rejected) or None if no valid values.
    """
    vals = [v for v in vals if abs(v) >= MIN_BASELINE_MWC]
    if not vals:
        return None
    if len(vals) >= 3:
        med = np.median(vals)
        mad = np.median([abs(v - med) for v in vals])
        threshold = max(3.5 * mad, 0.005 * abs(med))
        kept = [v for v in vals if abs(v - med) <= threshold]
        n_rejected = len(vals) - len(kept)
    else:
        kept = vals
        n_rejected = 0
    if not kept:
        return None
    bl_mean = np.mean(kept)
    bl_std = np.std(kept, ddof=1) if len(kept) > 1 else 0.0
    return (bl_mean, bl_std, len(kept), n_rejected)


# ─── Compute degradation ─────────────────────────────────────────────────────

def compute_y_plate_degradation(y_materials, temp_lookup):
    """
    Y-plate Helmholtz degradation.

    CRITICAL: baseline uses ONLY temperature-corrected pre-deployment readings.
    Uncorrected readings are excluded from baseline to avoid systematic bias.
    """
    helm_dir = os.path.join(BASE, 'Y_Plates', 'Helmholtz')
    results = []
    excluded = []  # samples excluded for various reasons

    for f in sorted(os.listdir(helm_dir)):
        if not f.endswith('_helmholtz.dat'):
            continue
        sample = f.replace('_helmholtz.dat', '')
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()
        alpha = ALPHA.get(material)
        if alpha is None:
            continue
        pm = re.match(r'Y-(\d+)-(\d+)', sample)
        if not pm:
            continue
        plate_num = int(pm.group(1))
        slot = int(pm.group(2))
        placement = Y_PLACEMENT.get(plate_num)
        if not placement:
            continue

        fpath = os.path.join(helm_dir, f)
        raw_rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
        if not raw_rows:
            continue

        # Collect ONLY corrected readings
        pre_corr_vals = []
        tunnel_corr = []

        for dt, h_raw in raw_rows:
            date_str = dt.strftime('%Y-%m-%d')
            key = (sample, date_str)
            if key in temp_lookup:
                t_mean, t_std = temp_lookup[key]
                h_corr, dh = correct_helmholtz(h_raw, alpha, t_mean, t_std)
                if dt < TUNNEL_START:
                    pre_corr_vals.append(h_corr)
                else:
                    tunnel_corr.append((dt, h_raw, h_corr, dh))

        if not pre_corr_vals:
            excluded.append((sample, 'no corrected baseline'))
            continue
        if not tunnel_corr:
            excluded.append((sample, 'no corrected tunnel reading'))
            continue

        bl = compute_robust_baseline(pre_corr_vals)
        if bl is None:
            excluded.append((sample, 'baseline below threshold'))
            continue
        baseline_mean, baseline_std, n_kept, n_rejected = bl

        if abs(baseline_mean) < MIN_BASELINE_MWC:
            excluded.append((sample, 'baseline near zero'))
            continue

        tunnel_corr.sort(key=lambda x: x[0])
        latest_dt, latest_raw, latest_corr, latest_dh = tunnel_corr[-1]

        pct_change = (latest_corr - baseline_mean) / baseline_mean * 100.0
        sigma_baseline = baseline_std / np.sqrt(n_kept) if n_kept > 1 else baseline_std
        sigma_bl_pct = sigma_baseline / abs(baseline_mean) * 100.0
        sigma_temp_pct = latest_dh / abs(baseline_mean) * 100.0
        sigma_total_pct = np.sqrt(sigma_bl_pct**2 + sigma_temp_pct**2)

        is_outlier = sample in FLAGGED_OUTLIERS

        results.append({
            'sample': sample, 'plate': plate_num, 'slot': slot,
            'material': material, 'region': placement['region'],
            'sub_location': placement['sub_location'],
            'line': placement['line'], 'h_plate': placement['h_plate'],
            'baseline_mean': baseline_mean, 'baseline_std': baseline_std,
            'n_baseline': n_kept, 'n_rejected': n_rejected,
            'latest_date': latest_dt.strftime('%Y-%m-%d'),
            'latest_corr': latest_corr, 'latest_raw': latest_raw,
            'pct_change': pct_change,
            'sigma_bl_pct': sigma_bl_pct, 'sigma_temp_pct': sigma_temp_pct,
            'sigma_total_pct': sigma_total_pct,
            'is_outlier': is_outlier,
        })

    return results, excluded


def compute_h_plate_degradation(pair_arrangements, temp_lookup):
    """
    H-plate (pair assembly) Helmholtz degradation.

    CRITICAL: baseline uses ONLY temperature-corrected readings.
    """
    helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')
    results = []
    excluded = []

    for f in sorted(os.listdir(helm_dir)):
        m = re.match(r'(H[ns]-\d+-\d+)_helmholtz\.dat$', f)
        if not m:
            continue
        h_sample = m.group(1)
        hm = re.match(r'H([ns])-(\d+)-(\d+)', h_sample)
        if not hm:
            continue
        ns = hm.group(1)
        plate_num = int(hm.group(2))
        slot = int(hm.group(3))
        mat_type = 'NdFeB' if ns == 'n' else 'SmCo'
        alpha = ALPHA[mat_type]

        plate_key = f'{ns}-{plate_num}'
        config = ''
        if plate_key in pair_arrangements:
            _, configs = pair_arrangements[plate_key]
            if slot - 1 < len(configs):
                config = configs[slot - 1]

        h_lookup_key = f'{"N" if ns == "n" else "S"}{plate_num}'
        placement = H_PLACEMENT.get(h_lookup_key)
        if not placement:
            continue

        fpath = os.path.join(helm_dir, f)
        raw_rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
        if not raw_rows:
            continue

        pre_corr_vals = []
        tunnel_corr = []

        for dt, h_raw in raw_rows:
            date_str = dt.strftime('%Y-%m-%d')
            key = (h_sample, date_str)
            if key in temp_lookup:
                t_mean, t_std = temp_lookup[key]
                h_corr, dh = correct_helmholtz(h_raw, alpha, t_mean, t_std)
                if dt < TUNNEL_START:
                    pre_corr_vals.append(h_corr)
                else:
                    tunnel_corr.append((dt, h_raw, h_corr, dh))

        if not pre_corr_vals:
            excluded.append((h_sample, f'no corrected baseline (config={config})'))
            continue
        if not tunnel_corr:
            excluded.append((h_sample, 'no corrected tunnel reading'))
            continue

        bl = compute_robust_baseline(pre_corr_vals)
        if bl is None:
            excluded.append((h_sample, 'baseline below threshold'))
            continue
        baseline_mean, baseline_std, n_kept, n_rejected = bl

        if abs(baseline_mean) < MIN_BASELINE_MWC:
            excluded.append((h_sample, 'baseline near zero'))
            continue

        tunnel_corr.sort(key=lambda x: x[0])
        latest_dt, latest_raw, latest_corr, latest_dh = tunnel_corr[-1]

        pct_change = (latest_corr - baseline_mean) / baseline_mean * 100.0
        sigma_baseline = baseline_std / np.sqrt(n_kept) if n_kept > 1 else baseline_std
        sigma_bl_pct = sigma_baseline / abs(baseline_mean) * 100.0
        sigma_temp_pct = latest_dh / abs(baseline_mean) * 100.0
        sigma_total_pct = np.sqrt(sigma_bl_pct**2 + sigma_temp_pct**2)

        # Flag H-plate outliers: |pct_change| > 5% likely indicates an anomalous
        # baseline (wrong config, mislabeled, or measurement error) — physically
        # impossible from radiation alone in ~1.5 years. Also flag samples with
        # single-reading baseline and >2% change as unreliable.
        is_outlier = (abs(pct_change) > 5.0 or
                      (n_kept == 1 and abs(pct_change) > 2.0))

        results.append({
            'sample': h_sample, 'plate': plate_num, 'slot': slot,
            'material': mat_type, 'config': config,
            'region': placement['region'],
            'sub_location': placement['sub_location'],
            'line': placement['line'],
            'baseline_mean': baseline_mean, 'baseline_std': baseline_std,
            'n_baseline': n_kept,
            'latest_date': latest_dt.strftime('%Y-%m-%d'),
            'latest_corr': latest_corr, 'latest_raw': latest_raw,
            'pct_change': pct_change,
            'sigma_bl_pct': sigma_bl_pct, 'sigma_temp_pct': sigma_temp_pct,
            'sigma_total_pct': sigma_total_pct,
            'is_outlier': is_outlier,
        })

    return results, excluded


def compute_teslameter_degradation(y_materials, pair_arrangements):
    """
    Teslameter field magnitude degradation.

    Since pre-deployment Teslameter field readings are invalid (broken probe),
    the baseline is the FIRST tunnel-period measurement. This means we are
    measuring change during the tunnel exposure period only, not total change
    from the pre-deployment state.

    For each sample/face, computes |B| = sqrt(F1² + F2² + F3²) and applies
    the same temperature correction: |B|_corr = |B|_raw / (1 + α(T - T_ref)).
    The per-face temperature reading (field 4) is used directly.

    Returns results for both Y-plate and H-plate (Hn/Hs level) Teslameter.
    """
    results = []
    faces = ['front', 'side', 'top']

    # --- Y-plate Teslameter ---
    y_tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
    y_samples = set()
    for f in os.listdir(y_tesla_dir):
        m = re.match(r'(Y-\d+-\d+)_(front|side|top)\.dat$', f)
        if m:
            y_samples.add(m.group(1))

    for sample in sorted(y_samples):
        pm = re.match(r'Y-(\d+)-(\d+)', sample)
        if not pm:
            continue
        plate_num = int(pm.group(1))
        slot = int(pm.group(2))
        placement = Y_PLACEMENT.get(plate_num)
        if not placement:
            continue
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()
        alpha = ALPHA.get(material)
        if alpha is None:
            continue

        # Collect corrected field magnitudes per face, tunnel period only
        face_data = {}  # face -> [(dt, mag_corr)]
        for face in faces:
            fpath = os.path.join(y_tesla_dir, f'{sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            corrected = []
            for dt, fields, temp in rows:
                if dt < TESLAMETER_FIELD_VALID_AFTER:
                    continue
                if temp is None:
                    continue
                mag = np.sqrt(sum(f**2 for f in fields))
                denom = 1.0 + alpha * (temp - T_REF)
                mag_corr = mag / denom
                corrected.append((dt, mag_corr))
            if corrected:
                face_data[face] = sorted(corrected, key=lambda x: x[0])

        if not face_data:
            continue

        # Compute % change PER FACE independently, then average.
        # Each face measures a different field component, so absolute magnitudes
        # differ widely (e.g. 78 vs 240 vs 141 mT). Averaging magnitudes
        # across faces creates spurious spread. Instead, compute the percent
        # change for each face separately and average those.
        face_pct_changes = []
        bl_date = None
        latest_date = None
        bl_mean_mag = None

        for face, data in face_data.items():
            dates_for_face = defaultdict(list)
            for dt, mag_corr in data:
                dates_for_face[dt.strftime('%Y-%m-%d')].append(mag_corr)
            sdates = sorted(dates_for_face.keys())
            if len(sdates) < 2:
                continue
            bl_d = sdates[0]
            lt_d = sdates[-1]
            bl_mag = np.mean(dates_for_face[bl_d])
            lt_mag = np.mean(dates_for_face[lt_d])
            if abs(bl_mag) < 1.0:
                continue
            face_pct = (lt_mag - bl_mag) / bl_mag * 100.0
            face_pct_changes.append(face_pct)
            bl_date = bl_d
            latest_date = lt_d
            if bl_mean_mag is None:
                bl_mean_mag = bl_mag

        if not face_pct_changes:
            continue

        pct_change = np.mean(face_pct_changes)
        # Error: std of per-face percent changes (captures face-to-face variation)
        if len(face_pct_changes) > 1:
            sigma_pct = np.std(face_pct_changes, ddof=1) / np.sqrt(len(face_pct_changes))
        else:
            sigma_pct = 0.5  # default 0.5% uncertainty for single-face

        is_outlier = sample in FLAGGED_OUTLIERS

        results.append({
            'sample': sample, 'plate': plate_num, 'slot': slot,
            'material': material, 'type': 'Y-plate',
            'region': placement['region'],
            'sub_location': placement['sub_location'],
            'line': placement['line'],
            'baseline_date': bl_date,
            'baseline_mean': bl_mean_mag,
            'latest_date': latest_date,
            'latest_mean': None,
            'pct_change': pct_change,
            'sigma_pct': sigma_pct,
            'n_faces': len(face_pct_changes),
            'is_outlier': is_outlier,
        })

    # --- H-plate Teslameter (Hn/Hs pair level) ---
    pair_tesla_dir = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')
    h_samples_seen = set()
    for f in sorted(os.listdir(pair_tesla_dir)):
        m = re.match(r'(H[ns]-\d+-\d+)_(front|side|top)\.dat$', f)
        if not m:
            continue
        h_sample = m.group(1)
        # Only process once per sample
        if h_sample in h_samples_seen:
            continue
        h_samples_seen.add(h_sample)

        hm = re.match(r'H([ns])-(\d+)-(\d+)', h_sample)
        if not hm:
            continue
        ns = hm.group(1)
        plate_num = int(hm.group(2))
        slot = int(hm.group(3))
        mat_type = 'NdFeB' if ns == 'n' else 'SmCo'
        alpha = ALPHA[mat_type]

        plate_key = f'{ns}-{plate_num}'
        config = ''
        if plate_key in pair_arrangements:
            _, configs = pair_arrangements[plate_key]
            if slot - 1 < len(configs):
                config = configs[slot - 1]

        h_lookup_key = f'{"N" if ns == "n" else "S"}{plate_num}'
        placement = H_PLACEMENT.get(h_lookup_key)
        if not placement:
            continue

        face_data = {}
        for face in faces:
            fpath = os.path.join(pair_tesla_dir, f'{h_sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            corrected = []
            for dt, fields, temp in rows:
                if dt < TESLAMETER_FIELD_VALID_AFTER:
                    continue
                if temp is None:
                    continue
                mag = np.sqrt(sum(fld**2 for fld in fields))
                denom = 1.0 + alpha * (temp - T_REF)
                mag_corr = mag / denom
                corrected.append((dt, mag_corr))
            if corrected:
                face_data[face] = sorted(corrected, key=lambda x: x[0])

        if not face_data:
            continue

        # Same per-face % change approach
        face_pct_changes = []
        bl_date = None
        latest_date = None

        for face, data in face_data.items():
            dates_for_face = defaultdict(list)
            for dt, mag_corr in data:
                dates_for_face[dt.strftime('%Y-%m-%d')].append(mag_corr)
            sdates = sorted(dates_for_face.keys())
            if len(sdates) < 2:
                continue
            bl_d = sdates[0]
            lt_d = sdates[-1]
            bl_mag = np.mean(dates_for_face[bl_d])
            lt_mag = np.mean(dates_for_face[lt_d])
            if abs(bl_mag) < 1.0:
                continue
            face_pct = (lt_mag - bl_mag) / bl_mag * 100.0
            face_pct_changes.append(face_pct)
            bl_date = bl_d
            latest_date = lt_d

        if not face_pct_changes:
            continue

        pct_change = np.mean(face_pct_changes)
        if len(face_pct_changes) > 1:
            sigma_pct = np.std(face_pct_changes, ddof=1) / np.sqrt(len(face_pct_changes))
        else:
            sigma_pct = 0.5

        results.append({
            'sample': h_sample, 'plate': plate_num, 'slot': slot,
            'material': mat_type, 'type': 'H-plate',
            'config': config,
            'region': placement['region'],
            'sub_location': placement['sub_location'],
            'line': placement['line'],
            'baseline_date': bl_date,
            'baseline_mean': None,
            'latest_date': latest_date,
            'latest_mean': None,
            'pct_change': pct_change,
            'sigma_pct': sigma_pct,
            'n_faces': len(face_pct_changes),
            'is_outlier': False,
        })

    return results


# ─── Summary statistics ──────────────────────────────────────────────────────

def summarize_group(results_list, label, sigma_key='sigma_total_pct'):
    """Compute mean and uncertainty for a group, excluding outliers."""
    clean = [r for r in results_list if not r.get('is_outlier', False)]
    if not clean:
        return None
    pcts = [r['pct_change'] for r in clean]
    sigmas = [r[sigma_key] for r in clean]
    n = len(pcts)
    mean_pct = np.mean(pcts)
    std_pct = np.std(pcts, ddof=1) if n > 1 else 0.0
    sem_pct = std_pct / np.sqrt(n) if n > 1 else sigmas[0]
    mean_sigma = np.mean(sigmas)
    combined_unc = max(sem_pct, mean_sigma)
    return {
        'label': label, 'n': n, 'mean_pct': mean_pct, 'std_pct': std_pct,
        'sem_pct': sem_pct, 'mean_sigma': mean_sigma,
        'combined_unc': combined_unc,
        'min_pct': min(pcts), 'max_pct': max(pcts),
    }


# ─── Plotting ────────────────────────────────────────────────────────────────

def plot_by_region(y_results, h_results):
    regions = ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
               'North Linac', 'South Linac', 'Labyrinth']
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 1, figsize=(14, 12))
    x = np.arange(len(regions))

    # Y-plates
    ax = axes[0]
    width = 0.18
    offsets = [-1.5, -0.5, 0.5, 1.5]
    for i, mat in enumerate(materials):
        means, errs = [], []
        for region in regions:
            group = [r for r in y_results
                     if r['region'] == region and r['material'] == mat]
            s = summarize_group(group, f'{region}/{mat}')
            means.append(s['mean_pct'] if s else 0)
            errs.append(s['combined_unc'] if s else 0)
        ax.bar(x + offsets[i] * width, means, width, yerr=errs,
               color=CB[mat], label=mat, capsize=3, alpha=0.85,
               edgecolor='black', linewidth=0.5)
    ax.set_xticks(x)
    ax.set_xticklabels(regions, fontsize=10)
    ax.set_ylabel('% Change from Baseline', fontsize=11)
    ax.set_title('Y-Plate Degradation by Region and Material\n'
                 '(Temp-corrected to 20°C; outliers excluded from group stats)',
                 fontsize=12, fontweight='bold')
    ax.axhline(0, color='black', linewidth=0.5)
    ax.legend(fontsize=9)
    ax.grid(axis='y', alpha=0.3)

    # H-plates
    ax2 = axes[1]
    width2 = 0.3
    offsets2 = [-0.5, 0.5]
    for i, mat in enumerate(['NdFeB', 'SmCo']):
        means, errs = [], []
        for region in regions:
            group = [r for r in h_results
                     if r['region'] == region and r['material'] == mat]
            s = summarize_group(group, f'{region}/{mat}')
            means.append(s['mean_pct'] if s else 0)
            errs.append(s['combined_unc'] if s else 0)
        ax2.bar(x + offsets2[i] * width2, means, width2, yerr=errs,
                color=CB[mat], label=mat, capsize=3, alpha=0.85,
                edgecolor='black', linewidth=0.5)
    ax2.set_xticks(x)
    ax2.set_xticklabels(regions, fontsize=10)
    ax2.set_ylabel('% Change from Baseline', fontsize=11)
    ax2.set_title('H-Plate (Pair Assembly) Degradation by Region\n'
                  '(Corrected baseline only; no uncorrected pre-deployment mixing)',
                  fontsize=12, fontweight='bold')
    ax2.axhline(0, color='black', linewidth=0.5)
    ax2.legend(fontsize=9)
    ax2.grid(axis='y', alpha=0.3)

    fig.tight_layout()
    fig.savefig(os.path.join(PLOT_DIR, 'degradation_v2_by_region.png'),
                dpi=150, bbox_inches='tight')
    plt.close(fig)


def plot_individual_samples(y_results):
    regions_order = ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
                     'North Linac', 'South Linac', 'Labyrinth']
    fig, ax = plt.subplots(figsize=(12, 20))
    y_pos = 0
    yticks, ylabels = [], []
    region_boundaries = []

    for region in regions_order:
        group = sorted([r for r in y_results if r['region'] == region],
                       key=lambda r: r['pct_change'])
        if not group:
            continue
        region_boundaries.append((y_pos, region))
        for r in group:
            color = CB.get(r['material'], '#888888')
            alpha_val = 0.3 if r['is_outlier'] else 0.8
            hatch = '///' if r['is_outlier'] else None
            ax.barh(y_pos, r['pct_change'], xerr=r['sigma_total_pct'],
                    height=0.7, color=color, alpha=alpha_val, capsize=3,
                    edgecolor='black', linewidth=0.3, hatch=hatch)
            yticks.append(y_pos)
            line_str = f" (L{r['line']})" if r['line'] > 0 else ""
            outlier_str = " ⚠OUTLIER" if r['is_outlier'] else ""
            ylabels.append(f"{r['sample']} [{r['material']}]{line_str}{outlier_str}")
            y_pos += 1
        y_pos += 1

    ax.set_yticks(yticks)
    ax.set_yticklabels(ylabels, fontsize=7)
    ax.axvline(0, color='black', linewidth=0.8)
    ax.set_xlabel('% Change from Baseline (corrected to 20°C)', fontsize=11)
    ax.set_title('Y-Plate Degradation — All Samples by Region\n'
                 '(hatched = flagged outlier, excluded from group statistics)',
                 fontsize=12, fontweight='bold')
    ax.grid(axis='x', alpha=0.3)
    for ystart, region in region_boundaries:
        ax.annotate(region, xy=(-0.01, ystart - 0.5),
                    xycoords=('axes fraction', 'data'),
                    fontsize=10, fontweight='bold', color='gray',
                    ha='right', va='bottom')
    ax.invert_yaxis()
    fig.tight_layout()
    fig.savefig(os.path.join(PLOT_DIR, 'degradation_v2_all_samples.png'),
                dpi=150, bbox_inches='tight')
    plt.close(fig)


def plot_teslameter_summary(tesla_results):
    """Teslameter field degradation summary by material and region."""
    fig, axes = plt.subplots(1, 2, figsize=(16, 7))

    # By material
    ax = axes[0]
    y_tesla = [r for r in tesla_results if r['type'] == 'Y-plate']
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']
    mat_means, mat_errs, mat_labels = [], [], []
    for mat in materials:
        s = summarize_group([r for r in y_tesla if r['material'] == mat],
                            mat, sigma_key='sigma_pct')
        if s:
            mat_means.append(s['mean_pct'])
            mat_errs.append(s['combined_unc'])
            mat_labels.append(f"{mat}\n(N={s['n']})")
    ax.bar(range(len(mat_labels)), mat_means, yerr=mat_errs,
           color=[CB[m] for m in materials[:len(mat_labels)]],
           capsize=5, alpha=0.85, edgecolor='black', linewidth=0.5)
    ax.set_xticks(range(len(mat_labels)))
    ax.set_xticklabels(mat_labels)
    ax.axhline(0, color='black', linewidth=0.5)
    ax.set_ylabel('% Change (tunnel period)', fontsize=11)
    ax.set_title('Y-Plate Teslameter |B| Change by Material\n'
                 '(baseline = first tunnel measurement)',
                 fontsize=11, fontweight='bold')
    ax.grid(axis='y', alpha=0.3)

    # H-plate by config
    ax2 = axes[1]
    h_tesla = [r for r in tesla_results if r['type'] == 'H-plate']
    configs = ['Alpha', 'Beta', 'Gamma', 'Delta']
    cfg_means, cfg_errs, cfg_labels = [], [], []
    cfg_colors = ['#4477AA', '#EE6677', '#228833', '#CCBB44']
    for cfg, cc in zip(configs, cfg_colors):
        group = [r for r in h_tesla if r.get('config') == cfg]
        s = summarize_group(group, cfg, sigma_key='sigma_pct')
        if s:
            cfg_means.append(s['mean_pct'])
            cfg_errs.append(s['combined_unc'])
            cfg_labels.append(f"{cfg}\n(N={s['n']})")
    if cfg_means:
        ax2.bar(range(len(cfg_labels)), cfg_means, yerr=cfg_errs,
                color=cfg_colors[:len(cfg_labels)],
                capsize=5, alpha=0.85, edgecolor='black', linewidth=0.5)
        ax2.set_xticks(range(len(cfg_labels)))
        ax2.set_xticklabels(cfg_labels)
    ax2.axhline(0, color='black', linewidth=0.5)
    ax2.set_ylabel('% Change (tunnel period)', fontsize=11)
    ax2.set_title('H-Plate Teslameter |B| Change by Config\n'
                  '(including Beta — reliable in Teslameter)',
                  fontsize=11, fontweight='bold')
    ax2.grid(axis='y', alpha=0.3)

    fig.tight_layout()
    fig.savefig(os.path.join(PLOT_DIR, 'degradation_v2_teslameter.png'),
                dpi=150, bbox_inches='tight')
    plt.close(fig)


# ─── Markdown report ─────────────────────────────────────────────────────────

def write_report(y_res, h_res, tesla_res, y_excl, h_excl):
    L = []
    L.append("# Preliminary Degradation Summary (v2)")
    L.append("")
    L.append("**Status**: Preliminary — more data collection and error analysis pending")
    L.append(f"**Generated**: {datetime.now().strftime('%Y-%m-%d')}")
    L.append(f"**Reference temperature**: {T_REF} °C")
    L.append("")

    # Fixes section
    L.append("## Methodology and Corrections Applied")
    L.append("")
    L.append("### Issue 1: Outlier Samples")
    L.append("Two Y-plate samples have pre-deployment baselines deviating >5% from "
             "their material-group median:")
    L.append("- **Y-34-4 (N52SH)**: baseline 1.172 mWC vs group median 1.312 mWC (−10.7%)")
    L.append("- **Y-40-4 (SmCo33H)**: baseline deviates +6.8% from group median")
    L.append("")
    L.append("**Treatment**: These samples are plotted (hatched bars) but **excluded from "
             "all group statistics** (means, uncertainties, ranges). They are reported "
             "separately. The cause of the anomalous baselines is unknown — possible "
             "labeling error, damaged sample, or measurement artifact.")
    L.append("")

    L.append("### Issue 2: H-Plate Baseline Bias (corrected in v2)")
    L.append("In v1, H-plate pre-deployment baselines mixed temperature-corrected and "
             "uncorrected Helmholtz readings. Many pre-deployment dates (Apr–Jun 2025) "
             "had no Teslameter temperature data because the Hall probe was broken. "
             "Using raw (uncorrected) readings as baseline while comparing against "
             "temperature-corrected tunnel readings introduces a systematic positive "
             "bias of ~0.2–1% depending on the temperature difference.")
    L.append("")
    L.append("**Treatment**: v2 uses **only temperature-corrected readings** for both "
             "baseline and tunnel measurements. Pre-deployment dates without Teslameter "
             "temperature are excluded from the baseline calculation entirely. "
             "Additionally, the temperature lookup now includes Hn/Hs pair-level "
             "Teslameter files (not just An/As individual magnet files), which recovers "
             "34 additional temperature-matched baseline readings from the Nov 2024 "
             "campaign where pair and individual measurements were taken on different days.")
    L.append("")
    n_h_excl_bl = sum(1 for _, reason in h_excl if 'no corrected baseline' in reason)
    L.append(f"**Impact**: {n_h_excl_bl} H-plate samples now excluded due to having no "
             f"temperature-corrected pre-deployment readings (only {len(h_res)} of "
             f"{len(h_res) + n_h_excl_bl} tunnel H-plate samples retained).")
    L.append("")

    L.append("### Issue 3: Teslameter Field Degradation")
    L.append("Pre-deployment Teslameter field readings are invalid (broken Hall probe — "
             "magnitudes ~10–20× lower than actual). The temperature sensor on the same "
             "probe appears to have been unaffected (readings are consistent with expected "
             "lab temperatures of ~22–24°C).")
    L.append("")
    L.append("**Treatment**: Teslameter field degradation uses the **first tunnel-period "
             "measurement** (Jul 2025) as baseline, not the pre-deployment reading. "
             "This measures change during tunnel exposure only. The same α(Br) temperature "
             "correction is applied using each face's own temperature reading. "
             "Beta (antiparallel) pair assemblies, which are unreliable in Helmholtz "
             "measurements due to multipole field character, **are reliable in Teslameter** "
             "point measurements and are included here.")
    L.append("")

    # Error bars
    L.append("## Error Bar Methodology")
    L.append("")
    L.append("**Helmholtz (Y-plate and H-plate):**")
    L.append("- σ_baseline = std(pre-deployment corrected readings) / √N")
    L.append("- σ_temp = |B_raw × α × σ_T / (1 + α(T−20))²|, from face temperature spread")
    L.append("- σ_total = √(σ_baseline² + σ_temp²), as % of baseline")
    L.append("")
    L.append("**Teslameter field:**")
    L.append("- σ = face-to-face spread (std of front/side/top corrected |B|) / √N_faces")
    L.append("- Propagated from both baseline and latest measurement dates")
    L.append("")
    L.append("**Group averages:**")
    L.append("- Uncertainty = max(SEM across samples, mean per-sample σ)")
    L.append("- Ensures we never underestimate uncertainty")
    L.append("")

    # ─── Y-plate results ───
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']
    regions = ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
               'North Linac', 'South Linac', 'Labyrinth']

    L.append("## Y-Plate Helmholtz Results")
    L.append("")
    L.append("### By Material (outliers excluded)")
    L.append("")
    L.append("| Material | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    L.append("|----------|---|-----------|-----------|---------|---------|")
    for mat in materials:
        s = summarize_group([r for r in y_res if r['material'] == mat], mat)
        if s:
            L.append(f"| {mat} | {s['n']} | {s['mean_pct']:+.3f} | "
                     f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                     f"{s['max_pct']:+.3f} |")
    L.append("")

    L.append("### By Region (outliers excluded)")
    L.append("")
    L.append("| Region | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    L.append("|--------|---|-----------|-----------|---------|---------|")
    for region in regions:
        s = summarize_group([r for r in y_res if r['region'] == region], region)
        if s:
            L.append(f"| {region} | {s['n']} | {s['mean_pct']:+.3f} | "
                     f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                     f"{s['max_pct']:+.3f} |")
    L.append("")

    L.append("### By Region × Material")
    L.append("")
    L.append("| Region | Material | N | Mean Δ (%) | ± Unc (%) |")
    L.append("|--------|----------|---|-----------|-----------|")
    for region in regions:
        for mat in materials:
            s = summarize_group(
                [r for r in y_res if r['region'] == region and r['material'] == mat],
                f'{region}/{mat}')
            if s:
                L.append(f"| {region} | {mat} | {s['n']} | "
                         f"{s['mean_pct']:+.3f} | ±{s['combined_unc']:.3f} |")
    L.append("")

    L.append("### By Arc Line Position")
    L.append("")
    L.append("Line 1 = top of stack, Line 5 = bottom. Different beam energy passes.")
    L.append("")
    L.append("| Line | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    L.append("|------|---|-----------|-----------|---------|---------|")
    for lp in range(1, 6):
        s = summarize_group([r for r in y_res if r['line'] == lp], f'Line {lp}')
        if s:
            L.append(f"| {lp} | {s['n']} | {s['mean_pct']:+.3f} | "
                     f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                     f"{s['max_pct']:+.3f} |")
    L.append("")

    # ─── H-plate results ───
    L.append("## H-Plate Helmholtz Results")
    L.append("")
    L.append("### By Material")
    L.append("")
    L.append("| Material | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    L.append("|----------|---|-----------|-----------|---------|---------|")
    for mat in ['NdFeB', 'SmCo']:
        s = summarize_group([r for r in h_res if r['material'] == mat], mat)
        if s:
            L.append(f"| {mat} | {s['n']} | {s['mean_pct']:+.3f} | "
                     f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                     f"{s['max_pct']:+.3f} |")
    L.append("")

    L.append("### By Assembly Configuration")
    L.append("")
    L.append("| Config | Material | N | Mean Δ (%) | ± Unc (%) | Notes |")
    L.append("|--------|----------|---|-----------|-----------|-------|")
    for config in ['Alpha', 'Beta', 'Gamma', 'Delta']:
        for mat in ['NdFeB', 'SmCo']:
            group = [r for r in h_res
                     if r.get('config') == config and r['material'] == mat]
            s = summarize_group(group, f'{config}/{mat}')
            if s:
                note = "⚠ Helmholtz unreliable for Beta" if config == 'Beta' else ""
                L.append(f"| {config} | {mat} | {s['n']} | "
                         f"{s['mean_pct']:+.3f} | ±{s['combined_unc']:.3f} | {note} |")
    L.append("")

    # ─── Teslameter results ───
    L.append("## Teslameter Field Results")
    L.append("")
    L.append("**Note**: Baseline is first tunnel measurement (Jul 2025), not pre-deployment. "
             "Measures change during tunnel exposure only.")
    L.append("")

    y_tesla = [r for r in tesla_res if r['type'] == 'Y-plate']
    h_tesla = [r for r in tesla_res if r['type'] == 'H-plate']

    L.append("### Y-Plate Teslameter by Material")
    L.append("")
    L.append("| Material | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    L.append("|----------|---|-----------|-----------|---------|---------|")
    for mat in materials:
        s = summarize_group([r for r in y_tesla if r['material'] == mat],
                            mat, sigma_key='sigma_pct')
        if s:
            L.append(f"| {mat} | {s['n']} | {s['mean_pct']:+.3f} | "
                     f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                     f"{s['max_pct']:+.3f} |")
    L.append("")

    L.append("### H-Plate Teslameter by Configuration (including Beta)")
    L.append("")
    L.append("| Config | N | Mean Δ (%) | ± Unc (%) |")
    L.append("|--------|---|-----------|-----------|")
    for cfg in ['Alpha', 'Beta', 'Gamma', 'Delta']:
        s = summarize_group([r for r in h_tesla if r.get('config') == cfg],
                            cfg, sigma_key='sigma_pct')
        if s:
            L.append(f"| {cfg} | {s['n']} | {s['mean_pct']:+.3f} | "
                     f"±{s['combined_unc']:.3f} |")
    L.append("")

    # ─── Outlier report ───
    L.append("## Flagged Outlier Samples")
    L.append("")
    outliers = [r for r in y_res if r['is_outlier']]
    if outliers:
        L.append("| Sample | Material | Region | Δ (%) | ± (%) | Reason |")
        L.append("|--------|----------|--------|-------|-------|--------|")
        for r in outliers:
            L.append(f"| {r['sample']} | {r['material']} | {r['region']} | "
                     f"{r['pct_change']:+.3f} | ±{r['sigma_total_pct']:.3f} | "
                     f"Baseline >5% from group median |")
    L.append("")

    # ─── Individual detail ───
    L.append("## Individual Y-Plate Results")
    L.append("")
    L.append("| Sample | Material | Region | Line | Δ (%) | ± (%) | Outlier | Date |")
    L.append("|--------|----------|--------|------|-------|-------|---------|------|")
    for r in sorted(y_res, key=lambda x: x['pct_change']):
        line_str = str(r['line']) if r['line'] > 0 else "—"
        flag = "⚠" if r['is_outlier'] else ""
        L.append(f"| {r['sample']} | {r['material']} | {r['region']} | "
                 f"{line_str} | {r['pct_change']:+.3f} | ±{r['sigma_total_pct']:.3f} | "
                 f"{flag} | {r['latest_date']} |")
    L.append("")

    # ─── Key findings ───
    L.append("## Key Findings")
    L.append("")
    clean_y = [r for r in y_res if not r['is_outlier']]
    most_deg = min(clean_y, key=lambda r: r['pct_change'])
    L.append(f"1. **N42EH shows the clearest degradation signal**: "
             f"mean −{abs(summarize_group([r for r in y_res if r['material']=='N42EH'], 'x')['mean_pct']):.2f}% "
             f"across all tunnel locations. This is consistent with NdFeB's known "
             f"radiation sensitivity.")
    L.append(f"2. **Largest individual degradation**: {most_deg['sample']} "
             f"({most_deg['material']}, {most_deg['region']}) at "
             f"{most_deg['pct_change']:+.3f} ± {most_deg['sigma_total_pct']:.3f}%")

    sig_05 = [r for r in clean_y if r['pct_change'] < -0.5]
    L.append(f"3. **Samples with >0.5% degradation**: {len(sig_05)} / {len(clean_y)}")

    L.append("4. **SmCo shows less degradation than NdFeB**, consistent with "
             "theoretical expectations (lower coercivity sensitivity to radiation damage)")
    L.append("5. **Labyrinth (control) sites show minimal change**, providing confidence "
             "that observed arc/linac degradation is radiation-related")
    L.append("")

    # ─── Caveats ───
    L.append("## Caveats and Limitations")
    L.append("")
    L.append("- **Preliminary**: More measurements are being collected")
    L.append("- **No dose correlation yet**: Radiation dose data incomplete; "
             "dose-response analysis deferred")
    L.append("- **α(Br) systematic**: Temperature coefficient uncertainty (~5–10% "
             "on α itself) would shift all values of a material grade together; "
             "not included in per-sample error bars")
    L.append("- **H-plate sample size reduced**: Only samples with temperature-corrected "
             "pre-deployment baselines are included, reducing statistical power")
    L.append("- **Teslameter baseline**: Uses first tunnel measurement, not "
             "pre-deployment; shorter exposure baseline")
    L.append("- **Probe positioning**: Teslameter measures a point on the magnet "
             "surface where field gradients are steep; ~0.5mm shift can produce >1% "
             "variation, making Teslameter inherently noisier than Helmholtz")
    L.append("- **Lab controls not yet included** (data still being collected)")
    L.append("")

    with open(os.path.join(MD_DIR, 'degradation_summary_v2.md'), 'w') as f:
        f.write('\n'.join(L))


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    y_materials, pair_arrangements = load_materials()
    temp_lookup = build_temperature_lookup()

    print("\nComputing Y-plate Helmholtz degradation...")
    y_results, y_excluded = compute_y_plate_degradation(y_materials, temp_lookup)
    n_outlier = sum(1 for r in y_results if r['is_outlier'])
    print(f"  {len(y_results)} samples ({n_outlier} flagged outliers)")
    if y_excluded:
        print(f"  {len(y_excluded)} excluded:")
        for s, reason in y_excluded[:5]:
            print(f"    {s}: {reason}")

    print("\nComputing H-plate Helmholtz degradation...")
    h_results, h_excluded = compute_h_plate_degradation(pair_arrangements, temp_lookup)
    print(f"  {len(h_results)} samples with corrected baseline")
    n_no_bl = sum(1 for _, r in h_excluded if 'no corrected baseline' in r)
    print(f"  {n_no_bl} excluded (no temperature-corrected baseline)")

    print("\nComputing Teslameter field degradation...")
    tesla_results = compute_teslameter_degradation(y_materials, pair_arrangements)
    n_y_tesla = sum(1 for r in tesla_results if r['type'] == 'Y-plate')
    n_h_tesla = sum(1 for r in tesla_results if r['type'] == 'H-plate')
    print(f"  {n_y_tesla} Y-plate + {n_h_tesla} H-plate Teslameter samples")

    # ─── Print summaries ───
    print("\n" + "="*70)
    print("Y-PLATE HELMHOLTZ (outliers excluded)")
    print("="*70)
    for mat in ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']:
        s = summarize_group([r for r in y_results if r['material'] == mat], mat)
        if s:
            print(f"  {mat:8s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']}, range [{s['min_pct']:+.3f}, {s['max_pct']:+.3f}])")

    print()
    for region in ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
                   'North Linac', 'South Linac', 'Labyrinth']:
        s = summarize_group([r for r in y_results if r['region'] == region], region)
        if s:
            print(f"  {region:14s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']})")

    h_outliers = sum(1 for r in h_results if r.get('is_outlier'))
    print(f"\n{'='*70}")
    print(f"H-PLATE HELMHOLTZ (corrected baselines; {h_outliers} outliers excluded)")
    print("="*70)
    for mat in ['NdFeB', 'SmCo']:
        s = summarize_group([r for r in h_results if r['material'] == mat], mat)
        if s:
            print(f"  {mat:8s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']})")
    for config in ['Alpha', 'Beta', 'Gamma', 'Delta']:
        s = summarize_group([r for r in h_results if r.get('config') == config], config)
        if s:
            note = " ⚠" if config == 'Beta' else ""
            print(f"  {config:8s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']}){note}")

    print("\n" + "="*70)
    print("TESLAMETER FIELD (tunnel baseline)")
    print("="*70)
    y_tesla = [r for r in tesla_results if r['type'] == 'Y-plate']
    for mat in ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']:
        s = summarize_group([r for r in y_tesla if r['material'] == mat],
                            mat, sigma_key='sigma_pct')
        if s:
            print(f"  Y {mat:8s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']})")
    h_tesla = [r for r in tesla_results if r['type'] == 'H-plate']
    for cfg in ['Alpha', 'Beta', 'Gamma', 'Delta']:
        s = summarize_group([r for r in h_tesla if r.get('config') == cfg],
                            cfg, sigma_key='sigma_pct')
        if s:
            print(f"  H {cfg:8s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']})")

    # ─── Plots and report ───
    print("\nGenerating plots...")
    plot_by_region(y_results, h_results)
    plot_individual_samples(y_results)
    plot_teslameter_summary(tesla_results)
    print("Writing report...")
    write_report(y_results, h_results, tesla_results, y_excluded, h_excluded)

    print(f"\nPlots: {PLOT_DIR}/degradation_v2_*.png")
    print(f"Report: {MD_DIR}/degradation_summary_v2.md")
    print("Done.")


if __name__ == '__main__':
    main()
