#!/usr/bin/env python3
"""
Temperature-corrected Helmholtz analysis.

Corrects Helmholtz coil readings using co-located Teslameter temperature
measurements, removing the dominant thermal systematic to reveal radiation-
induced magnet degradation.

Produces:
  A. Temperature-corrected Y-plate time-series by material
  B. Temperature-corrected % change from baseline
  C. Corrected Jul 17 vs Jul 30 group comparison
  D. Corrected intra-plate NdFeB-SmCo differential
  E. Temperature history from Teslameter readings
  F. Per-material mean degradation curve ("money plot")
"""

import os
import re
import glob
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from datetime import datetime
from collections import defaultdict
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
PLOT_DIR = os.path.join(BASE, 'TempCorrected_Plots')
os.makedirs(PLOT_DIR, exist_ok=True)

# Reference temperature for correction (°C)
T_REF = 20.0

# Temperature coefficients of remanence α(Br) [per °C]
ALPHA = {
    'N42EH':   -0.0010,
    'N52SH':   -0.0011,
    'SmCo33H': -0.0004,
    'SmCo35':  -0.0004,
}

# Colorblind-safe palette
CB = {
    'N42EH':   '#EE6677',
    'N52SH':   '#4477AA',
    'SmCo33H': '#228833',
    'SmCo35':  '#CCBB44',
    'NdFeB':   '#AA3377',
    'SmCo':    '#66CCEE',
}
CB_LIST = ['#4477AA', '#EE6677', '#228833', '#CCBB44',
           '#66CCEE', '#AA3377', '#BBBBBB', '#000000']

# Jul 17 plates vs Jul 30 plates
JUL17_PLATES = {4, 6, 7, 9, 12, 16, 17, 18, 21, 22, 25, 34, 36, 38, 39}
JUL30_PLATES = {1, 3, 5, 10, 11, 13, 15, 19, 20, 23, 24, 26, 30, 32, 40}

SENTINEL = 1337  # missing-data sentinel
MIN_BASELINE_MWC = 0.1  # minimum valid Helmholtz reading for % change calc

# ─── Material assignments ────────────────────────────────────────────────────

def load_materials():
    """Load material assignments from the Materials spreadsheet."""
    mat_path = os.path.join(BASE, 'Materials_Arrangements.xlsx')
    wb = openpyxl.load_workbook(mat_path)

    y_materials = {}
    ws = wb['Tunnel - Y Materials']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            plate_num = plate_id.replace('y-', '')
            y_materials[plate_num] = [str(row[i]).strip() if row[i] else ''
                                      for i in range(1, 5)]

    pair_arrangements = {}
    ws2 = wb['Tunnel - Pair Arrangements']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            mat_type = str(row[1]).strip() if row[1] else ''
            configs = [str(row[i]).strip() if row[i] else '' for i in range(2, 6)]
            pair_arrangements[plate_id] = (mat_type, configs)

    return y_materials, pair_arrangements


def get_y_material(y_materials, sample_name):
    """Get material for a Y-plate sample like 'Y-22-4'."""
    m = re.match(r'Y-(\d+)-(\d+)', sample_name)
    if not m:
        return None
    plate = m.group(1)
    slot = int(m.group(2))
    if plate in y_materials and 1 <= slot <= 4:
        return y_materials[plate][slot - 1]
    return None


# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_helmholtz_file(filepath):
    """Parse a Helmholtz .dat file. Returns list of (datetime, value, unit)."""
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            val_match = re.search(r'([+-]?\d+\.?\d*)\s*(mWC|kT|kBG)', rest)
            if val_match:
                value = float(val_match.group(1))
                unit = val_match.group(2)
                rows.append((dt, value, unit))
    return rows


def parse_teslameter_file(filepath):
    """Parse a Teslameter .dat file. Returns list of (datetime, [f1,f2,f3], temp)."""
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            nums = re.findall(r'(-?\d+\.\d+)', rest)
            if len(nums) >= 4:
                rows.append((dt, [float(x) for x in nums[:3]], float(nums[3])))
            elif len(nums) >= 3:
                rows.append((dt, [float(x) for x in nums[:3]], None))
    return rows


def _clean_retakes(data):
    """Remove obvious bad re-takes from a list of (dt, val) tuples."""
    if len(data) < 3:
        return data
    median_val = np.median([v for _, v in data])
    cleaned = []
    skip = set()
    for i in range(len(data)):
        if i in skip:
            continue
        if i + 1 < len(data):
            dt1, v1 = data[i]
            dt2, v2 = data[i + 1]
            if abs((dt2 - dt1).total_seconds()) < 300:
                if abs(v1 - median_val) > abs(v2 - median_val):
                    skip.add(i)
                    continue
                else:
                    skip.add(i + 1)
        cleaned.append(data[i])
    for i in range(len(data)):
        if i not in skip and data[i] not in cleaned:
            cleaned.append(data[i])
    return sorted(cleaned, key=lambda x: x[0])


def get_all_mwc_rows(rows):
    """All mWC rows with re-take cleaning. Excludes sentinel values."""
    all_mwc = [(dt, val) for dt, val, unit in rows
               if unit == 'mWC' and abs(val - SENTINEL) > 1]
    return _clean_retakes(all_mwc)


# ─── Temperature lookup ──────────────────────────────────────────────────────

def build_temperature_lookup():
    """
    Build a temperature lookup dict from all Teslameter files.

    For Y-plate samples (Y-XX-N):
        key = ('Y-XX-N', 'YYYY-MM-DD')
        value = (T_mean, T_std) from front/side/top temperatures on that date

    For pair assemblies (Hn-XX-YY):
        key = ('Hn-XX-YY', 'YYYY-MM-DD')
        value = (T_mean, T_std) from An-XX-YY-1 and An-XX-YY-2 front/side/top

    Also returns a flat list of (datetime, temperature, sample) for Plot E.
    """
    temp_lookup = {}
    all_temp_records = []  # (datetime, temp, sample_name)

    faces = ['front', 'side', 'top']

    # --- Y-plate temperatures ---
    y_tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
    # Find all unique Y-plate samples
    y_samples = set()
    for f in os.listdir(y_tesla_dir):
        m = re.match(r'(Y-\d+-\d+)_(front|side|top)\.dat$', f)
        if m:
            y_samples.add(m.group(1))

    for sample in sorted(y_samples):
        # Collect temperatures by date from all 3 faces
        date_temps = defaultdict(list)  # date_str -> [temps]
        for face in faces:
            fpath = os.path.join(y_tesla_dir, f'{sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            for dt, fields, temp in rows:
                if temp is not None:
                    date_str = dt.strftime('%Y-%m-%d')
                    date_temps[date_str].append(temp)
                    all_temp_records.append((dt, temp, sample))

        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(sample, date_str)] = (np.mean(temps), np.std(temps))

    # --- Pair assembly temperatures ---
    # Hn-XX-YY maps to An-XX-YY-1 and An-XX-YY-2
    pair_tesla_dir = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')
    pair_helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')

    # Find all Hn-XX-YY samples from Helmholtz files
    hn_samples = set()
    for f in os.listdir(pair_helm_dir):
        m = re.match(r'(Hn-\d+-\d+)_helmholtz\.dat$', f)
        if m:
            hn_samples.add(m.group(1))
    # Also Hs
    hs_samples = set()
    for f in os.listdir(pair_helm_dir):
        m = re.match(r'(Hs-\d+-\d+)_helmholtz\.dat$', f)
        if m:
            hs_samples.add(m.group(1))

    for sample_set, prefix_map in [(hn_samples, 'An'), (hs_samples, 'As')]:
        for h_sample in sorted(sample_set):
            # Parse Hn-XX-YY -> plate=XX, slot=YY
            m = re.match(r'H[ns]-(\d+)-(\d+)', h_sample)
            if not m:
                continue
            plate, slot = m.group(1), m.group(2)

            date_temps = defaultdict(list)
            # Look for An-XX-YY-1 and An-XX-YY-2 (or As-)
            for magnet_idx in ['1', '2']:
                a_sample = f'{prefix_map}-{plate}-{slot}-{magnet_idx}'
                for face in faces:
                    fpath = os.path.join(pair_tesla_dir,
                                         f'{a_sample}_{face}.dat')
                    if not os.path.exists(fpath):
                        continue
                    rows = parse_teslameter_file(fpath)
                    for dt, fields, temp in rows:
                        if temp is not None:
                            date_str = dt.strftime('%Y-%m-%d')
                            date_temps[date_str].append(temp)
                            all_temp_records.append((dt, temp, h_sample))

            for date_str, temps in date_temps.items():
                if temps:
                    temp_lookup[(h_sample, date_str)] = (np.mean(temps),
                                                          np.std(temps))

    print(f"  Temperature lookup: {len(temp_lookup)} (sample, date) entries")
    print(f"  Temperature records: {len(all_temp_records)} individual readings")

    return temp_lookup, all_temp_records


# ─── Temperature correction ──────────────────────────────────────────────────

def correct_helmholtz(h_raw, alpha, t_mean, t_std):
    """
    Correct a Helmholtz reading to T_REF.

    H_corr = H_raw / (1 + alpha * (T - T_ref))

    Since alpha < 0 and T > T_ref, denominator < 1, so H_corr > H_raw.

    Returns (H_corrected, dH_uncertainty).
    """
    denom = 1.0 + alpha * (t_mean - T_REF)
    h_corr = h_raw / denom
    # Uncertainty from temperature spread
    dh = abs(h_raw * alpha * t_std / denom**2)
    return h_corr, dh


# ─── Data loading helper ─────────────────────────────────────────────────────

# Pre-deployment cutoff: measurements before this are "baseline"
TUNNEL_START = datetime(2025, 7, 1)

def load_y_plate_data(y_materials, temp_lookup):
    """
    Load all Y-plate Helmholtz data with temperature corrections.

    Returns:
      corrected:   sample -> [(dt, h_raw, h_corr, dh)]
      uncorrected: sample -> [(dt, h_raw)]  (dates without temp)
      baselines:   sample -> (baseline_corr, n_baseline_pts)
      flagged:     set of sample names flagged as outliers
    """
    helm_dir = os.path.join(BASE, 'Y_Plates', 'Helmholtz')
    corrected = {}
    uncorrected = {}

    for f in sorted(os.listdir(helm_dir)):
        if not f.endswith('_helmholtz.dat'):
            continue
        sample = f.replace('_helmholtz.dat', '')
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()
        alpha = ALPHA.get(material)
        if alpha is None:
            continue

        fpath = os.path.join(helm_dir, f)
        rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
        if not rows:
            continue

        corr_list = []
        uncorr_list = []
        for dt, h_raw in rows:
            date_str = dt.strftime('%Y-%m-%d')
            key = (sample, date_str)
            if key in temp_lookup:
                t_mean, t_std = temp_lookup[key]
                h_corr, dh = correct_helmholtz(h_raw, alpha, t_mean, t_std)
                corr_list.append((dt, h_raw, h_corr, dh))
            else:
                uncorr_list.append((dt, h_raw))

        if corr_list:
            corrected[sample] = corr_list
        if uncorr_list:
            uncorrected[sample] = uncorr_list

    # Compute robust baselines: mean of pre-deployment corrected values
    # with intra-sample outlier rejection (reject readings >3σ from median
    # when ≥3 pre-deployment points exist; for 2 points, reject if spread >5%)
    baselines = {}  # sample -> (baseline_mean, baseline_std, n_used, n_rejected)
    for sample, data in corrected.items():
        pre_vals = [d[2] for d in data if d[0] < TUNNEL_START]
        # Filter out near-zero readings first
        pre_vals = [v for v in pre_vals if abs(v) >= MIN_BASELINE_MWC]
        if not pre_vals:
            continue

        if len(pre_vals) >= 3:
            # Median absolute deviation rejection
            med = np.median(pre_vals)
            mad = np.median([abs(v - med) for v in pre_vals])
            # MAD-based threshold (3.5× MAD, with minimum floor of 0.5% of median)
            threshold = max(3.5 * mad, 0.005 * abs(med))
            kept = [v for v in pre_vals if abs(v - med) <= threshold]
            n_rejected = len(pre_vals) - len(kept)
        elif len(pre_vals) == 2:
            # For 2 points: reject if they differ by >5%
            spread = abs(pre_vals[0] - pre_vals[1]) / np.mean(pre_vals) * 100
            if spread > 5.0:
                # Keep the one closer to the material group (handled below)
                # For now, keep both but flag
                kept = pre_vals
                n_rejected = 0
            else:
                kept = pre_vals
                n_rejected = 0
        else:
            kept = pre_vals
            n_rejected = 0

        if kept:
            bl_mean = np.mean(kept)
            bl_std = np.std(kept) if len(kept) > 1 else 0.0
            baselines[sample] = (bl_mean, bl_std, len(kept), n_rejected)

    # Flag outlier samples: baseline deviates >5% from material-group median
    flagged = set()
    materials_list = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']
    for material in materials_list:
        mat_baselines = []
        mat_samples = []
        for sample, (bl, bl_sd, n, n_rej) in baselines.items():
            mat = get_y_material(y_materials, sample)
            if mat and mat.strip() == material:
                mat_baselines.append(bl)
                mat_samples.append(sample)
        if len(mat_baselines) < 3:
            continue
        med = np.median(mat_baselines)
        for sample, bl in zip(mat_samples, mat_baselines):
            if abs(bl - med) / abs(med) > 0.05:
                flagged.add(sample)

    return corrected, uncorrected, baselines, flagged


# ─── Plot A: Temperature-corrected time-series by material ───────────────────

def plot_A_corrected_timeseries(y_materials, corrected, uncorrected, flagged):
    """4 panels, one per material. Raw (dashed) vs corrected (solid)."""
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 2, figsize=(18, 12), sharex=True)
    axes = axes.flatten()

    for ax, material in zip(axes, materials):
        n_samples = 0
        for sample in sorted(corrected.keys()):
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            n_samples += 1
            data = corrected[sample]
            dates = [d[0] for d in data]
            raw_vals = [d[1] for d in data]
            corr_vals = [d[2] for d in data]

            is_flagged = sample in flagged
            # Raw: dashed, faint
            ax.plot(dates, raw_vals, '--', color=CB[material],
                    alpha=0.1 if is_flagged else 0.2, linewidth=0.8)
            # Corrected: solid (flagged samples shown as x markers)
            ax.plot(dates, corr_vals,
                    'x-' if is_flagged else 'o-',
                    color='gray' if is_flagged else CB[material],
                    alpha=0.3 if is_flagged else 0.4,
                    markersize=3 if is_flagged else 2, linewidth=0.8)

        # Compute mean corrected trace (exclude flagged and near-zero samples)
        all_dates_set = set()
        sample_traces = []
        for sample in corrected:
            if sample in flagged:
                continue
            s_data = corrected[sample]
            if any(abs(d[2]) < MIN_BASELINE_MWC for d in s_data[:1]):
                continue
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            trace = {d[0].strftime('%Y-%m-%d'): d[2] for d in s_data}
            sample_traces.append(trace)
            all_dates_set.update(trace.keys())

        if sample_traces:
            sorted_dates = sorted(all_dates_set)
            mean_vals = []
            std_vals = []
            plot_dates = []
            for ds in sorted_dates:
                vals = [t[ds] for t in sample_traces if ds in t]
                if len(vals) >= 2:
                    mean_vals.append(np.mean(vals))
                    std_vals.append(np.std(vals))
                    plot_dates.append(datetime.strptime(ds, '%Y-%m-%d'))

            if plot_dates:
                mean_vals = np.array(mean_vals)
                std_vals = np.array(std_vals)
                ax.plot(plot_dates, mean_vals, 'k-', linewidth=2.5,
                        label='Mean (corrected)', zorder=10)
                ax.fill_between(plot_dates, mean_vals - std_vals,
                                mean_vals + std_vals,
                                color='gray', alpha=0.2, label='±1σ')

        ax.set_title(f'{material} (n={n_samples})', fontsize=13,
                     fontweight='bold')
        ax.set_ylabel('mWC (corrected to 20°C)', fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.legend(fontsize=8, loc='best')

    for ax in axes[2:]:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in ax.get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('Y-Plate Helmholtz Time-Series — Temperature-Corrected to 20°C\n'
                 '(dashed = raw, solid = corrected)',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.94])

    outpath = os.path.join(PLOT_DIR, 'A_corrected_timeseries_by_material.png')
    plt.savefig(outpath, dpi=150)
    print(f"  Saved: {outpath}")
    plt.close()


# ─── Plot B: Temperature-corrected % change from baseline ────────────────────

def plot_B_corrected_pct_change(y_materials, corrected, baselines, flagged):
    """% change from robust pre-deployment baseline, by material."""
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 2, figsize=(18, 12), sharex=True)
    axes = axes.flatten()

    for ax, material in zip(axes, materials):
        pct_traces = []  # unflagged only, for mean
        all_pct_traces = []  # all, for individual lines
        for sample in sorted(corrected.keys()):
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            data = corrected[sample]
            if len(data) < 2:
                continue

            # Use robust baseline (mean of pre-deployment)
            if sample not in baselines:
                continue
            base = baselines[sample][0]
            if abs(base) < MIN_BASELINE_MWC:
                continue

            dates = [d[0] for d in data]
            pcts = [(d[2] - base) / abs(base) * 100 for d in data]

            is_flagged = sample in flagged
            all_pct_traces.append((sample, dates, pcts, is_flagged))
            if not is_flagged:
                pct_traces.append((sample, dates, pcts))

            ax.plot(dates, pcts,
                    'x--' if is_flagged else 'o-',
                    markersize=3 if is_flagged else 2,
                    linewidth=0.7, alpha=0.2 if is_flagged else 0.4,
                    color='gray' if is_flagged else None)

        # Mean trace (unflagged only)
        if pct_traces:
            all_dates = sorted(set(d for _, dates, _ in pct_traces for d in dates))
            mean_by_date = {}
            std_by_date = {}
            for d in all_dates:
                vals = []
                for _, dates, pcts in pct_traces:
                    for dd, pp in zip(dates, pcts):
                        if dd == d:
                            vals.append(pp)
                if len(vals) >= 2:
                    mean_by_date[d] = np.mean(vals)
                    std_by_date[d] = np.std(vals) / np.sqrt(len(vals))

            if mean_by_date:
                md = sorted(mean_by_date.keys())
                mv = [mean_by_date[d] for d in md]
                se = [std_by_date[d] for d in md]
                ax.plot(md, mv, 'k-', linewidth=2.5, label='Mean', zorder=10)
                ax.fill_between(md, np.array(mv) - np.array(se),
                                np.array(mv) + np.array(se),
                                color='gray', alpha=0.3, label='±1 SE')

        n_unflagged = len(pct_traces)
        n_flagged_mat = len(all_pct_traces) - n_unflagged
        title = f'{material} (n={n_unflagged}'
        if n_flagged_mat:
            title += f', {n_flagged_mat} flagged'
        title += ')'
        ax.axhline(0, color='gray', linestyle='--', linewidth=0.8)
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.set_ylabel('% change from baseline (corrected)', fontsize=10)
        ax.grid(True, alpha=0.3)
        if ax.get_legend_handles_labels()[1]:
            ax.legend(fontsize=8)

    for ax in axes[2:]:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in ax.get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('Y-Plate % Change from Baseline — Temperature-Corrected to 20°C',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.96])

    outpath = os.path.join(PLOT_DIR, 'B_corrected_pct_change_by_material.png')
    plt.savefig(outpath, dpi=150)
    print(f"  Saved: {outpath}")
    plt.close()


# ─── Plot C: Corrected Jul 17 vs Jul 30 group comparison ─────────────────────

def plot_C_jul17_vs_jul30_corrected(y_materials, corrected, flagged):
    """
    Bar chart of mean first→last corrected % change, Jul17 vs Jul30 groups.
    Compare with raw version from full_validation_plots.py.
    Excludes flagged outlier samples.
    """
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    # Also compute raw % change for comparison
    group_data_corr = {'Jul 17': defaultdict(list), 'Jul 30': defaultdict(list)}
    group_data_raw = {'Jul 17': defaultdict(list), 'Jul 30': defaultdict(list)}

    for sample, data in corrected.items():
        if sample in flagged:
            continue
        m = re.match(r'Y-(\d+)-(\d+)', sample)
        if not m:
            continue
        plate_num = int(m.group(1))
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()
        if material not in materials:
            continue

        if plate_num in JUL17_PLATES:
            group = 'Jul 17'
        elif plate_num in JUL30_PLATES:
            group = 'Jul 30'
        else:
            continue

        # Only tunnel data (Jul 2025+)
        tunnel_data = [d for d in data if d[0] >= TUNNEL_START]
        if len(tunnel_data) < 2:
            continue

        first_corr = tunnel_data[0][2]
        last_corr = tunnel_data[-1][2]
        first_raw = tunnel_data[0][1]
        last_raw = tunnel_data[-1][1]

        if abs(first_corr) < MIN_BASELINE_MWC or abs(first_raw) < MIN_BASELINE_MWC:
            continue

        pct_corr = (last_corr - first_corr) / abs(first_corr) * 100
        pct_raw = (last_raw - first_raw) / abs(first_raw) * 100

        group_data_corr[group][material].append(pct_corr)
        group_data_raw[group][material].append(pct_raw)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 8))
    x = np.arange(len(materials))
    width = 0.35

    # Left panel: Raw
    for ax, data, title_suffix in [(ax1, group_data_raw, 'Raw'),
                                    (ax2, group_data_corr, 'Corrected to 20°C')]:
        j17_means = [np.mean(data['Jul 17'].get(m, [0])) for m in materials]
        j30_means = [np.mean(data['Jul 30'].get(m, [0])) for m in materials]
        j17_stds = [np.std(data['Jul 17'].get(m, [0])) for m in materials]
        j30_stds = [np.std(data['Jul 30'].get(m, [0])) for m in materials]
        j17_n = [len(data['Jul 17'].get(m, [])) for m in materials]
        j30_n = [len(data['Jul 30'].get(m, [])) for m in materials]

        bars1 = ax.bar(x - width/2, j17_means, width, yerr=j17_stds,
                       capsize=4, label='Jul 17 group', color=CB_LIST[0],
                       alpha=0.8)
        bars2 = ax.bar(x + width/2, j30_means, width, yerr=j30_stds,
                       capsize=4, label='Jul 30 group', color=CB_LIST[1],
                       alpha=0.8)

        for bar, val, n in zip(bars1, j17_means, j17_n):
            ax.annotate(f'{val:+.3f}%\n(n={n})',
                        xy=(bar.get_x() + bar.get_width()/2,
                            bar.get_height()),
                        xytext=(0, 8), textcoords='offset points',
                        ha='center', fontsize=8, fontweight='bold')
        for bar, val, n in zip(bars2, j30_means, j30_n):
            ax.annotate(f'{val:+.3f}%\n(n={n})',
                        xy=(bar.get_x() + bar.get_width()/2,
                            bar.get_height()),
                        xytext=(0, 8), textcoords='offset points',
                        ha='center', fontsize=8, fontweight='bold')

        ax.set_xticks(x)
        ax.set_xticklabels(materials, fontsize=11)
        ax.set_ylabel('Mean % change (first → last tunnel)', fontsize=11)
        ax.set_title(f'{title_suffix}', fontsize=13, fontweight='bold')
        ax.legend(fontsize=10)
        ax.grid(axis='y', alpha=0.3)
        ax.axhline(0, color='black', linewidth=0.8, linestyle='--')

    fig.suptitle('Jul 17 vs Jul 30 Group Systematic — Raw vs Temperature-Corrected',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.95])

    outpath = os.path.join(PLOT_DIR, 'C_jul17_vs_jul30_corrected.png')
    plt.savefig(outpath, dpi=200)
    print(f"  Saved: {outpath}")
    plt.close()

    # Print comparison
    print("\n  Jul17 vs Jul30 — Raw vs Corrected comparison:")
    print(f"  {'Material':10s} | {'Jul17 raw':>10s} | {'Jul17 corr':>11s} | "
          f"{'Jul30 raw':>10s} | {'Jul30 corr':>11s} | {'Δ(groups) raw':>14s} | "
          f"{'Δ(groups) corr':>15s}")
    print("  " + "-" * 95)
    for m_name in materials:
        r17 = np.mean(group_data_raw['Jul 17'].get(m_name, [0]))
        c17 = np.mean(group_data_corr['Jul 17'].get(m_name, [0]))
        r30 = np.mean(group_data_raw['Jul 30'].get(m_name, [0]))
        c30 = np.mean(group_data_corr['Jul 30'].get(m_name, [0]))
        print(f"  {m_name:10s} | {r17:+.4f}% | {c17:+.4f}%  | "
              f"{r30:+.4f}% | {c30:+.4f}%  | {r17-r30:+.4f}%       | "
              f"{c17-c30:+.4f}%")


# ─── Plot D: Corrected intra-plate differential ──────────────────────────────

def plot_D_intraplate_differential(y_materials, corrected, flagged):
    """
    Intra-plate (NdFeB avg - SmCo avg) differential, raw vs corrected.
    Excludes flagged outlier samples.
    """
    # Build per-plate per-material % changes for both raw and corrected
    plate_data_raw = defaultdict(dict)
    plate_data_corr = defaultdict(dict)

    for sample, data in corrected.items():
        if sample in flagged:
            continue
        m_match = re.match(r'Y-(\d+)-(\d+)', sample)
        if not m_match:
            continue
        plate_num = int(m_match.group(1))
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()

        tunnel_data = [d for d in data if d[0] >= TUNNEL_START]
        if len(tunnel_data) < 2:
            continue

        first_raw, last_raw = tunnel_data[0][1], tunnel_data[-1][1]
        first_corr, last_corr = tunnel_data[0][2], tunnel_data[-1][2]

        if abs(first_raw) < MIN_BASELINE_MWC or abs(first_corr) < MIN_BASELINE_MWC:
            continue

        plate_data_raw[plate_num][material] = \
            (last_raw - first_raw) / abs(first_raw) * 100
        plate_data_corr[plate_num][material] = \
            (last_corr - first_corr) / abs(first_corr) * 100

    # Compute differentials
    def compute_diffs(plate_data):
        diffs = []
        labels = []
        for pn in sorted(plate_data.keys()):
            d = plate_data[pn]
            ndfeb = [d.get('N42EH'), d.get('N52SH')]
            smco = [d.get('SmCo33H'), d.get('SmCo35')]
            ndfeb = [v for v in ndfeb if v is not None]
            smco = [v for v in smco if v is not None]
            if ndfeb and smco:
                diffs.append(np.mean(ndfeb) - np.mean(smco))
                labels.append(f'Y-{pn}')
        return diffs, labels

    raw_diffs, raw_labels = compute_diffs(plate_data_raw)
    corr_diffs, corr_labels = compute_diffs(plate_data_corr)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 8))

    # Raw differential
    for ax, diffs, labels, title in [(ax1, raw_diffs, raw_labels, 'Raw'),
                                      (ax2, corr_diffs, corr_labels,
                                       'Corrected to 20°C')]:
        x = np.arange(len(labels))
        colors = [CB['NdFeB'] if d >= 0 else CB['SmCo'] for d in diffs]
        ax.bar(x, diffs, color=colors, alpha=0.8)
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
        ax.axhline(0, color='black', linewidth=0.8, linestyle='--')

        mean_d = np.mean(diffs)
        std_d = np.std(diffs)
        ax.axhline(mean_d, color=CB_LIST[2], linewidth=2, linestyle='-',
                    label=f'Mean: {mean_d:+.3f}% ± {std_d:.3f}%')
        ax.set_ylabel('NdFeB - SmCo differential (%)', fontsize=11)
        ax.set_title(f'{title} (n={len(diffs)})', fontsize=13,
                     fontweight='bold')
        ax.legend(fontsize=10)
        ax.grid(axis='y', alpha=0.3)

    fig.suptitle('Intra-Plate NdFeB - SmCo Differential — Raw vs Temperature-Corrected',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.95])

    outpath = os.path.join(PLOT_DIR, 'D_intraplate_differential_corrected.png')
    plt.savefig(outpath, dpi=200)
    print(f"  Saved: {outpath}")
    plt.close()

    if raw_diffs and corr_diffs:
        print(f"  Intra-plate differential — "
              f"Raw: {np.mean(raw_diffs):+.4f}% ± {np.std(raw_diffs):.4f}%, "
              f"Corrected: {np.mean(corr_diffs):+.4f}% ± {np.std(corr_diffs):.4f}%")


# ─── Plot E: Temperature history ─────────────────────────────────────────────

def plot_E_temperature_history(all_temp_records, y_materials):
    """
    Teslameter temperature readings over time, colored by material/plate.
    """
    # Group Y-plate temps by plate number
    plate_temps = defaultdict(list)  # plate_num -> [(dt, temp)]
    for dt, temp, sample in all_temp_records:
        m = re.match(r'Y-(\d+)-', sample)
        if m:
            plate_temps[int(m.group(1))].append((dt, temp))

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(18, 12), sharex=True)

    # Top: all temperatures, colored by Jul17/Jul30 group
    for plate_num in sorted(plate_temps.keys()):
        pts = sorted(plate_temps[plate_num])
        dates = [p[0] for p in pts]
        temps = [p[1] for p in pts]
        if plate_num in JUL17_PLATES:
            ax1.plot(dates, temps, '.', color=CB_LIST[0], alpha=0.15,
                     markersize=3)
        elif plate_num in JUL30_PLATES:
            ax1.plot(dates, temps, '.', color=CB_LIST[1], alpha=0.15,
                     markersize=3)
        else:
            ax1.plot(dates, temps, '.', color=CB_LIST[6], alpha=0.15,
                     markersize=3)

    # Add dummy entries for legend
    ax1.plot([], [], '.', color=CB_LIST[0], markersize=8, label='Jul 17 group')
    ax1.plot([], [], '.', color=CB_LIST[1], markersize=8, label='Jul 30 group')

    # Compute daily mean temperature
    daily_temps = defaultdict(list)
    for dt, temp, sample in all_temp_records:
        if re.match(r'Y-', sample):
            daily_temps[dt.strftime('%Y-%m-%d')].append(temp)

    daily_dates = sorted(daily_temps.keys())
    daily_means = [np.mean(daily_temps[d]) for d in daily_dates]
    daily_stds = [np.std(daily_temps[d]) for d in daily_dates]
    daily_dt = [datetime.strptime(d, '%Y-%m-%d') for d in daily_dates]

    ax1.plot(daily_dt, daily_means, 'k-', linewidth=2, label='Daily mean',
             zorder=10)
    ax1.fill_between(daily_dt,
                     np.array(daily_means) - np.array(daily_stds),
                     np.array(daily_means) + np.array(daily_stds),
                     color='gray', alpha=0.3)

    ax1.axhline(T_REF, color='red', linewidth=1.5, linestyle='--',
                label=f'T_ref = {T_REF}°C', zorder=5)
    ax1.set_ylabel('Temperature (°C)', fontsize=12)
    ax1.set_title('Teslameter Temperature Readings — Y-Plates', fontsize=13,
                  fontweight='bold')
    ax1.legend(fontsize=9, loc='upper right')
    ax1.grid(True, alpha=0.3)

    # Bottom: temperature correction magnitude (for NdFeB α=-0.11%)
    # Correction factor = 1/(1 + alpha*(T - T_ref)) - 1, as percentage
    corr_pct_n52 = [(1.0 / (1.0 + ALPHA['N52SH'] * (t - T_REF)) - 1.0) * 100
                    for t in daily_means]
    corr_pct_smco = [(1.0 / (1.0 + ALPHA['SmCo35'] * (t - T_REF)) - 1.0) * 100
                     for t in daily_means]

    ax2.plot(daily_dt, corr_pct_n52, 'o-', color=CB['N52SH'], markersize=4,
             linewidth=1.5, label=f'N52SH (α={ALPHA["N52SH"]*100:.2f}%/°C)')
    ax2.plot(daily_dt, corr_pct_smco, 's-', color=CB['SmCo35'], markersize=4,
             linewidth=1.5, label=f'SmCo35 (α={ALPHA["SmCo35"]*100:.2f}%/°C)')
    ax2.axhline(0, color='gray', linewidth=0.8, linestyle='--')
    ax2.set_ylabel('Correction magnitude (%)', fontsize=12)
    ax2.set_xlabel('Date', fontsize=12)
    ax2.set_title('Temperature Correction Magnitude (% added to raw reading)',
                  fontsize=13, fontweight='bold')
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)

    ax2.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
    ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    for label in ax2.get_xticklabels():
        label.set_rotation(45)

    fig.suptitle('Temperature History and Correction Magnitude',
                 fontsize=15, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.96])

    outpath = os.path.join(PLOT_DIR, 'E_temperature_history.png')
    plt.savefig(outpath, dpi=150)
    print(f"  Saved: {outpath}")
    plt.close()


# ─── Plot F: Per-material mean degradation curve ("money plot") ──────────────

def _collect_pct_traces(y_materials, corrected, baselines, flagged, material,
                        use_raw=False):
    """
    Helper: collect per-sample % change traces for a material.
    If use_raw, computes % change from raw values instead of corrected.
    Returns dict: sample -> {date_str: pct_change}
    Only includes unflagged samples with valid baselines.
    """
    traces = {}
    for sample in sorted(corrected.keys()):
        if sample in flagged:
            continue
        mat = get_y_material(y_materials, sample)
        if not mat or mat.strip() != material:
            continue
        data = corrected[sample]
        if len(data) < 2:
            continue
        if sample not in baselines:
            continue
        base_corr = baselines[sample][0]
        if abs(base_corr) < MIN_BASELINE_MWC:
            continue

        # For raw baseline: mean of pre-deployment raw values (same filtering)
        if use_raw:
            pre_raw = [d[1] for d in data
                       if d[0] < TUNNEL_START and abs(d[1]) >= MIN_BASELINE_MWC]
            if not pre_raw:
                continue
            # Apply same MAD outlier rejection as corrected baseline
            if len(pre_raw) >= 3:
                med = np.median(pre_raw)
                mad = np.median([abs(v - med) for v in pre_raw])
                threshold = max(3.5 * mad, 0.005 * abs(med))
                pre_raw = [v for v in pre_raw if abs(v - med) <= threshold]
            base = np.mean(pre_raw) if pre_raw else 0
            if abs(base) < MIN_BASELINE_MWC:
                continue
            trace = {}
            for d in data:
                ds = d[0].strftime('%Y-%m-%d')
                trace[ds] = (d[1] - base) / abs(base) * 100
        else:
            base = base_corr
            trace = {}
            for d in data:
                ds = d[0].strftime('%Y-%m-%d')
                trace[ds] = (d[2] - base) / abs(base) * 100
        traces[sample] = trace
    return traces


def _compute_mean_timeseries(traces):
    """From per-sample trace dicts, compute mean ± SE at each date."""
    all_dates = sorted(set(ds for t in traces.values() for ds in t.keys()))
    results = []  # (datetime, mean, se, n)
    for ds in all_dates:
        vals = [t[ds] for t in traces.values() if ds in t]
        if len(vals) >= 2:
            results.append((datetime.strptime(ds, '%Y-%m-%d'),
                            np.mean(vals),
                            np.std(vals) / np.sqrt(len(vals)),
                            len(vals)))
    return results


def plot_F_mean_degradation(y_materials, corrected, baselines, flagged):
    """
    Mean corrected AND raw % change vs time for each material.
    Error bars = standard error of the mean.
    Two panels: top = corrected, bottom = raw, for direct comparison.
    """
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 14), sharex=True)

    for ax, use_raw, title in [(ax1, False, 'Temperature-Corrected to 20°C'),
                                (ax2, True, 'Raw (no temperature correction)')]:
        for material in materials:
            traces = _collect_pct_traces(y_materials, corrected, baselines,
                                         flagged, material, use_raw=use_raw)
            if not traces:
                continue

            results = _compute_mean_timeseries(traces)
            if not results:
                continue

            dates = [r[0] for r in results]
            means = np.array([r[1] for r in results])
            ses = np.array([r[2] for r in results])

            ax.errorbar(dates, means, yerr=ses,
                        fmt='o-', color=CB[material], markersize=6,
                        linewidth=2, capsize=4, capthick=1.5,
                        label=f'{material} (n={len(traces)})')

        ax.axhline(0, color='gray', linewidth=1, linestyle='--')
        beam_on = datetime(2025, 7, 17)
        ax.axvline(beam_on, color='red', linewidth=1, linestyle=':',
                   alpha=0.7, label='First tunnel campaign')
        ax.set_ylabel('Mean % change from pre-deployment baseline', fontsize=12)
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.legend(fontsize=10, loc='best')
        ax.grid(True, alpha=0.3)

    ax2.set_xlabel('Date', fontsize=13)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
    ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    for label in ax2.get_xticklabels():
        label.set_rotation(45)

    fig.suptitle('Per-Material Mean Degradation — Raw vs Temperature-Corrected\n'
                 '(error bars = standard error of the mean; flagged outliers excluded)',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.95])

    outpath = os.path.join(PLOT_DIR, 'F_mean_degradation_curve.png')
    plt.savefig(outpath, dpi=200)
    print(f"  Saved: {outpath}")
    plt.close()


def print_summary_table(y_materials, corrected, baselines, flagged):
    """Print comprehensive summary tables for both raw and corrected data."""
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    for label, use_raw in [('CORRECTED (to 20°C)', False), ('RAW', True)]:
        print(f"\n  === {label} per-material mean % change from pre-deployment baseline ===")
        for material in materials:
            traces = _collect_pct_traces(y_materials, corrected, baselines,
                                         flagged, material, use_raw=use_raw)
            if not traces:
                continue
            results = _compute_mean_timeseries(traces)
            print(f"\n  {material} (n={len(traces)} samples):")
            for dt, mean, se, n in results:
                ds = dt.strftime('%Y-%m-%d')
                print(f"    {ds}: {mean:+.4f}% ± {se:.4f}% (n={n})")

    # Print flagged samples
    print("\n  === FLAGGED OUTLIER SAMPLES (excluded from group means) ===")
    if not flagged:
        print("    None")
    else:
        for sample in sorted(flagged):
            mat = get_y_material(y_materials, sample) or '?'
            if sample in baselines:
                bl, bl_sd, n_used, n_rej = baselines[sample]
                # Find group median for comparison
                mat_baselines = [baselines[s][0] for s in baselines
                                 if get_y_material(y_materials, s) and
                                 get_y_material(y_materials, s).strip() == mat.strip()
                                 and s not in flagged]
                med = np.median(mat_baselines) if mat_baselines else 0
                dev = (bl - med) / med * 100 if med != 0 else 0
                print(f"    {sample} ({mat.strip()}): baseline={bl:.4f} ± {bl_sd:.4f} mWC "
                      f"(n={n_used}, {n_rej} rejected), "
                      f"group median={med:.4f}, deviation={dev:+.1f}%")


# ─── Teslameter field analysis ────────────────────────────────────────────────

def load_teslameter_field_data(y_materials, temp_lookup):
    """
    Load Teslameter field magnitude data for all Y-plate samples.
    Computes magnitude = sqrt(F1² + F2² + F3²) for each face.

    Returns:
      tesla_data: sample -> [(dt, mag_raw, mag_corr, temp, face)]
        Only tunnel-period data (pre-deployment geometry is different).
    """
    tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
    faces = ['front', 'side', 'top']
    tesla_data = defaultdict(list)  # sample -> [(dt, mag_raw, mag_corr, temp)]

    for f in sorted(os.listdir(tesla_dir)):
        m = re.match(r'(Y-\d+-\d+)_(front|side|top)\.dat$', f)
        if not m:
            continue
        sample = m.group(1)
        face = m.group(2)
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()
        alpha = ALPHA.get(material)
        if alpha is None:
            continue

        fpath = os.path.join(tesla_dir, f)
        rows = parse_teslameter_file(fpath)

        for dt, fields, temp in rows:
            if dt < TUNNEL_START:
                continue
            if temp is None:
                continue
            mag = np.sqrt(sum(fi**2 for fi in fields))
            if mag < 1.0:  # skip near-zero readings
                continue
            # Temperature-correct the magnitude
            denom = 1.0 + alpha * (temp - T_REF)
            mag_corr = mag / denom
            tesla_data[sample].append((dt, mag, mag_corr, temp, face))

    return tesla_data


def plot_G_teslameter_timeseries(y_materials, tesla_data, flagged):
    """
    Teslameter field magnitude time-series by material.
    Shows raw and corrected, similar structure to Helmholtz Plot A.
    Uses average across 3 faces per sample per date.
    """
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 2, figsize=(18, 12), sharex=True)
    axes = axes.flatten()

    for ax, material in zip(axes, materials):
        # Aggregate by (sample, date): average magnitude across faces
        sample_date_data = defaultdict(lambda: {'raw': [], 'corr': []})
        for sample, readings in tesla_data.items():
            if sample in flagged:
                continue
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            for dt, mag_raw, mag_corr, temp, face in readings:
                ds = dt.strftime('%Y-%m-%d')
                key = (sample, ds)
                sample_date_data[key]['raw'].append(mag_raw)
                sample_date_data[key]['corr'].append(mag_corr)
                sample_date_data[key]['dt'] = dt

        # Build per-sample time-series (face-averaged)
        sample_traces = defaultdict(list)  # sample -> [(dt, raw_avg, corr_avg)]
        for (sample, ds), vals in sample_date_data.items():
            sample_traces[sample].append(
                (vals['dt'], np.mean(vals['raw']), np.mean(vals['corr'])))

        n_samples = 0
        for sample in sorted(sample_traces.keys()):
            trace = sorted(sample_traces[sample])
            if len(trace) < 2:
                continue
            n_samples += 1
            dates = [t[0] for t in trace]
            raw_vals = [t[1] for t in trace]
            corr_vals = [t[2] for t in trace]

            ax.plot(dates, raw_vals, '--', color=CB[material], alpha=0.15,
                    linewidth=0.8)
            ax.plot(dates, corr_vals, 'o-', color=CB[material], alpha=0.3,
                    markersize=2, linewidth=0.8)

        ax.set_title(f'{material} (n={n_samples})', fontsize=13,
                     fontweight='bold')
        ax.set_ylabel('Field magnitude (mT, corrected to 20°C)', fontsize=10)
        ax.grid(True, alpha=0.3)

    for ax in axes[2:]:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in ax.get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('Y-Plate Teslameter Field Magnitude — Temperature-Corrected to 20°C\n'
                 '(face-averaged, tunnel period only; dashed = raw, solid = corrected)',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.94])

    outpath = os.path.join(PLOT_DIR, 'G_teslameter_field_timeseries.png')
    plt.savefig(outpath, dpi=150)
    print(f"  Saved: {outpath}")
    plt.close()


def plot_H_teslameter_pct_change(y_materials, tesla_data, flagged):
    """
    Teslameter % change from first tunnel measurement, by material.
    Parallels Helmholtz Plot B for direct comparison.
    """
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 2, figsize=(18, 12), sharex=True)
    axes = axes.flatten()

    for ax, material in zip(axes, materials):
        # Build per-sample face-averaged time-series
        sample_date_data = defaultdict(lambda: {'corr': [], 'dt': None})
        for sample, readings in tesla_data.items():
            if sample in flagged:
                continue
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            for dt, mag_raw, mag_corr, temp, face in readings:
                ds = dt.strftime('%Y-%m-%d')
                key = (sample, ds)
                sample_date_data[key]['corr'].append(mag_corr)
                sample_date_data[key]['dt'] = dt

        sample_traces = defaultdict(list)
        for (sample, ds), vals in sample_date_data.items():
            sample_traces[sample].append(
                (vals['dt'], np.mean(vals['corr'])))

        pct_traces = []
        for sample in sorted(sample_traces.keys()):
            trace = sorted(sample_traces[sample])
            if len(trace) < 2:
                continue
            base = trace[0][1]
            if base < 1.0:
                continue
            dates = [t[0] for t in trace]
            pcts = [(t[1] - base) / base * 100 for t in trace]
            pct_traces.append((sample, dates, pcts))
            ax.plot(dates, pcts, 'o-', markersize=2, linewidth=0.7, alpha=0.4)

        # Mean trace
        if pct_traces:
            all_dates = sorted(set(d for _, dates, _ in pct_traces for d in dates))
            mean_by_date = {}
            std_by_date = {}
            for d in all_dates:
                vals = []
                for _, dates, pcts in pct_traces:
                    for dd, pp in zip(dates, pcts):
                        if dd == d:
                            vals.append(pp)
                if len(vals) >= 2:
                    mean_by_date[d] = np.mean(vals)
                    std_by_date[d] = np.std(vals) / np.sqrt(len(vals))

            if mean_by_date:
                md = sorted(mean_by_date.keys())
                mv = [mean_by_date[d] for d in md]
                se = [std_by_date[d] for d in md]
                ax.plot(md, mv, 'k-', linewidth=2.5, label='Mean', zorder=10)
                ax.fill_between(md, np.array(mv) - np.array(se),
                                np.array(mv) + np.array(se),
                                color='gray', alpha=0.3, label='±1 SE')

        ax.axhline(0, color='gray', linestyle='--', linewidth=0.8)
        n_t = len(pct_traces)
        ax.set_title(f'{material} (n={n_t})', fontsize=13, fontweight='bold')
        ax.set_ylabel('% change from first tunnel meas. (corrected)', fontsize=10)
        ax.grid(True, alpha=0.3)
        if ax.get_legend_handles_labels()[1]:
            ax.legend(fontsize=8)

    for ax in axes[2:]:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in ax.get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('Y-Plate Teslameter % Change — Temperature-Corrected to 20°C\n'
                 '(baseline = first tunnel measurement per sample)',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.94])

    outpath = os.path.join(PLOT_DIR, 'H_teslameter_pct_change.png')
    plt.savefig(outpath, dpi=150)
    print(f"  Saved: {outpath}")
    plt.close()

    # Print summary
    print("\n  Teslameter corrected % change from first tunnel measurement:")
    for material in materials:
        sample_date_data = defaultdict(lambda: {'corr': [], 'dt': None})
        for sample, readings in tesla_data.items():
            if sample in flagged:
                continue
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            for dt, mag_raw, mag_corr, temp, face in readings:
                ds = dt.strftime('%Y-%m-%d')
                sample_date_data[(sample, ds)]['corr'].append(mag_corr)
                sample_date_data[(sample, ds)]['dt'] = dt

        sample_traces = defaultdict(list)
        for (sample, ds), vals in sample_date_data.items():
            sample_traces[sample].append((vals['dt'], np.mean(vals['corr'])))

        by_date = defaultdict(list)
        for sample in sorted(sample_traces.keys()):
            trace = sorted(sample_traces[sample])
            if len(trace) < 2:
                continue
            base = trace[0][1]
            if base < 1.0:
                continue
            for dt, val in trace:
                by_date[dt.strftime('%Y-%m-%d')].append((val - base) / base * 100)

        if by_date:
            print(f"\n  {material}:")
            for ds in sorted(by_date.keys()):
                vals = by_date[ds]
                if len(vals) >= 2:
                    print(f"    {ds}: {np.mean(vals):+.4f}% ± "
                          f"{np.std(vals)/np.sqrt(len(vals)):.4f}% (n={len(vals)})")


def plot_I_helmholtz_vs_teslameter(y_materials, corrected, baselines, flagged,
                                    tesla_data):
    """
    Direct comparison: Helmholtz vs Teslameter % change at each date.
    Both temperature-corrected. Uses first tunnel measurement as baseline
    for both instruments.
    """
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 2, figsize=(18, 12))
    axes = axes.flatten()

    for ax, material in zip(axes, materials):
        # Build Teslameter face-averaged per-sample per-date
        t_sample_date = defaultdict(lambda: {'corr': [], 'dt': None})
        for sample, readings in tesla_data.items():
            if sample in flagged:
                continue
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            for dt, mag_raw, mag_corr, temp, face in readings:
                ds = dt.strftime('%Y-%m-%d')
                t_sample_date[(sample, ds)]['corr'].append(mag_corr)
                t_sample_date[(sample, ds)]['dt'] = dt

        # Build per-sample Teslameter % change from first tunnel
        t_pct = {}  # sample -> {date_str: pct}
        for sample in set(k[0] for k in t_sample_date.keys()):
            trace = []
            for (s, ds), vals in t_sample_date.items():
                if s == sample:
                    trace.append((vals['dt'], np.mean(vals['corr'])))
            trace.sort()
            if len(trace) < 2 or trace[0][1] < 1.0:
                continue
            base = trace[0][1]
            t_pct[sample] = {t[0].strftime('%Y-%m-%d'): (t[1]-base)/base*100
                             for t in trace}

        # Build per-sample Helmholtz % change from first tunnel
        h_pct = {}
        for sample, data in corrected.items():
            if sample in flagged:
                continue
            mat = get_y_material(y_materials, sample)
            if not mat or mat.strip() != material:
                continue
            tunnel = [d for d in data if d[0] >= TUNNEL_START]
            if len(tunnel) < 2:
                continue
            base = tunnel[0][2]
            if abs(base) < MIN_BASELINE_MWC:
                continue
            h_pct[sample] = {d[0].strftime('%Y-%m-%d'): (d[2]-base)/abs(base)*100
                             for d in tunnel}

        # Find samples with both
        common = set(t_pct.keys()) & set(h_pct.keys())
        if not common:
            ax.text(0.5, 0.5, 'No common data', transform=ax.transAxes,
                    ha='center')
            continue

        # Collect paired (helmholtz_pct, teslameter_pct) for each sample+date
        h_vals = []
        t_vals = []
        for sample in common:
            for ds in h_pct[sample]:
                if ds in t_pct[sample]:
                    h_vals.append(h_pct[sample][ds])
                    t_vals.append(t_pct[sample][ds])

        ax.scatter(h_vals, t_vals, alpha=0.3, s=15, color=CB[material])

        # Fit line
        if len(h_vals) > 5:
            coeffs = np.polyfit(h_vals, t_vals, 1)
            xfit = np.array([min(h_vals), max(h_vals)])
            ax.plot(xfit, coeffs[0]*xfit + coeffs[1], 'k--', linewidth=1.5,
                    label=f'slope={coeffs[0]:.2f}')

        # 1:1 line
        lims = [min(min(h_vals), min(t_vals)), max(max(h_vals), max(t_vals))]
        ax.plot(lims, lims, 'gray', linewidth=0.8, linestyle=':', label='1:1')

        ax.set_xlabel('Helmholtz % change (corrected)', fontsize=10)
        ax.set_ylabel('Teslameter % change (corrected)', fontsize=10)
        ax.set_title(f'{material} (n={len(common)} samples)', fontsize=13,
                     fontweight='bold')
        ax.legend(fontsize=8)
        ax.grid(True, alpha=0.3)

    fig.suptitle('Helmholtz vs Teslameter % Change — Temperature-Corrected\n'
                 '(both baselined to first tunnel measurement)',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.94])

    outpath = os.path.join(PLOT_DIR, 'I_helmholtz_vs_teslameter.png')
    plt.savefig(outpath, dpi=200)
    print(f"  Saved: {outpath}")
    plt.close()


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("Temperature-Corrected Helmholtz Analysis")
    print(f"Reference temperature: {T_REF}°C")
    print("=" * 70)

    print("\nStep 1: Loading material assignments...")
    y_materials, pair_arrangements = load_materials()
    print(f"  {len(y_materials)} Y-plate, {len(pair_arrangements)} pair plate entries")

    print("\nStep 2: Building temperature lookup from Teslameter data...")
    temp_lookup, all_temp_records = build_temperature_lookup()

    print("\nStep 3: Loading and correcting Y-plate Helmholtz data...")
    corrected, uncorrected, baselines, flagged = load_y_plate_data(
        y_materials, temp_lookup)
    print(f"  {len(corrected)} samples with temperature-corrected data")
    print(f"  {len(uncorrected)} samples with some uncorrected dates")
    print(f"  {len(baselines)} samples with valid pre-deployment baselines")
    # Report baseline statistics
    n_rejected_total = sum(b[3] for b in baselines.values())
    n_multi = sum(1 for b in baselines.values() if b[2] > 1)
    print(f"    {n_multi} baselines from multiple readings, "
          f"{n_rejected_total} individual readings rejected as outliers")
    print(f"  {len(flagged)} samples flagged as group outliers: {sorted(flagged)}")

    # Verification: check correction direction
    print("\n  Correction direction check (first few non-flagged samples):")
    checked = 0
    for sample in sorted(corrected.keys()):
        if sample in flagged or checked >= 3:
            break
        data = corrected[sample]
        if data:
            d = data[0]
            date_str = d[0].strftime('%Y-%m-%d')
            key = (sample, date_str)
            if key in temp_lookup:
                t_mean = temp_lookup[key][0]
                direction = '↑' if d[2] > d[1] else '↓' if d[2] < d[1] else '='
                expect = '↑' if t_mean > T_REF else '↓'
                ok = '✓' if direction == expect else '✗ UNEXPECTED'
                print(f"    {sample} on {date_str}: T={t_mean:.1f}°C, "
                      f"raw={d[1]:.4f}, corr={d[2]:.4f} ({direction} {ok})")
                checked += 1

    print("\nStep 4: Generating plots...")

    print("\n  Plot A: Corrected time-series by material...")
    plot_A_corrected_timeseries(y_materials, corrected, uncorrected, flagged)

    print("\n  Plot B: Corrected % change from baseline...")
    plot_B_corrected_pct_change(y_materials, corrected, baselines, flagged)

    print("\n  Plot C: Jul 17 vs Jul 30 group — raw vs corrected...")
    plot_C_jul17_vs_jul30_corrected(y_materials, corrected, flagged)

    print("\n  Plot D: Intra-plate differential — raw vs corrected...")
    plot_D_intraplate_differential(y_materials, corrected, flagged)

    print("\n  Plot E: Temperature history...")
    plot_E_temperature_history(all_temp_records, y_materials)

    print("\n  Plot F: Mean degradation curve (money plot)...")
    plot_F_mean_degradation(y_materials, corrected, baselines, flagged)

    print("\nStep 5: Loading Teslameter field data...")
    tesla_data = load_teslameter_field_data(y_materials, temp_lookup)
    n_tesla_samples = len(tesla_data)
    n_tesla_readings = sum(len(v) for v in tesla_data.values())
    print(f"  {n_tesla_samples} Y-plate samples with tunnel Teslameter data")
    print(f"  {n_tesla_readings} individual face readings")

    print("\n  Plot G: Teslameter field time-series...")
    plot_G_teslameter_timeseries(y_materials, tesla_data, flagged)

    print("\n  Plot H: Teslameter % change...")
    plot_H_teslameter_pct_change(y_materials, tesla_data, flagged)

    print("\n  Plot I: Helmholtz vs Teslameter correlation...")
    plot_I_helmholtz_vs_teslameter(y_materials, corrected, baselines, flagged,
                                    tesla_data)

    print("\nStep 6: Summary tables...")
    print_summary_table(y_materials, corrected, baselines, flagged)

    print("\n" + "=" * 70)
    print(f"All plots saved to: {PLOT_DIR}/")
    print("See MD_Files/temperature_correction_analysis.md for full documentation.")
    print("=" * 70)


if __name__ == '__main__':
    main()
