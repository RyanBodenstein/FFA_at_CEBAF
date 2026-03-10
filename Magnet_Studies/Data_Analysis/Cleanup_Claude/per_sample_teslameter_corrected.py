#!/usr/bin/env python3
"""
Per-sample Teslameter plots: raw vs temperature-corrected field magnitude.

For every Y-plate, H-plate, and A-sample, produces figures showing
Teslameter |B| raw (dashed) vs corrected to 20°C (solid) for all 3 faces.

Correction: B_corr = B_raw / (1 + alpha * (T_measured - T_ref))
where T_measured is that face's own temperature reading (field 4).

Y-plates:  4 rows (material slots), 1 column per figure
H-plates:  4 rows (pair assembly slots), 2 columns
           Left: Hn/Hs pair-level Teslameter
           Right: An/As individual magnet Teslameter (both magnets overlaid)

Output: Cleanup_Claude/PerSample_Plots/Teslameter_Corrected/{Y,Hn,Hs}_Plates/
"""

import os
import re
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from collections import defaultdict
import openpyxl

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
PLOT_DIR = os.path.join(BASE, 'PerSample_Plots', 'Teslameter_Corrected')
for sub in ('Y_Plates', 'Hn_Plates', 'Hs_Plates'):
    os.makedirs(os.path.join(PLOT_DIR, sub), exist_ok=True)

# ─── Constants ───────────────────────────────────────────────────────────────
T_REF = 20.0

ALPHA = {
    'N42EH':   -0.0010,
    'N52SH':   -0.0011,
    'SmCo33H': -0.0004,
    'SmCo35':  -0.0004,
    'NdFeB':   -0.00105,
    'SmCo':    -0.0004,
}

SENTINEL = 1337
TUNNEL_START = datetime(2025, 7, 1)

CB = {
    'N42EH':   '#EE6677',
    'N52SH':   '#4477AA',
    'SmCo33H': '#228833',
    'SmCo35':  '#CCBB44',
    'NdFeB':   '#AA3377',
    'SmCo':    '#66CCEE',
}

FACE_STYLE = {
    'front': ('s', '#4477AA', 'Front'),
    'side':  ('^', '#EE6677', 'Side'),
    'top':   ('D', '#228833', 'Top'),
}

# ─── Material assignments ────────────────────────────────────────────────────

def load_materials():
    mat_path = os.path.join(BASE, 'Materials_Arrangements.xlsx')
    wb = openpyxl.load_workbook(mat_path)

    y_materials = {}
    ws = wb['Tunnel - Y Materials']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            plate_num = plate_id.replace('y-', '')
            y_materials[plate_num] = [str(row[i]).strip() if row[i] else ''
                                      for i in range(1, 5)]

    pair_arrangements = {}
    ws2 = wb['Tunnel - Pair Arrangements']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            mat_type = str(row[1]).strip() if row[1] else ''
            configs = [str(row[i]).strip() if row[i] else '' for i in range(2, 6)]
            pair_arrangements[plate_id] = (mat_type, configs)

    return y_materials, pair_arrangements


def get_y_material(y_materials, sample_name):
    m = re.match(r'Y-(\d+)-(\d+)', sample_name)
    if not m:
        return None
    plate = m.group(1)
    slot = int(m.group(2))
    if plate in y_materials and 1 <= slot <= 4:
        return y_materials[plate][slot - 1]
    return None


# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_teslameter_file(filepath):
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            nums = re.findall(r'(-?\d+\.\d+)', rest)
            if len(nums) >= 4:
                rows.append((dt, [float(x) for x in nums[:3]], float(nums[3])))
            elif len(nums) >= 3:
                rows.append((dt, [float(x) for x in nums[:3]], None))
    return rows


# ─── Teslameter loader with correction ───────────────────────────────────────

def load_teslameter_raw_and_corrected(filepath, alpha, tunnel_only=True):
    """
    Load one Teslameter face file, returning raw and corrected |B|.

    Each data point is corrected using its own temperature reading:
      B_corr = B_raw / (1 + alpha * (T - T_ref))

    Returns:
      dates:     list of datetime
      mags_raw:  list of float (raw |B| in mT)
      mags_corr: list of float (corrected |B|, or NaN if no temperature)
      temps:     list of float (temperature in °C, or NaN)
    """
    rows = parse_teslameter_file(filepath)
    dates, mags_raw, mags_corr, temps = [], [], [], []
    for dt, fields, temp in rows:
        if tunnel_only and dt < TUNNEL_START:
            continue
        mag = np.sqrt(sum(f**2 for f in fields))
        dates.append(dt)
        mags_raw.append(mag)
        if temp is not None:
            denom = 1.0 + alpha * (temp - T_REF)
            mags_corr.append(mag / denom)
            temps.append(temp)
        else:
            mags_corr.append(np.nan)
            temps.append(np.nan)
    return dates, mags_raw, mags_corr, temps


# ─── Shared formatting ───────────────────────────────────────────────────────

def format_axes(ax, ylabel, title, show_legend=True):
    ax.set_ylabel(ylabel, fontsize=9)
    ax.set_title(title, fontsize=10, fontweight='bold')
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax.tick_params(axis='x', rotation=45, labelsize=8)
    ax.tick_params(axis='y', labelsize=8)
    ax.grid(True, alpha=0.3)
    if show_legend:
        ax.legend(fontsize=7, loc='best')


# ─── Y-plate figures ─────────────────────────────────────────────────────────

def plot_y_plate_tesla(plate_num, y_materials):
    """One figure per Y-plate: 4 rows (slots), each showing 3 faces raw+corrected."""
    fig, axes = plt.subplots(4, 1, figsize=(12, 18))
    has_data = False

    for slot_idx in range(4):
        slot = slot_idx + 1
        sample = f'Y-{plate_num}-{slot}'
        material = get_y_material(y_materials, sample)
        mat_label = material.strip() if material else '?'
        alpha = ALPHA.get(mat_label, -0.001)

        ax = axes[slot_idx]
        tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
        any_data = False

        for face, (marker, fcolor, flabel) in FACE_STYLE.items():
            fpath = os.path.join(tesla_dir, f'{sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            dates, raw, corr, temps = load_teslameter_raw_and_corrected(
                fpath, alpha, tunnel_only=True)
            if not dates:
                continue
            any_data = True
            has_data = True
            ax.plot(dates, raw, marker=marker, linestyle='--', color=fcolor,
                    markersize=4, linewidth=0.8, alpha=0.4,
                    label=f'{flabel} raw')
            ax.plot(dates, corr, marker=marker, linestyle='-', color=fcolor,
                    markersize=5, linewidth=1.2,
                    label=f'{flabel} corr (20 °C)')

        # Temperature on twin axis
        if any_data:
            ax_temp = ax.twinx()
            for face, (marker, fcolor, flabel) in FACE_STYLE.items():
                fpath = os.path.join(tesla_dir, f'{sample}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                dates, raw, corr, temps = load_teslameter_raw_and_corrected(
                    fpath, alpha, tunnel_only=True)
                if dates and not all(np.isnan(t) for t in temps):
                    ax_temp.plot(dates, temps, marker=marker, linestyle=':',
                                 color=fcolor, markersize=3, alpha=0.3,
                                 linewidth=0.7)
            ax_temp.set_ylabel('Temp (°C)', fontsize=8, color='gray')
            ax_temp.tick_params(axis='y', labelsize=7, colors='gray')

        format_axes(ax, '|B| (mT)',
                    f'{sample}  [{mat_label}] — Teslameter Raw vs Corrected',
                    show_legend=any_data)

    fig.suptitle(f'Y-Plate {plate_num} — Teslameter Temperature Correction',
                 fontsize=14, fontweight='bold', y=0.995)
    fig.tight_layout(rect=[0, 0, 1, 0.98])
    outpath = os.path.join(PLOT_DIR, 'Y_Plates',
                           f'Y-{plate_num:02d}_teslameter.png')
    fig.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return has_data


# ─── H-plate figures ─────────────────────────────────────────────────────────

def plot_h_plate_tesla(plate_num, ns, pair_arrangements):
    """
    One figure per H-plate: 4 rows (pair slots), 2 columns.
    Left: Hn/Hs pair-level Teslameter raw vs corrected.
    Right: An/As individual magnet Teslameter raw vs corrected.
    """
    h_prefix = f'H{ns}'
    a_prefix = f'A{ns}'
    plate_key = f'{ns}-{plate_num}'
    mat_type = 'NdFeB' if ns == 'n' else 'SmCo'
    alpha = ALPHA[mat_type]

    configs = ['', '', '', '']
    if plate_key in pair_arrangements:
        _, configs = pair_arrangements[plate_key]

    fig, axes = plt.subplots(4, 2, figsize=(16, 18))
    has_data = False

    for slot_idx in range(4):
        slot = slot_idx + 1
        h_sample = f'{h_prefix}-{plate_num}-{slot}'
        config = configs[slot_idx] if slot_idx < len(configs) else ''
        config_label = f'  ({config})' if config else ''

        ax_pair = axes[slot_idx, 0]
        ax_indiv = axes[slot_idx, 1]

        tesla_dir = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')

        # --- Left: H-sample pair-level Teslameter ---
        any_pair = False
        for face, (marker, fcolor, flabel) in FACE_STYLE.items():
            fpath = os.path.join(tesla_dir, f'{h_sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            dates, raw, corr, temps = load_teslameter_raw_and_corrected(
                fpath, alpha, tunnel_only=True)
            if not dates:
                continue
            any_pair = True
            has_data = True
            ax_pair.plot(dates, raw, marker=marker, linestyle='--',
                         color=fcolor, markersize=4, linewidth=0.8,
                         alpha=0.4, label=f'{flabel} raw')
            ax_pair.plot(dates, corr, marker=marker, linestyle='-',
                         color=fcolor, markersize=5, linewidth=1.2,
                         label=f'{flabel} corr (20 °C)')

        if any_pair:
            ax_temp = ax_pair.twinx()
            for face, (marker, fcolor, flabel) in FACE_STYLE.items():
                fpath = os.path.join(tesla_dir, f'{h_sample}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                dates, raw, corr, temps = load_teslameter_raw_and_corrected(
                    fpath, alpha, tunnel_only=True)
                if dates and not all(np.isnan(t) for t in temps):
                    ax_temp.plot(dates, temps, marker=marker, linestyle=':',
                                 color=fcolor, markersize=3, alpha=0.3,
                                 linewidth=0.7)
            ax_temp.set_ylabel('Temp (°C)', fontsize=8, color='gray')
            ax_temp.tick_params(axis='y', labelsize=7, colors='gray')

        format_axes(ax_pair, '|B| (mT)',
                    f'{h_sample}{config_label} — Pair Teslameter',
                    show_legend=any_pair)

        # --- Right: A-sample individual magnet Teslameter ---
        any_indiv = False
        mag_markers_extra = {
            '1': {'suffix': ' (mag1)', 'ms_raw': 3, 'ms_corr': 5,
                   'lw_raw': 0.7, 'lw_corr': 1.1},
            '2': {'suffix': ' (mag2)', 'ms_raw': 3, 'ms_corr': 4,
                   'lw_raw': 0.7, 'lw_corr': 1.0},
        }
        # Use slightly shifted colors for mag2 to distinguish
        mag_face_colors = {
            '1': {'front': '#4477AA', 'side': '#EE6677', 'top': '#228833'},
            '2': {'front': '#77AADD', 'side': '#EE9988', 'top': '#66BB66'},
        }

        all_temp_dates, all_temp_vals = [], []

        for mag_idx in ['1', '2']:
            a_sample = f'{a_prefix}-{plate_num}-{slot}-{mag_idx}'
            sty = mag_markers_extra[mag_idx]
            fcols = mag_face_colors[mag_idx]
            for face, (marker, _, flabel) in FACE_STYLE.items():
                fpath = os.path.join(tesla_dir, f'{a_sample}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                dates, raw, corr, temps = load_teslameter_raw_and_corrected(
                    fpath, alpha, tunnel_only=True)
                if not dates:
                    continue
                any_indiv = True
                has_data = True
                fc = fcols[face]
                ax_indiv.plot(dates, raw, marker=marker, linestyle='--',
                              color=fc, markersize=sty['ms_raw'],
                              linewidth=sty['lw_raw'], alpha=0.35,
                              label=f'{flabel}{sty["suffix"]} raw')
                ax_indiv.plot(dates, corr, marker=marker, linestyle='-',
                              color=fc, markersize=sty['ms_corr'],
                              linewidth=sty['lw_corr'],
                              label=f'{flabel}{sty["suffix"]} corr')
                for d, t in zip(dates, temps):
                    if not np.isnan(t):
                        all_temp_dates.append(d)
                        all_temp_vals.append(t)

        if all_temp_dates:
            ax_temp2 = ax_indiv.twinx()
            ax_temp2.plot(all_temp_dates, all_temp_vals, 'x:',
                          color='gray', markersize=3, alpha=0.3,
                          linewidth=0.5)
            ax_temp2.set_ylabel('Temp (°C)', fontsize=8, color='gray')
            ax_temp2.tick_params(axis='y', labelsize=7, colors='gray')

        format_axes(ax_indiv, '|B| (mT)',
                    f'{a_prefix}-{plate_num}-{slot}-{{1,2}}{config_label}'
                    f' — Individual Magnets',
                    show_legend=any_indiv)

    fig.suptitle(f'{h_prefix}-Plate {plate_num}  ({mat_type})'
                 f' — Teslameter Temperature Correction',
                 fontsize=14, fontweight='bold', y=0.995)
    fig.tight_layout(rect=[0, 0, 1, 0.98])

    subdir = 'Hn_Plates' if ns == 'n' else 'Hs_Plates'
    outpath = os.path.join(PLOT_DIR, subdir,
                           f'{h_prefix}-{plate_num:02d}_teslameter.png')
    fig.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return has_data


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Loading materials...")
    y_materials, pair_arrangements = load_materials()

    # ── Y-plates ──
    helm_dir_y = os.path.join(BASE, 'Y_Plates', 'Teslameter')
    y_plate_nums = set()
    for f in os.listdir(helm_dir_y):
        m = re.match(r'Y-(\d+)-\d+_(front|side|top)\.dat$', f)
        if m:
            y_plate_nums.add(int(m.group(1)))

    print(f"\nGenerating Y-plate Teslameter figures ({len(y_plate_nums)} plates)...")
    for pn in sorted(y_plate_nums):
        ok = plot_y_plate_tesla(pn, y_materials)
        status = "OK" if ok else "no data"
        print(f"  Y-{pn:02d}: {status}")

    # ── H-plates ──
    tesla_dir_p = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')

    for ns, label in [('n', 'NdFeB'), ('s', 'SmCo')]:
        h_prefix = f'H{ns}'
        a_prefix = f'A{ns}'
        # Collect plate numbers from both Hn/Hs and An/As Teslameter files
        plate_nums = set()
        for f in os.listdir(tesla_dir_p):
            m = re.match(rf'{h_prefix}-(\d+)-\d+_(front|side|top)\.dat$', f)
            if m:
                plate_nums.add(int(m.group(1)))
            m = re.match(rf'{a_prefix}-(\d+)-\d+-\d+_(front|side|top)\.dat$', f)
            if m:
                plate_nums.add(int(m.group(1)))

        print(f"\nGenerating {h_prefix}-plate Teslameter figures "
              f"({len(plate_nums)} plates)...")
        for pn in sorted(plate_nums):
            ok = plot_h_plate_tesla(pn, ns, pair_arrangements)
            status = "OK" if ok else "no data"
            print(f"  {h_prefix}-{pn:02d}: {status}")

    print(f"\nAll plots saved to {PLOT_DIR}/")
    print("Done.")


if __name__ == '__main__':
    main()
