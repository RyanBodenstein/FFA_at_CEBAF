#!/usr/bin/env python3
"""
Phase 2: Parse rod dosimetry spreadsheet.

Reads LDRD_11102025_commentfix.xlsx sheet DataInput.
Identifies LDRD rods by matching R-### in Notes column against rod IDs
found in Teslameter/Helmholtz .dat files (Phase 1).

Columns:
  1  ProjectNumber     — project ID (39, 20, etc.)
  2  LocationNumber    — sequential index (Project 20) or R-### (Project 39)
  3  Range             — 'Low' or 'High' (rod sensitivity)
  4  DateRead          — date rod was read
  5  600nm-1           — absorbance at 600nm, replicate 1
  6  656nm-1           — absorbance at 656nm, replicate 1
  7  600nm-2           — absorbance at 600nm, replicate 2
  8  656nm-2           — absorbance at 656nm, replicate 2
  9  Notes             — rod ID (R-###) or descriptive text
  10 600_slope         — calibration slope for 600nm
  11 600_intercept     — calibration intercept for 600nm
  12 656_slope         — calibration slope for 656nm
  13 656_intercept     — calibration intercept for 656nm
  14 600_rad-1         — dose from 600nm reading 1 (krad)
  15 656_rad-1         — dose from 656nm reading 1 (krad)
  16 600_rad-2         — dose from 600nm reading 2 (krad)
  17 656_rad-2         — dose from 656nm reading 2 (krad)

Dose formula: dose_krad = OD * slope - intercept

Validity ranges (from Kirsten's calibration, units = krad):
  600nm Low:   2 – 39 krad
  656nm Low:  30 – ∞ krad
  600nm High: 12 – 300 krad
  656nm High: 132 – ∞ krad

Output:
  Rod_Dosimetry/rod_doses.csv — per-rod doses with validity flags
"""

import os
import re
import csv
import openpyxl
import numpy as np

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CLEANUP = os.path.join(BASE, 'Cleanup_Claude')
OUT_DIR = os.path.join(CLEANUP, 'Rod_Dosimetry')
os.makedirs(OUT_DIR, exist_ok=True)

SPREADSHEET = os.path.join(BASE, 'Radiation Info', 'Dosimetry Reports', 'Rods',
                           'LDRD_11102025_commentfix.xlsx')

# Validity ranges: (min_krad, max_krad) — None means no upper bound
VALIDITY = {
    ('Low', '600nm'):  (2.0, 39.0),
    ('Low', '656nm'):  (30.0, None),
    ('High', '600nm'): (12.0, 300.0),
    ('High', '656nm'): (132.0, None),
}

# Known data entry errors (row numbers, 1-indexed including header)
KNOWN_OUTLIER_ROWS = {1424}  # flagged in plan


def load_dat_rod_ids():
    """Load the set of R-### IDs found in .dat files (Phase 1 output)."""
    rod_map_csv = os.path.join(OUT_DIR, 'rod_plate_map.csv')
    rod_ids = set()
    with open(rod_map_csv) as f:
        for row in csv.DictReader(f):
            rod_ids.add(row['rod_id'])
    return rod_ids


def extract_rod_id(notes_str):
    """Extract R-### from Notes column string.
    Handles variations: 'R-525', 'R525', 'R- 525', 'R-525 - Low', etc.
    Returns normalized 'R-###' or None.
    """
    if not notes_str or str(notes_str) == 'None':
        return None
    notes_str = str(notes_str)
    m = re.search(r'R-?\s*(\d+)', notes_str, re.IGNORECASE)
    if m:
        return 'R-%s' % m.group(1)
    return None


def flag_validity(dose_krad, rod_range, wavelength):
    """Check if a dose reading is within validity range.

    Returns:
      'valid'       — within calibration range
      'below_min'   — below minimum detection (noise)
      'above_max'   — above maximum calibration range
      'no_data'     — dose is None or NaN
    """
    if dose_krad is None or (isinstance(dose_krad, float) and np.isnan(dose_krad)):
        return 'no_data'
    key = (rod_range, wavelength)
    if key not in VALIDITY:
        return 'no_data'
    vmin, vmax = VALIDITY[key]
    if dose_krad < vmin:
        return 'below_min'
    if vmax is not None and dose_krad > vmax:
        return 'above_max'
    return 'valid'


def main():
    print("=" * 70)
    print("Phase 2: Parse Rod Spreadsheet")
    print("=" * 70)

    # Load R-### IDs from .dat files
    dat_rod_ids = load_dat_rod_ids()
    print("  R-### IDs from .dat files: %d" % len(dat_rod_ids))

    # Load spreadsheet
    print("\n  Loading spreadsheet...")
    wb = openpyxl.load_workbook(SPREADSHEET, read_only=True, data_only=True)
    ws = wb['DataInput']

    rows_out = []
    stats = {
        'total': 0, 'with_rod_id': 0, 'ldrd_match': 0,
        'low_range': 0, 'high_range': 0,
        'valid_600_low': 0, 'valid_600_high': 0,
        'valid_656_low': 0, 'valid_656_high': 0,
        'skipped_outlier': 0,
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        stats['total'] += 1
        row_num = row[0].row

        # Skip known outlier rows
        if row_num in KNOWN_OUTLIER_ROWS:
            stats['skipped_outlier'] += 1
            continue

        proj = row[0].value
        loc = row[1].value
        rod_range = str(row[2].value).strip() if row[2].value else ''
        date_read = row[3].value
        od600_1 = row[4].value
        od656_1 = row[5].value
        od600_2 = row[6].value
        od656_2 = row[7].value
        notes = row[8].value
        slope600 = row[9].value
        int600 = row[10].value
        slope656 = row[11].value
        int656 = row[12].value
        dose600_1 = row[13].value  # krad
        dose656_1 = row[14].value  # krad
        dose600_2 = row[15].value  # krad
        dose656_2 = row[16].value  # krad

        # Extract rod ID
        rod_id = extract_rod_id(notes)
        if rod_id:
            stats['with_rod_id'] += 1

        # Check if this is an LDRD rod
        is_ldrd = rod_id is not None and rod_id in dat_rod_ids
        if is_ldrd:
            stats['ldrd_match'] += 1

        if rod_range == 'Low':
            stats['low_range'] += 1
        elif rod_range == 'High':
            stats['high_range'] += 1

        # Compute average of replicates for each wavelength
        def safe_avg(a, b):
            vals = [v for v in [a, b] if v is not None and isinstance(v, (int, float))]
            if not vals:
                return None
            return sum(vals) / len(vals)

        avg_600 = safe_avg(dose600_1, dose600_2)
        avg_656 = safe_avg(dose656_1, dose656_2)

        # Flag validity
        flag_600 = flag_validity(avg_600, rod_range, '600nm')
        flag_656 = flag_validity(avg_656, rod_range, '656nm')

        if is_ldrd:
            if flag_600 == 'valid' and rod_range == 'Low':
                stats['valid_600_low'] += 1
            elif flag_600 == 'valid' and rod_range == 'High':
                stats['valid_600_high'] += 1
            if flag_656 == 'valid' and rod_range == 'Low':
                stats['valid_656_low'] += 1
            elif flag_656 == 'valid' and rod_range == 'High':
                stats['valid_656_high'] += 1

        # Determine best dose estimate:
        # Priority: 600nm preferred (lower threshold), then 656nm
        # Only use readings within validity range
        best_dose_krad = None
        best_source = ''
        if flag_600 == 'valid':
            best_dose_krad = avg_600
            best_source = '600nm'
        elif flag_656 == 'valid':
            best_dose_krad = avg_656
            best_source = '656nm'

        # Format date
        date_str = ''
        if date_read:
            if hasattr(date_read, 'strftime'):
                date_str = date_read.strftime('%Y-%m-%d')
            else:
                date_str = str(date_read)

        rows_out.append({
            'row_num': row_num,
            'project': proj if proj else '',
            'location': loc if loc else '',
            'rod_id': rod_id if rod_id else '',
            'is_ldrd': is_ldrd,
            'range': rod_range,
            'date_read': date_str,
            'od_600nm_1': '%.4f' % od600_1 if isinstance(od600_1, (int, float)) else '',
            'od_600nm_2': '%.4f' % od600_2 if isinstance(od600_2, (int, float)) else '',
            'od_656nm_1': '%.4f' % od656_1 if isinstance(od656_1, (int, float)) else '',
            'od_656nm_2': '%.4f' % od656_2 if isinstance(od656_2, (int, float)) else '',
            'dose_600_1_krad': '%.3f' % dose600_1 if isinstance(dose600_1, (int, float)) else '',
            'dose_600_2_krad': '%.3f' % dose600_2 if isinstance(dose600_2, (int, float)) else '',
            'dose_656_1_krad': '%.3f' % dose656_1 if isinstance(dose656_1, (int, float)) else '',
            'dose_656_2_krad': '%.3f' % dose656_2 if isinstance(dose656_2, (int, float)) else '',
            'avg_dose_600_krad': '%.3f' % avg_600 if avg_600 is not None else '',
            'avg_dose_656_krad': '%.3f' % avg_656 if avg_656 is not None else '',
            'flag_600': flag_600,
            'flag_656': flag_656,
            'best_dose_krad': '%.3f' % best_dose_krad if best_dose_krad is not None else '',
            'best_source': best_source,
            'notes': str(notes)[:80] if notes else '',
        })

    wb.close()

    # Write output
    out_csv = os.path.join(OUT_DIR, 'rod_doses.csv')
    fieldnames = [
        'row_num', 'project', 'location', 'rod_id', 'is_ldrd', 'range', 'date_read',
        'od_600nm_1', 'od_600nm_2', 'od_656nm_1', 'od_656nm_2',
        'dose_600_1_krad', 'dose_600_2_krad', 'dose_656_1_krad', 'dose_656_2_krad',
        'avg_dose_600_krad', 'avg_dose_656_krad',
        'flag_600', 'flag_656',
        'best_dose_krad', 'best_source', 'notes',
    ]
    with open(out_csv, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows_out)
    print("\n  Wrote %s (%d rows)" % (out_csv, len(rows_out)))

    # Summary
    print("\n--- Statistics ---")
    print("  Total spreadsheet rows: %d" % stats['total'])
    print("  Rows with R-### in Notes: %d" % stats['with_rod_id'])
    print("  LDRD rods (match .dat files): %d" % stats['ldrd_match'])
    print("  Low-range rows: %d" % stats['low_range'])
    print("  High-range rows: %d" % stats['high_range'])
    print("  Skipped outlier rows: %d" % stats['skipped_outlier'])

    ldrd_rows = [r for r in rows_out if r['is_ldrd']]
    print("\n--- LDRD Rod Validity (of %d LDRD rows) ---" % len(ldrd_rows))
    print("  600nm valid, Low range:  %d" % stats['valid_600_low'])
    print("  600nm valid, High range: %d" % stats['valid_600_high'])
    print("  656nm valid, Low range:  %d" % stats['valid_656_low'])
    print("  656nm valid, High range: %d" % stats['valid_656_high'])

    # Show LDRD rods with valid readings
    ldrd_valid = [r for r in ldrd_rows if r['best_dose_krad']]
    print("\n  LDRD rods with valid dose: %d / %d (%.1f%%)" % (
        len(ldrd_valid), len(ldrd_rows),
        100.0 * len(ldrd_valid) / len(ldrd_rows) if ldrd_rows else 0))

    # Show dose range for valid LDRD rods
    if ldrd_valid:
        doses = [float(r['best_dose_krad']) for r in ldrd_valid]
        print("  Valid dose range: %.1f – %.1f krad" % (min(doses), max(doses)))
        print("  Valid dose median: %.1f krad" % np.median(doses))

    # Check: show some high-dose LDRD rods
    print("\n--- Top 15 LDRD rods by dose ---")
    ldrd_valid_sorted = sorted(ldrd_valid, key=lambda r: -float(r['best_dose_krad']))
    for r in ldrd_valid_sorted[:15]:
        print("  %s (%s, %s): %.1f krad via %s  [flag600=%s, flag656=%s]" % (
            r['rod_id'], r['range'], r['date_read'],
            float(r['best_dose_krad']), r['best_source'],
            r['flag_600'], r['flag_656']))

    print("\nDone.")


if __name__ == '__main__':
    main()
