#!/usr/bin/env python3
"""
Per-sample plots: raw vs temperature-corrected Helmholtz + Teslameter faces.

For every Y-plate and H-plate, produces a multi-panel figure showing:
  - Left column:  Helmholtz readings (raw as gray dashed, corrected as solid
                   with error bars from temperature uncertainty)
  - Right column: Teslameter field magnitude per face (front, side, top),
                   with measurement temperature on a twin y-axis

Y-plates:  4 rows (one per material slot) × 2 columns per figure, 30 figures
H-plates:  4 rows (one per pair assembly slot) × 2 columns, with Hn (pair)
           and An/As (individual magnets) overlaid

Output: Cleanup_Claude/PerSample_Plots/{Y_Plates,Hn_Plates,Hs_Plates}/
"""

import os
import re
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from collections import defaultdict
import openpyxl

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
PLOT_DIR = os.path.join(BASE, 'PerSample_Plots')
for sub in ('Y_Plates', 'Hn_Plates', 'Hs_Plates'):
    os.makedirs(os.path.join(PLOT_DIR, sub), exist_ok=True)

# ─── Constants ───────────────────────────────────────────────────────────────
T_REF = 20.0

# Temperature coefficients of remanence α(Br) [per °C]
ALPHA = {
    'N42EH':   -0.0010,
    'N52SH':   -0.0011,
    'SmCo33H': -0.0004,
    'SmCo35':  -0.0004,
    'NdFeB':   -0.00105,   # average of N42EH/N52SH for pair assemblies
    'SmCo':    -0.0004,
}

SENTINEL = 1337
MIN_BASELINE_MWC = 0.1

# Pre-deployment Teslameter readings are invalid (broken Hall probe).
# Only plot Teslameter data from tunnel period onward.
TUNNEL_START = datetime(2025, 7, 1)

# Colorblind-safe palette
CB = {
    'N42EH':   '#EE6677',
    'N52SH':   '#4477AA',
    'SmCo33H': '#228833',
    'SmCo35':  '#CCBB44',
    'NdFeB':   '#AA3377',
    'SmCo':    '#66CCEE',
}

# Teslameter face colors/markers
FACE_STYLE = {
    'front': ('s', '#4477AA', 'Front'),
    'side':  ('^', '#EE6677', 'Side'),
    'top':   ('D', '#228833', 'Top'),
}

# ─── Material assignments ────────────────────────────────────────────────────

def load_materials():
    """Load material assignments from the Materials spreadsheet."""
    mat_path = os.path.join(BASE, 'Materials_Arrangements.xlsx')
    wb = openpyxl.load_workbook(mat_path)

    y_materials = {}
    ws = wb['Tunnel - Y Materials']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            plate_num = plate_id.replace('y-', '')
            y_materials[plate_num] = [str(row[i]).strip() if row[i] else ''
                                      for i in range(1, 5)]

    pair_arrangements = {}
    ws2 = wb['Tunnel - Pair Arrangements']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            mat_type = str(row[1]).strip() if row[1] else ''
            configs = [str(row[i]).strip() if row[i] else '' for i in range(2, 6)]
            pair_arrangements[plate_id] = (mat_type, configs)

    return y_materials, pair_arrangements


def get_y_material(y_materials, sample_name):
    """Get material for a Y-plate sample like 'Y-22-4'."""
    m = re.match(r'Y-(\d+)-(\d+)', sample_name)
    if not m:
        return None
    plate = m.group(1)
    slot = int(m.group(2))
    if plate in y_materials and 1 <= slot <= 4:
        return y_materials[plate][slot - 1]
    return None


# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_helmholtz_file(filepath):
    """Parse a Helmholtz .dat file. Returns list of (datetime, value, unit)."""
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            val_match = re.search(r'([+-]?\d+\.?\d*)\s*(mWC|kT|kBG)', rest)
            if val_match:
                value = float(val_match.group(1))
                unit = val_match.group(2)
                rows.append((dt, value, unit))
    return rows


def parse_teslameter_file(filepath):
    """Parse a Teslameter .dat file. Returns list of (datetime, [f1,f2,f3], temp)."""
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            nums = re.findall(r'(-?\d+\.\d+)', rest)
            if len(nums) >= 4:
                rows.append((dt, [float(x) for x in nums[:3]], float(nums[3])))
            elif len(nums) >= 3:
                rows.append((dt, [float(x) for x in nums[:3]], None))
    return rows


def _clean_retakes(data):
    """Remove obvious bad re-takes from a list of (dt, val) tuples."""
    if len(data) < 3:
        return data
    median_val = np.median([v for _, v in data])
    cleaned = []
    skip = set()
    for i in range(len(data)):
        if i in skip:
            continue
        if i + 1 < len(data):
            dt1, v1 = data[i]
            dt2, v2 = data[i + 1]
            if abs((dt2 - dt1).total_seconds()) < 300:
                if abs(v1 - median_val) > abs(v2 - median_val):
                    skip.add(i)
                    continue
                else:
                    skip.add(i + 1)
        cleaned.append(data[i])
    for i in range(len(data)):
        if i not in skip and data[i] not in cleaned:
            cleaned.append(data[i])
    return sorted(cleaned, key=lambda x: x[0])


def get_all_mwc_rows(rows):
    """All mWC rows with re-take cleaning. Excludes sentinel values."""
    all_mwc = [(dt, val) for dt, val, unit in rows
               if unit == 'mWC' and abs(val - SENTINEL) > 1]
    return _clean_retakes(all_mwc)


# ─── Temperature lookup ──────────────────────────────────────────────────────

def build_temperature_lookup():
    """
    Build per-sample, per-date temperature lookup from Teslameter files.

    Returns dict: (sample_name, 'YYYY-MM-DD') -> (T_mean, T_std)
    where T_mean/T_std are computed from up to 3 face temperatures for that
    specific sample on that date.
    """
    temp_lookup = {}
    faces = ['front', 'side', 'top']

    # --- Y-plate temperatures ---
    y_tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
    y_samples = set()
    for f in os.listdir(y_tesla_dir):
        m = re.match(r'(Y-\d+-\d+)_(front|side|top)\.dat$', f)
        if m:
            y_samples.add(m.group(1))

    for sample in sorted(y_samples):
        date_temps = defaultdict(list)
        for face in faces:
            fpath = os.path.join(y_tesla_dir, f'{sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            for dt, fields, temp in rows:
                if temp is not None:
                    date_str = dt.strftime('%Y-%m-%d')
                    date_temps[date_str].append(temp)
        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(sample, date_str)] = (np.mean(temps), np.std(temps))

    # --- Pair assembly temperatures ---
    # For Hn-XX-YY, average An-XX-YY-1 and An-XX-YY-2 temperatures
    # For individual An-XX-YY-Z, use that magnet's own temperatures
    pair_tesla_dir = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')

    # Build A-sample temperature lookup (individual magnets)
    a_samples = set()
    for f in os.listdir(pair_tesla_dir):
        m = re.match(r'(A[ns]?-\d+-\d+-\d+)_(front|side|top)\.dat$', f)
        if m:
            a_samples.add(m.group(1))

    for a_sample in sorted(a_samples):
        date_temps = defaultdict(list)
        for face in faces:
            fpath = os.path.join(pair_tesla_dir, f'{a_sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            for dt, fields, temp in rows:
                if temp is not None:
                    date_str = dt.strftime('%Y-%m-%d')
                    date_temps[date_str].append(temp)
        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(a_sample, date_str)] = (np.mean(temps),
                                                      np.std(temps))

    # Build H-sample temperature lookup from constituent A magnets
    pair_helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')
    for f in os.listdir(pair_helm_dir):
        m = re.match(r'(H[ns]-\d+-\d+)_helmholtz\.dat$', f)
        if not m:
            continue
        h_sample = m.group(1)
        # Parse Hn-XX-YY -> prefix, plate, slot
        hm = re.match(r'H([ns])-(\d+)-(\d+)', h_sample)
        if not hm:
            continue
        ns, plate, slot = hm.group(1), hm.group(2), hm.group(3)
        a_prefix = 'An' if ns == 'n' else 'As'

        date_temps = defaultdict(list)
        for magnet_idx in ['1', '2']:
            a_name = f'{a_prefix}-{plate}-{slot}-{magnet_idx}'
            for face in faces:
                fpath = os.path.join(pair_tesla_dir, f'{a_name}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                rows = parse_teslameter_file(fpath)
                for dt, fields, temp in rows:
                    if temp is not None:
                        date_str = dt.strftime('%Y-%m-%d')
                        date_temps[date_str].append(temp)

        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(h_sample, date_str)] = (np.mean(temps),
                                                      np.std(temps))

    print(f"  Temperature lookup: {len(temp_lookup)} (sample, date) entries")
    return temp_lookup


# ─── Temperature correction ──────────────────────────────────────────────────

def correct_helmholtz(h_raw, alpha, t_mean, t_std):
    """
    Correct a Helmholtz reading to T_REF.

    H_corr = H_raw / (1 + alpha * (T - T_ref))

    Returns (H_corrected, dH_uncertainty).
    """
    denom = 1.0 + alpha * (t_mean - T_REF)
    h_corr = h_raw / denom
    dh = abs(h_raw * alpha * t_std / denom**2)
    return h_corr, dh


# ─── Helmholtz data loader ───────────────────────────────────────────────────

def load_helmholtz_sample(filepath, sample_name, alpha, temp_lookup):
    """
    Load one Helmholtz sample, returning raw and corrected data.

    Returns:
      dates_raw:   list of datetime
      vals_raw:    list of float (mWC)
      dates_corr:  list of datetime  (subset with temperature available)
      vals_corr:   list of float (corrected mWC)
      errs_corr:   list of float (uncertainty from temperature spread)
    """
    rows = get_all_mwc_rows(parse_helmholtz_file(filepath))
    dates_raw, vals_raw = [], []
    dates_corr, vals_corr, errs_corr = [], [], []

    for dt, h_raw in rows:
        dates_raw.append(dt)
        vals_raw.append(h_raw)
        date_str = dt.strftime('%Y-%m-%d')
        key = (sample_name, date_str)
        if key in temp_lookup:
            t_mean, t_std = temp_lookup[key]
            h_corr, dh = correct_helmholtz(h_raw, alpha, t_mean, t_std)
            dates_corr.append(dt)
            vals_corr.append(h_corr)
            errs_corr.append(dh)

    return dates_raw, vals_raw, dates_corr, vals_corr, errs_corr


# ─── Teslameter data loader ──────────────────────────────────────────────────

def load_teslameter_face(filepath, tunnel_only=True):
    """
    Load one Teslameter face file.

    Returns:
      dates:  list of datetime
      mags:   list of float (field magnitude |B| = sqrt(f1²+f2²+f3²))
      temps:  list of float (temperature in °C)
    """
    rows = parse_teslameter_file(filepath)
    dates, mags, temps = [], [], []
    for dt, fields, temp in rows:
        if tunnel_only and dt < TUNNEL_START:
            continue
        dates.append(dt)
        mags.append(np.sqrt(sum(f**2 for f in fields)))
        temps.append(temp if temp is not None else np.nan)
    return dates, mags, temps


# ─── Shared formatting ───────────────────────────────────────────────────────

def format_axes(ax, ylabel, title, show_legend=True):
    """Apply consistent formatting to an axis."""
    ax.set_ylabel(ylabel, fontsize=9)
    ax.set_title(title, fontsize=10, fontweight='bold')
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax.tick_params(axis='x', rotation=45, labelsize=8)
    ax.tick_params(axis='y', labelsize=8)
    ax.grid(True, alpha=0.3)
    if show_legend:
        ax.legend(fontsize=7, loc='best')


# ─── Y-plate figures ─────────────────────────────────────────────────────────

def plot_y_plate(plate_num, y_materials, temp_lookup):
    """Generate per-plate figure for Y-plates (4 slots × 2 columns)."""
    fig, axes = plt.subplots(4, 2, figsize=(16, 18))

    plate_str = str(plate_num)
    has_data = False

    for slot_idx in range(4):
        slot = slot_idx + 1
        sample = f'Y-{plate_num}-{slot}'
        material = get_y_material(y_materials, sample)
        mat_label = material.strip() if material else '?'
        alpha = ALPHA.get(mat_label, -0.001)
        color = CB.get(mat_label, '#888888')

        ax_helm = axes[slot_idx, 0]
        ax_tesla = axes[slot_idx, 1]

        # --- Helmholtz panel ---
        helm_path = os.path.join(BASE, 'Y_Plates', 'Helmholtz',
                                 f'{sample}_helmholtz.dat')
        if os.path.exists(helm_path):
            dr, vr, dc, vc, ec = load_helmholtz_sample(
                helm_path, sample, alpha, temp_lookup)
            if dr:
                has_data = True
                ax_helm.plot(dr, vr, 'o--', color='gray', alpha=0.5,
                             markersize=4, label='Raw')
            if dc:
                ax_helm.errorbar(dc, vc, yerr=ec, fmt='o-', color=color,
                                 markersize=5, capsize=3, linewidth=1.2,
                                 label='Corrected (20 °C)')

        format_axes(ax_helm, 'Helmholtz (mWC)',
                    f'{sample}  [{mat_label}] — Helmholtz')

        # --- Teslameter panel ---
        tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
        any_tesla = False
        for face, (marker, fcolor, flabel) in FACE_STYLE.items():
            fpath = os.path.join(tesla_dir, f'{sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            dates, mags, temps = load_teslameter_face(fpath, tunnel_only=True)
            if dates:
                any_tesla = True
                ax_tesla.plot(dates, mags, f'{marker}-', color=fcolor,
                              markersize=5, linewidth=1, label=f'|B| {flabel}')

        # Temperature on twin axis
        if any_tesla:
            ax_temp = ax_tesla.twinx()
            for face, (marker, fcolor, flabel) in FACE_STYLE.items():
                fpath = os.path.join(tesla_dir, f'{sample}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                dates, mags, temps = load_teslameter_face(fpath, tunnel_only=True)
                if dates and not all(np.isnan(t) for t in temps):
                    ax_temp.plot(dates, temps, f'{marker}:', color=fcolor,
                                 markersize=3, alpha=0.4, linewidth=0.8)
            ax_temp.set_ylabel('Temp (°C)', fontsize=8, color='gray')
            ax_temp.tick_params(axis='y', labelsize=7, colors='gray')

        format_axes(ax_tesla, '|B| (mT)',
                    f'{sample}  [{mat_label}] — Teslameter',
                    show_legend=any_tesla)

    fig.suptitle(f'Y-Plate {plate_num}', fontsize=14, fontweight='bold', y=0.995)
    fig.tight_layout(rect=[0, 0, 1, 0.98])
    outpath = os.path.join(PLOT_DIR, 'Y_Plates',
                           f'Y-{plate_num:02d}_plate.png')
    fig.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return has_data


# ─── H-plate (pair assembly) figures ─────────────────────────────────────────

def plot_h_plate(plate_num, ns, pair_arrangements, temp_lookup):
    """
    Generate per-plate figure for H-plates (Hn or Hs).

    For each of the 4 pair assembly slots:
      - Left panel: Hn/Hs (pair) Helmholtz + An/As-1 + An/As-2 (individual)
                    raw (dashed gray) and corrected (solid with error bars)
      - Right panel: Teslameter field magnitude per face for both A magnets

    ns: 'n' (NdFeB) or 's' (SmCo)
    """
    h_prefix = f'H{ns}'
    a_prefix = f'A{ns}'
    plate_key = f'{ns}-{plate_num}'
    mat_type = 'NdFeB' if ns == 'n' else 'SmCo'
    alpha = ALPHA[mat_type]
    color = CB[mat_type]

    # Get pair configurations if available
    configs = ['', '', '', '']
    if plate_key in pair_arrangements:
        _, configs = pair_arrangements[plate_key]

    fig, axes = plt.subplots(4, 2, figsize=(16, 18))
    has_data = False

    for slot_idx in range(4):
        slot = slot_idx + 1
        h_sample = f'{h_prefix}-{plate_num}-{slot}'
        config = configs[slot_idx] if slot_idx < len(configs) else ''
        config_label = f'  ({config})' if config else ''

        ax_helm = axes[slot_idx, 0]
        ax_tesla = axes[slot_idx, 1]

        # --- Helmholtz panel: H-sample (pair) + two A-samples (individual) ---
        helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')

        # H-sample (pair integrated reading)
        h_helm_path = os.path.join(helm_dir, f'{h_sample}_helmholtz.dat')
        if os.path.exists(h_helm_path):
            dr, vr, dc, vc, ec = load_helmholtz_sample(
                h_helm_path, h_sample, alpha, temp_lookup)
            if dr:
                has_data = True
                ax_helm.plot(dr, vr, 'o--', color='gray', alpha=0.4,
                             markersize=4, label=f'{h_sample} raw')
            if dc:
                ax_helm.errorbar(dc, vc, yerr=ec, fmt='o-', color=color,
                                 markersize=5, capsize=3, linewidth=1.2,
                                 label=f'{h_sample} corr')

        # A-samples (individual magnets in the pair)
        a_colors = ['#0077BB', '#CC3311']  # blue, red for magnet 1, 2
        for mag_idx, ac in zip(['1', '2'], a_colors):
            a_sample = f'{a_prefix}-{plate_num}-{slot}-{mag_idx}'
            a_helm_path = os.path.join(helm_dir, f'{a_sample}_helmholtz.dat')
            if not os.path.exists(a_helm_path):
                continue
            dr, vr, dc, vc, ec = load_helmholtz_sample(
                a_helm_path, a_sample, alpha, temp_lookup)
            if dr:
                has_data = True
                ax_helm.plot(dr, vr, 's--', color=ac, alpha=0.3,
                             markersize=3, label=f'{a_sample} raw')
            if dc:
                ax_helm.errorbar(dc, vc, yerr=ec, fmt='s-', color=ac,
                                 markersize=4, capsize=2, linewidth=0.9,
                                 label=f'{a_sample} corr')

        format_axes(ax_helm, 'Helmholtz (mWC)',
                    f'{h_sample}  [{mat_type}{config_label}] — Helmholtz')

        # --- Teslameter panel: H-sample (pair) + both A magnets × 3 faces ---
        tesla_dir = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')
        any_tesla = False
        all_temp_dates = []
        all_temp_vals = []

        # H-sample pair-level Teslameter (Hn-XX-YY or Hs-XX-YY)
        for face, (marker, fcolor, flabel) in FACE_STYLE.items():
            fpath = os.path.join(tesla_dir, f'{h_sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            dates, mags, temps = load_teslameter_face(fpath,
                                                       tunnel_only=True)
            if dates:
                any_tesla = True
                ax_tesla.plot(dates, mags, marker=marker,
                              linestyle='-', color=fcolor,
                              markersize=6, linewidth=1.2,
                              label=f'{h_sample} {flabel}')
                for d, t in zip(dates, temps):
                    if not np.isnan(t):
                        all_temp_dates.append(d)
                        all_temp_vals.append(t)

        # A-sample individual magnet Teslameter (An-XX-YY-1, An-XX-YY-2)
        mag_styles = {
            '1': {'ls': '--', 'alpha': 0.6, 'ms': 4},
            '2': {'ls': ':',  'alpha': 0.5, 'ms': 3},
        }

        for mag_idx in ['1', '2']:
            a_sample = f'{a_prefix}-{plate_num}-{slot}-{mag_idx}'
            sty = mag_styles[mag_idx]
            for face, (marker, fcolor, flabel) in FACE_STYLE.items():
                fpath = os.path.join(tesla_dir, f'{a_sample}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                dates, mags, temps = load_teslameter_face(fpath,
                                                           tunnel_only=True)
                if dates:
                    any_tesla = True
                    ax_tesla.plot(dates, mags, marker=marker,
                                 linestyle=sty['ls'], color=fcolor,
                                 markersize=sty['ms'], linewidth=0.8,
                                 alpha=sty['alpha'],
                                 label=f'{a_sample} {flabel}')
                    for d, t in zip(dates, temps):
                        if not np.isnan(t):
                            all_temp_dates.append(d)
                            all_temp_vals.append(t)

        # Temperature on twin axis
        if all_temp_dates:
            ax_temp = ax_tesla.twinx()
            ax_temp.plot(all_temp_dates, all_temp_vals, 'x:',
                         color='gray', markersize=3, alpha=0.35,
                         linewidth=0.5, label='Temp')
            ax_temp.set_ylabel('Temp (°C)', fontsize=8, color='gray')
            ax_temp.tick_params(axis='y', labelsize=7, colors='gray')

        format_axes(ax_tesla, '|B| (mT)',
                    f'{h_sample}  [{mat_type}{config_label}] — Teslameter',
                    show_legend=any_tesla)

    fig.suptitle(f'{h_prefix}-Plate {plate_num}  ({mat_type})',
                 fontsize=14, fontweight='bold', y=0.995)
    fig.tight_layout(rect=[0, 0, 1, 0.98])

    subdir = 'Hn_Plates' if ns == 'n' else 'Hs_Plates'
    outpath = os.path.join(PLOT_DIR, subdir,
                           f'{h_prefix}-{plate_num:02d}_plate.png')
    fig.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return has_data


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Loading materials and temperature data...")
    y_materials, pair_arrangements = load_materials()
    temp_lookup = build_temperature_lookup()

    # ── Y-plates ──
    # Find all Y-plate numbers with data
    helm_dir_y = os.path.join(BASE, 'Y_Plates', 'Helmholtz')
    y_plate_nums = set()
    for f in os.listdir(helm_dir_y):
        m = re.match(r'Y-(\d+)-\d+_helmholtz\.dat$', f)
        if m:
            y_plate_nums.add(int(m.group(1)))

    print(f"\nGenerating Y-plate figures ({len(y_plate_nums)} plates)...")
    for pn in sorted(y_plate_nums):
        ok = plot_y_plate(pn, y_materials, temp_lookup)
        status = "OK" if ok else "no data"
        print(f"  Y-{pn:02d}: {status}")

    # ── H-plates (NdFeB = Hn, SmCo = Hs) ──
    helm_dir_p = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')

    for ns, label in [('n', 'NdFeB'), ('s', 'SmCo')]:
        prefix = f'H{ns}'
        plate_nums = set()
        for f in os.listdir(helm_dir_p):
            m = re.match(rf'{prefix}-(\d+)-\d+_helmholtz\.dat$', f)
            if m:
                plate_nums.add(int(m.group(1)))

        print(f"\nGenerating {prefix}-plate figures ({len(plate_nums)} plates)...")
        for pn in sorted(plate_nums):
            ok = plot_h_plate(pn, ns, pair_arrangements, temp_lookup)
            status = "OK" if ok else "no data"
            print(f"  {prefix}-{pn:02d}: {status}")

    print(f"\nAll plots saved to {PLOT_DIR}/")
    print("Done.")


if __name__ == '__main__':
    main()
