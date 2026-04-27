#!/usr/bin/env python3
"""Lab Control H-plate and A-sample Analysis

Analyzes lab-only (never irradiated) H-plates and A-samples to measure
Helmholtz drift over time. This provides the critical baseline for
interpreting tunnel radiation effects.

Data sources:
  - Baselines: Cleanup_Claude/Pair_Assemblies/Helmholtz/ (Nov 2024 - Jul 2025)
  - Upstairs 2026: Cleanup_Claude/Lab_Measurements/Upstairs_2026/ (Jan-Mar 2026)
  - Dec 2025: Cleanup_Claude/Lab_Measurements/2025-12-17_AdamR/

TEMPERATURE HANDLING:
  Lab measurements generally lack co-incident temperature data.
  For 5/8 dates, Teslameter temps are available from H/A measurements.
  For 3/8 dates, temp is estimated from nearby measured dates.
  All readings are corrected to T_ref=20°C using:
    h_corr = h_raw / (1 + alpha * (T_est - T_ref))
  Temperature coefficients: NdFeB α=-0.10%/°C, SmCo α=-0.04%/°C
  Each result carries sigma_temp_pct = temp uncertainty on % change.
  Uncertainty is larger for estimated dates (~±0.14% NdFeB) than
  teslameter dates (~±0.09% NdFeB). Raw values also stored (pct_change_raw).

Assembly configurations (from Materials_Arrangements_Spreadsheet.xlsx):
  - Beta (antiparallel): H-plate reads ~zero (unreliable), A-samples individual
  - Delta (single + slug): slug is always the 2nd A-sample (e.g., An-35-2-2)
  - Alpha (parallel): clean signals at both H and A level
  - Gamma (90°): clean signals at both H and A level

Output: Cleanup_Claude/Lab_Controls/
"""

import os
import sys
import re
import numpy as np
from collections import defaultdict
from datetime import datetime

# Paths
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLEANUP = os.path.dirname(os.path.abspath(__file__))
PAIR_HELM = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')
UPSTAIRS = os.path.join(BASE, 'Lab_Measurements', 'Upstairs_2026')
DEC17 = os.path.join(BASE, 'Lab_Measurements', '2025-12-17_AdamR')
PLOT_DIR = os.path.join(CLEANUP)
os.makedirs(PLOT_DIR, exist_ok=True)

# Import parsers from existing codebase
sys.path.insert(0, BASE)
from degradation_summary_v2 import parse_helmholtz_file

# Constants
SENTINEL = 1337.0
MIN_READING = 0.05  # mWC below this = slug or zero (Beta antiparallel)

# Tunnel plate numbers — these are EXCLUDED from lab analysis
TUNNEL_N = {6, 8, 9, 10, 11, 12, 15, 16, 17, 18, 19, 20, 37, 39}
TUNNEL_S = {1, 2, 3, 4, 5, 6, 7, 10, 12, 13, 14, 15, 16, 17, 18, 20}

# Temperature coefficients (per °C, fractional)
ALPHA = {'NdFeB': -0.0010, 'SmCo': -0.0004}  # per °C
T_REF = 20.0  # Reference temperature for all corrections (°C)

# ─── Lab temperature lookup ────────────────────────────────────────────────
# For each lab measurement date, provide (T_est, sigma_T, source).
# - 'teslameter': actual measured temp from H/A Teslameter on that date
# - 'estimated': inferred from nearby dates or seasonal pattern
#
# Measured temps (from lab H/A Teslameter readings):
#   2024-12-04: 23.08 ± 0.62°C (N=63 readings)
#   2025-02-05: 23.01 ± 0.33°C (N=8)
#   2025-02-14: 22.04 ± 0.29°C (N=3)
#   2025-08-26: 24.04 ± 0.63°C (N=36)
#   2025-08-27: 31.68 ± 0.13°C (N=33) — facility-wide hot day
#
# Estimated temps (for dates without Teslameter):
#   2025-02-06: 23.0 ± 1.0°C (day after Feb 5 = 23.01)
#   2025-02-26: 22.5 ± 1.0°C (12d after Feb 14 = 22.04)
#   2025-12-17: 23.0 ± 1.0°C (winter, cf Dec 4 = 23.08)
#
LAB_TEMP_LOOKUP = {
    # ── Baseline dates (2024-2025) ──
    '2024-12-04': (23.08, 0.62, 'teslameter'),
    '2025-02-05': (23.01, 0.33, 'teslameter'),
    '2025-02-06': (23.0,  1.0,  'estimated'),   # day after Feb 5 = 23.01
    '2025-02-14': (22.04, 0.29, 'teslameter'),
    '2025-02-26': (22.5,  1.0,  'estimated'),   # 12d after Feb 14 = 22.04
    '2025-08-26': (24.04, 0.63, 'teslameter'),
    '2025-08-27': (31.68, 0.13, 'teslameter'),  # facility-wide hot day
    # ── Latest measurement dates (2025-2026, all Helmholtz-only) ──
    '2025-12-17': (23.0,  1.0,  'estimated'),   # winter, cf Dec 4 = 23.08
    '2026-01-30': (23.0,  1.0,  'estimated'),   # winter
    '2026-02-06': (23.0,  1.0,  'estimated'),   # winter
    '2026-02-13': (23.0,  1.0,  'estimated'),   # winter
    '2026-02-20': (23.0,  1.0,  'estimated'),   # winter
    '2026-03-02': (23.0,  1.0,  'estimated'),   # winter/spring
    '2026-03-09': (23.0,  1.0,  'estimated'),   # winter/spring
    '2026-03-20': (21.0,  1.0,  'estimated'),   # user-reported 21°C on this date
}
# Normal-day mean: 23.0 ± 0.8°C (from 4 non-anomalous teslameter dates)
LAB_TEMP_DEFAULT = (22.5, 2.5, 'default')  # fallback if date not in lookup


def get_lab_temp(date_str):
    """Get estimated lab temperature for a given date.

    Returns (T_est, sigma_T, source) where source is 'teslameter', 'estimated',
    or 'default'.
    """
    return LAB_TEMP_LOOKUP.get(date_str, LAB_TEMP_DEFAULT)


def temp_correct_reading(h_raw, temp, alpha):
    """Apply temperature correction: normalize reading to T_REF.

    h_corr = h_raw / (1 + alpha * (T - T_REF))
    """
    return h_raw / (1 + alpha * (temp - T_REF))


def compute_sigma_temp_pct(alpha, sigma_T):
    """Compute % uncertainty on a SINGLE reading due to temperature uncertainty.

    For a % change between two readings, the uncertainty is sqrt(2) * this value
    (assuming independent temp errors on baseline and latest).
    """
    return abs(alpha * sigma_T) * 100.0  # in percent


def is_tunnel_plate(sample_name):
    """Check if a sample belongs to a tunnel plate."""
    m = re.match(r'[HA]([ns])-(\d+)', sample_name)
    if not m:
        return False
    ns, pnum = m.group(1), int(m.group(2))
    if ns == 'n':
        return pnum in TUNNEL_N
    else:
        return pnum in TUNNEL_S


def load_assembly_configs():
    """Load assembly configurations from spreadsheet.

    Returns dict: 'n-35' -> ['Beta', 'Delta', 'Alpha', 'Gamma']
    """
    import openpyxl
    mat_path = os.path.join(os.path.dirname(BASE), 'Materials_Arrangements_Spreadsheet.xlsx')
    if not os.path.exists(mat_path):
        # Try alternate location
        mat_path = os.path.join(os.path.dirname(BASE), 'Materials_Arrangements.xlsx')
    if not os.path.exists(mat_path):
        print("WARNING: Materials spreadsheet not found, configs unavailable")
        return {}

    configs = {}
    wb = openpyxl.load_workbook(mat_path, data_only=True)
    for sheet_name in ['Lab - Pair Arrangements', 'Tunnel - Pair Arrangements']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            plate_id = str(row[0]).strip().lower()
            # Expect format like 'n-35' or 's-22'
            m = re.match(r'([ns])-?(\d+)', plate_id)
            if not m:
                continue
            key = '%s-%s' % (m.group(1), m.group(2))
            slot_configs = []
            for i in range(2, min(6, len(row))):
                if row[i]:
                    slot_configs.append(str(row[i]).strip())
                else:
                    slot_configs.append('')
            configs[key] = slot_configs
    return configs


def parse_dat_file(filepath):
    """Parse a .dat file, returning list of (datetime, value_mWC).

    Handles both standard parse_helmholtz_file format and raw reading.
    Filters out sentinel values and non-mWC units.
    """
    results = []
    try:
        rows = parse_helmholtz_file(filepath)
        for dt, v, u in rows:
            if u == 'mWC' and abs(v - SENTINEL) > 1:
                results.append((dt, v))
    except Exception:
        # Fallback: manual parse
        with open(filepath, 'r', errors='replace') as f:
            for line in f:
                m = re.match(
                    r'\s*(\d{4}-\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2})\s+'
                    r'.*?([+-]?\d+\.?\d*)\s+mWC', line)
                if m:
                    dt = datetime.strptime(
                        '%s %s' % (m.group(1), m.group(2)),
                        '%Y-%m-%d %H:%M:%S')
                    val = float(m.group(3))
                    if abs(val - SENTINEL) > 1:
                        results.append((dt, val))
    return sorted(results, key=lambda x: x[0])


def collect_all_lab_data():
    """Collect all lab H-plate and A-sample Helmholtz data.

    Merges baselines (Pair_Assemblies), Dec 2025, and Upstairs 2026 data.
    Returns dict: sample_name -> [(datetime, value_mWC), ...]
    """
    all_data = defaultdict(list)

    # Known corrupt filenames to skip
    SKIP_FILES = {'Hs-19-1Hs-19-1_helmholtz.dat'}  # duplicate-name artifact

    # Source 1: Baselines from Pair_Assemblies/Helmholtz
    if os.path.isdir(PAIR_HELM):
        for f in os.listdir(PAIR_HELM):
            if not f.endswith('_helmholtz.dat') or f in SKIP_FILES:
                continue
            sample = f.replace('_helmholtz.dat', '')
            if not re.match(r'[HA][ns]-', sample):
                continue
            if is_tunnel_plate(sample):
                continue
            readings = parse_dat_file(os.path.join(PAIR_HELM, f))
            all_data[sample].extend(readings)

    # Source 2: Dec 2025 lab measurements
    if os.path.isdir(DEC17):
        for f in os.listdir(DEC17):
            if not f.endswith('_helmholtz.dat') or f in SKIP_FILES:
                continue
            sample = f.replace('_helmholtz.dat', '')
            if not re.match(r'[HA][ns]-', sample):
                continue
            if is_tunnel_plate(sample):
                continue
            readings = parse_dat_file(os.path.join(DEC17, f))
            all_data[sample].extend(readings)

    # Source 3: Upstairs 2026
    if os.path.isdir(UPSTAIRS):
        for f in os.listdir(UPSTAIRS):
            if not f.endswith('_helmholtz.dat') or f in SKIP_FILES:
                continue
            sample = f.replace('_helmholtz.dat', '')
            if not re.match(r'[HA][ns]-', sample):
                continue
            if is_tunnel_plate(sample):
                continue
            readings = parse_dat_file(os.path.join(UPSTAIRS, f))
            all_data[sample].extend(readings)

    # Deduplicate by datetime (same measurement from multiple sources)
    deduped = {}
    for sample, readings in all_data.items():
        seen = {}
        for dt, val in readings:
            dt_key = dt.strftime('%Y-%m-%d %H:%M')
            if dt_key not in seen:
                seen[dt_key] = (dt, val)
        deduped[sample] = sorted(seen.values(), key=lambda x: x[0])

    return deduped


def classify_sample(sample_name):
    """Classify a sample.

    Returns dict with: type (H or A), material (NdFeB or SmCo),
    plate_num, slot, magnet_idx (A only), is_slug (A only)
    """
    info = {}

    m_h = re.match(r'H([ns])-(\d+)-(\d+)', sample_name)
    m_a = re.match(r'A([ns])-(\d+)-(\d+)-(\d+)', sample_name)

    if m_h:
        info['type'] = 'H'
        info['ns'] = m_h.group(1)
        info['material'] = 'NdFeB' if m_h.group(1) == 'n' else 'SmCo'
        info['plate_num'] = int(m_h.group(2))
        info['slot'] = int(m_h.group(3))
    elif m_a:
        info['type'] = 'A'
        info['ns'] = m_a.group(1)
        info['material'] = 'NdFeB' if m_a.group(1) == 'n' else 'SmCo'
        info['plate_num'] = int(m_a.group(2))
        info['slot'] = int(m_a.group(3))
        info['magnet_idx'] = int(m_a.group(4))
    else:
        return None

    return info


def analyze_lab_samples(all_data, configs, apply_temp_correction=True):
    """Compute percent changes for all lab samples.

    Uses earliest reading as baseline, latest as final.
    When apply_temp_correction=True (default), corrects each reading to T_REF
    using the lab temperature lookup, and propagates temperature uncertainty
    into per-measurement error bars.

    Returns list of result dicts. Each dict includes:
      - pct_change: % change (temp-corrected if enabled)
      - pct_change_raw: % change WITHOUT temp correction (always present)
      - temp_corrected: True if correction was applied
      - sigma_temp_pct: temperature-induced uncertainty on pct_change (%)
      - temp_source_baseline: how baseline temp was determined
      - temp_source_latest: how latest temp was determined
    """
    results = []

    for sample, readings in sorted(all_data.items()):
        if len(readings) < 2:
            continue

        info = classify_sample(sample)
        if not info:
            continue

        # Get config for this slot
        config_key = '%s-%d' % (info['ns'], info['plate_num'])
        slot_configs = configs.get(config_key, [])
        slot_idx = info['slot'] - 1
        config = slot_configs[slot_idx] if slot_idx < len(slot_configs) else ''

        baseline_val = readings[0][1]
        baseline_dt = readings[0][0]
        latest_val = readings[-1][1]
        latest_dt = readings[-1][0]

        # Skip near-zero readings (Beta H-plates, Delta slugs)
        if abs(baseline_val) < MIN_READING:
            continue

        # Detect slugs (A-samples with ~zero)
        # Delta slug is always the 2nd A-sample in the pair
        is_slug = (info['type'] == 'A' and
                   info.get('magnet_idx') == 2 and
                   config.lower().startswith('delta') and
                   abs(latest_val) < MIN_READING)
        if is_slug:
            continue

        # Robust baseline selection: find the consensus value
        # If a reading is >10% away from the median of all readings,
        # it's likely a measurement error. Use the earliest non-anomalous
        # reading as baseline.
        is_anomalous_baseline = False
        all_vals = [v for _, v in readings]
        median_val = np.median(all_vals)

        # Select baseline: earliest reading within 20% of median
        baseline_val = None
        baseline_dt = None
        for dt, val in readings:
            if abs(val - median_val) / abs(median_val) < 0.20:
                baseline_val = val
                baseline_dt = dt
                break

        if baseline_val is None:
            continue

        if baseline_dt != readings[0][0]:
            is_anomalous_baseline = True

        # Similarly, use latest non-anomalous reading
        latest_val = None
        latest_dt = None
        for dt, val in reversed(readings):
            if abs(val - median_val) / abs(median_val) < 0.20:
                latest_val = val
                latest_dt = dt
                break

        if latest_val is None or latest_dt == baseline_dt:
            continue

        # RAW pct_change (always computed)
        pct_change_raw = (latest_val - baseline_val) / baseline_val * 100.0

        # Temperature correction
        mat_class = info['material']  # 'NdFeB' or 'SmCo'
        alpha = ALPHA[mat_class]

        if apply_temp_correction:
            bl_date_str = baseline_dt.strftime('%Y-%m-%d')
            lt_date_str = latest_dt.strftime('%Y-%m-%d')
            bl_temp, bl_sigma_T, bl_source = get_lab_temp(bl_date_str)
            lt_temp, lt_sigma_T, lt_source = get_lab_temp(lt_date_str)

            bl_corr = temp_correct_reading(baseline_val, bl_temp, alpha)
            lt_corr = temp_correct_reading(latest_val, lt_temp, alpha)
            pct_change = (lt_corr - bl_corr) / bl_corr * 100.0

            # Temperature uncertainty on % change:
            # sigma_pct = |alpha| * sqrt(sigma_bl^2 + sigma_lt^2) * 100
            sigma_temp_pct = abs(alpha) * np.sqrt(
                bl_sigma_T**2 + lt_sigma_T**2) * 100.0
            temp_corrected = True
        else:
            pct_change = pct_change_raw
            sigma_temp_pct = 0.0
            bl_source = 'none'
            lt_source = 'none'
            bl_corr = baseline_val
            lt_corr = latest_val
            temp_corrected = False

        # Time span in days
        span_days = (latest_dt - baseline_dt).days

        # Build time series (corrected if enabled)
        date_pcts = []
        for dt, val in readings:
            if dt >= baseline_dt:
                if apply_temp_correction:
                    rd_str = dt.strftime('%Y-%m-%d')
                    rd_temp, _, _ = get_lab_temp(rd_str)
                    val_c = temp_correct_reading(val, rd_temp, alpha)
                    pct = (val_c - bl_corr) / bl_corr * 100.0
                else:
                    pct = (val - baseline_val) / baseline_val * 100.0
                date_pcts.append((dt, pct))

        results.append({
            'sample': sample,
            'type': info['type'],
            'material': info['material'],
            'plate_num': info['plate_num'],
            'slot': info['slot'],
            'magnet_idx': info.get('magnet_idx'),
            'config': config,
            'baseline_val': baseline_val,
            'baseline_dt': baseline_dt,
            'latest_val': latest_val,
            'latest_dt': latest_dt,
            'pct_change': pct_change,
            'pct_change_raw': pct_change_raw,
            'span_days': span_days,
            'n_readings': len(readings),
            'date_pcts': date_pcts,
            'is_anomalous_baseline': is_anomalous_baseline,
            'is_delta_hplate': (info['type'] == 'H' and
                                config.lower().startswith('delta')),
            'temp_corrected': temp_corrected,
            'sigma_temp_pct': sigma_temp_pct,
            'temp_source_baseline': bl_source if apply_temp_correction else 'none',
            'temp_source_latest': lt_source if apply_temp_correction else 'none',
        })

    return results


def print_summary(results):
    """Print comprehensive summary statistics."""

    print("=" * 80)
    print("LAB CONTROL H-PLATE AND A-SAMPLE ANALYSIS")
    print("=" * 80)
    print()
    n_corr = sum(1 for r in results if r.get('temp_corrected', False))
    n_raw = len(results) - n_corr
    if n_corr > 0:
        print("TEMPERATURE CORRECTION APPLIED: %d/%d samples corrected to T_ref=%.0f°C." %
              (n_corr, len(results), T_REF))
        print("Sources: LAB_TEMP_LOOKUP (teslameter where available, estimated otherwise).")
        # Show temp uncertainty range
        sigmas = [r.get('sigma_temp_pct', 0) for r in results if r.get('temp_corrected')]
        if sigmas:
            print("Temp uncertainty on %% change: %.3f%% – %.3f%% (per sample)." %
                  (min(sigmas), max(sigmas)))
    else:
        print("IMPORTANT: All values are RAW (no temperature correction).")
        print("Lab temp estimated 20-25°C. A 1°C error biases NdFeB by ~0.10%,")
        print("SmCo by ~0.04%, and the NdFeB-SmCo differential by ~0.06%.")
    print()

    # Separate H and A
    h_results = [r for r in results if r['type'] == 'H']
    a_results = [r for r in results if r['type'] == 'A']

    for label, subset in [('H-PLATE (assembled)', h_results),
                          ('A-SAMPLE (individual)', a_results)]:
        print("-" * 70)
        print(f"{label}: {len(subset)} samples")
        print("-" * 70)

        if not subset:
            print("  No data\n")
            continue

        # By material (exclude anomalous baselines and Delta H-plates)
        for mat in ['NdFeB', 'SmCo']:
            mat_data = [r for r in subset if r['material'] == mat
                        and not r['is_anomalous_baseline']
                        and not r.get('is_delta_hplate', False)]
            if not mat_data:
                continue
            pcts = [r['pct_change'] for r in mat_data]
            mean_pct = np.mean(pcts)
            std_pct = np.std(pcts, ddof=1)
            sem = std_pct / np.sqrt(len(pcts))
            med_pct = np.median(pcts)
            span_range = (min(r['span_days'] for r in mat_data),
                          max(r['span_days'] for r in mat_data))

            print(f"  {mat} (N={len(pcts)}):")
            print(f"    Mean:   {mean_pct:+.3f}% +/- {sem:.3f}% (SEM)")
            print(f"    StdDev: {std_pct:.3f}%")
            print(f"    Median: {med_pct:+.3f}%")
            print(f"    Range:  [{min(pcts):+.3f}%, {max(pcts):+.3f}%]")
            print(f"    Span:   {span_range[0]}-{span_range[1]} days")
            n_pos = sum(1 for p in pcts if p > 0)
            n_neg = sum(1 for p in pcts if p < 0)
            print(f"    Sign:   {n_neg} negative, {n_pos} positive")
            print()

        # NdFeB - SmCo differential
        # Exclude: Beta (unreliable), anomalous baselines, Delta H-plates
        nd_data = [r for r in subset if r['material'] == 'NdFeB'
                   and not r['is_anomalous_baseline']
                   and not r.get('is_delta_hplate', False)
                   and r['config'].lower() not in ('beta', '')]
        sm_data = [r for r in subset if r['material'] == 'SmCo'
                   and not r['is_anomalous_baseline']
                   and not r.get('is_delta_hplate', False)
                   and r['config'].lower() not in ('beta', '')]

        if nd_data and sm_data:
            nd_mean = np.mean([r['pct_change'] for r in nd_data])
            sm_mean = np.mean([r['pct_change'] for r in sm_data])
            nd_sem = (np.std([r['pct_change'] for r in nd_data], ddof=1)
                      / np.sqrt(len(nd_data)))
            sm_sem = (np.std([r['pct_change'] for r in sm_data], ddof=1)
                      / np.sqrt(len(sm_data)))
            diff = nd_mean - sm_mean
            diff_err = np.sqrt(nd_sem**2 + sm_sem**2)
            sigma = abs(diff / diff_err) if diff_err > 0 else 0

            print(f"  NdFeB-SmCo differential (excl. Beta, anomalous):")
            print(f"    NdFeB mean: {nd_mean:+.3f}% +/- {nd_sem:.3f}%"
                  f" (N={len(nd_data)})")
            print(f"    SmCo mean:  {sm_mean:+.3f}% +/- {sm_sem:.3f}%"
                  f" (N={len(sm_data)})")
            print(f"    Differential: {diff:+.3f}% +/- {diff_err:.3f}%"
                  f" ({sigma:.1f} sigma)")
            print()

        # By config
        print(f"  By assembly config:")
        for cfg in ['Alpha', 'Beta', 'Delta', 'Gamma']:
            cfg_data = [r for r in subset
                        if r['config'].lower() == cfg.lower()
                        and not r['is_anomalous_baseline']]
            if not cfg_data:
                continue
            pcts = [r['pct_change'] for r in cfg_data]
            mean = np.mean(pcts)
            sem = np.std(pcts, ddof=1) / np.sqrt(len(pcts)) if len(pcts) > 1 else 0
            print(f"    {cfg:6s} (N={len(cfg_data):3d}): "
                  f"{mean:+.3f}% +/- {sem:.3f}%")
        print()

    # Anomalous baselines
    anomalous = [r for r in results if r['is_anomalous_baseline']]
    if anomalous:
        print("-" * 70)
        print(f"ANOMALOUS BASELINES DETECTED: {len(anomalous)} samples")
        print("(First reading >10% from second; using second as baseline)")
        print("-" * 70)
        for r in anomalous:
            print(f"  {r['sample']}: config={r['config']}, "
                  f"baseline={r['baseline_val']:.4f} mWC")
        print()


def print_per_plate_summary(results, configs):
    """Print per-plate matched NdFeB-SmCo differential."""

    print("=" * 80)
    print("PER-PLATE NdFeB-SmCo MATCHED DIFFERENTIAL (lab controls)")
    print("=" * 80)
    print()
    print("For each NdFeB plate, find the corresponding SmCo plate measured")
    print("on the same dates. This is the closest analog to the Y-plate")
    print("intra-plate differential (which was 9.7 sigma in the tunnel).")
    print()
    print("NOTE: H-plates are SINGLE-MATERIAL, so NdFeB-SmCo pairing requires")
    print("matching across separate plates, unlike Y-plates which have all 4")
    print("materials on one plate. This introduces cross-plate variability.")
    print()

    # Group A-samples by plate and material
    a_results = [r for r in results if r['type'] == 'A'
                 and not r['is_anomalous_baseline']
                 and r['config'].lower() not in ('beta', '')]

    # For A-samples, compute per-plate means
    plate_means = {}  # (material, plate_num) -> mean pct
    for r in a_results:
        key = (r['material'], r['plate_num'])
        if key not in plate_means:
            plate_means[key] = []
        plate_means[key].append(r['pct_change'])

    for key in plate_means:
        plate_means[key] = np.mean(plate_means[key])

    # List NdFeB and SmCo plate means
    nd_plates = {pn: pct for (mat, pn), pct in plate_means.items()
                 if mat == 'NdFeB'}
    sm_plates = {pn: pct for (mat, pn), pct in plate_means.items()
                 if mat == 'SmCo'}

    print(f"NdFeB plates with A-sample data: {len(nd_plates)}")
    print(f"SmCo plates with A-sample data: {len(sm_plates)}")
    print()

    if nd_plates and sm_plates:
        nd_vals = list(nd_plates.values())
        sm_vals = list(sm_plates.values())
        nd_mean = np.mean(nd_vals)
        sm_mean = np.mean(sm_vals)
        nd_sem = np.std(nd_vals, ddof=1) / np.sqrt(len(nd_vals))
        sm_sem = np.std(sm_vals, ddof=1) / np.sqrt(len(sm_vals))
        diff = nd_mean - sm_mean
        diff_err = np.sqrt(nd_sem**2 + sm_sem**2)
        sigma = abs(diff / diff_err) if diff_err > 0 else 0

        print(f"NdFeB plate mean: {nd_mean:+.3f}% +/- {nd_sem:.3f}% "
              f"(N={len(nd_plates)} plates)")
        print(f"SmCo plate mean:  {sm_mean:+.3f}% +/- {sm_sem:.3f}% "
              f"(N={len(sm_plates)} plates)")
        print(f"Differential:     {diff:+.3f}% +/- {diff_err:.3f}% "
              f"({sigma:.1f} sigma)")
        print()
        print("For comparison:")
        print("  Tunnel Y-plate NdFeB-SmCo: -0.266% +/- 0.027% (9.7 sigma)")
        print("  Lab Y-plate NdFeB-SmCo:    -0.006% +/- 0.019% (0.3 sigma)")
        print()

    # Individual plate table
    print(f"{'Plate':>8s} {'Material':>8s} {'N_samples':>9s} "
          f"{'Mean%':>8s} {'StdDev':>8s}")
    print("-" * 50)
    for mat_label, plates in [('NdFeB', nd_plates), ('SmCo', sm_plates)]:
        for pn in sorted(plates.keys()):
            key = (mat_label, pn)
            pcts = [r['pct_change'] for r in a_results
                    if r['material'] == mat_label and r['plate_num'] == pn]
            std = np.std(pcts, ddof=1) if len(pcts) > 1 else 0
            print(f"  {mat_label[0]}-{pn:>2d}   {mat_label:>8s} {len(pcts):>9d} "
                  f"{np.mean(pcts):>+8.3f} {std:>8.3f}")


def main():
    print("Loading assembly configurations...")
    configs = load_assembly_configs()
    print(f"  Found configs for {len(configs)} plates")

    print("\nCollecting all lab H/A data...")
    all_data = collect_all_lab_data()
    n_h = sum(1 for s in all_data if s.startswith('H'))
    n_a = sum(1 for s in all_data if s.startswith('A'))
    print(f"  {n_h} H-plate samples, {n_a} A-samples")
    print(f"  Dates range: ", end='')
    all_dates = set()
    for readings in all_data.values():
        for dt, _ in readings:
            all_dates.add(dt.strftime('%Y-%m-%d'))
    print(', '.join(sorted(all_dates)))

    # Count samples with >=2 readings
    multi = sum(1 for readings in all_data.values() if len(readings) >= 2)
    single = sum(1 for readings in all_data.values() if len(readings) == 1)
    print(f"  {multi} with >=2 readings (can compute % change)")
    print(f"  {single} with only 1 reading (baseline only)")

    print("\nAnalyzing lab samples (with temp correction to T_ref=%.0f°C)..." % T_REF)
    results = analyze_lab_samples(all_data, configs, apply_temp_correction=True)
    n_corr = sum(1 for r in results if r.get('temp_corrected'))
    print(f"  {len(results)} samples with valid % change ({n_corr} temp-corrected)")

    print()
    print_summary(results)
    print()
    print_per_plate_summary(results, configs)


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    main()
