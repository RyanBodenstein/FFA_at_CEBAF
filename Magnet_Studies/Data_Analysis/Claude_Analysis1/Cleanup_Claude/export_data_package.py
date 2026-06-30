#!/usr/bin/env python3
"""
Export Data Package — Generates all CSVs for the reproducibility package.

Imports from the existing analysis scripts (v3, v2, v5_polish) and exports
clean, well-documented CSVs to ../Data_Package/.

Run from the Cleanup_Claude/ directory:
    python3 export_data_package.py

Output: ../Data_Package/01_Sample_Configuration/*.csv
        ../Data_Package/02_Magnetic_Measurements/*.csv
        ../Data_Package/04_Results/headline_results.csv
"""

import os
import sys
import re
import csv
import numpy as np
import openpyxl
from datetime import datetime
from collections import defaultdict

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
PACKAGE_DIR = os.path.join(os.path.dirname(BASE), 'Data_Package')

if BASE not in sys.path:
    sys.path.insert(0, BASE)

# ─── Imports from existing scripts ───────────────────────────────────────────
from manager_summary_v3 import (
    load_all, get_gain_syst, compute_intra_plate_diffs,
    PLACEMENTS as PLACEMENTS_SIMPLE,
    Y_BASELINE_TEMP_LOOKUP, Y_BASELINE_TEMP_DEFAULT,
    ALPHA, MAT_BY_SLOT, FLAGGED, T_REF, SENTINEL,
)
from degradation_summary_v2 import (
    load_materials, build_temperature_lookup,
    compute_h_plate_degradation,
    PLACEMENTS as PLACEMENTS_FULL,
    H_PLACEMENT, Y_PLACEMENT,
)
from manager_summary_v5_polish import (
    load_a_sample_helmholtz, load_lab_y_plates,
    load_lab_y_materials, LAB_Y_PLATES,
)


def write_csv(filepath, rows, fieldnames):
    """Write a list of dicts to CSV."""
    with open(filepath, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Wrote {filepath} ({len(rows)} rows)")


# ═══════════════════════════════════════════════════════════════════════════════
# 01_Sample_Configuration
# ═══════════════════════════════════════════════════════════════════════════════

def export_materials_arrangements():
    """Parse Materials_Arrangements.xlsx → materials_arrangements.csv"""
    wb = openpyxl.load_workbook(os.path.join(BASE, 'Materials_Arrangements.xlsx'),
                                data_only=True)
    rows = []

    # Tunnel Y-plates
    for row in wb['Tunnel - Y Materials'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pm = re.match(r'[yY]-?(\d+)', str(row[0]).strip())
        if not pm:
            continue
        plate = int(pm.group(1))
        for slot in range(1, 5):
            mat = str(row[slot]).strip() if row[slot] else ''
            if mat:
                rows.append({
                    'sample_id': 'Y-%d-%d' % (plate, slot),
                    'plate_number': plate,
                    'slot_number': slot,
                    'material': mat,
                    'sample_type': 'Y',
                    'environment': 'tunnel',
                })

    # Lab Y-plates
    for row in wb['Lab - Y Materials'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pm = re.match(r'[yY]-?(\d+)', str(row[0]).strip())
        if not pm:
            continue
        plate = int(pm.group(1))
        for slot in range(1, 5):
            mat = str(row[slot]).strip() if row[slot] else ''
            if mat:
                rows.append({
                    'sample_id': 'Y-%d-%d' % (plate, slot),
                    'plate_number': plate,
                    'slot_number': slot,
                    'material': mat,
                    'sample_type': 'Y',
                    'environment': 'lab',
                })

    # Tunnel pair arrangements (H-plates and A-samples)
    for row in wb['Tunnel - Pair Arrangements'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        plate_id = str(row[0]).strip().lower()
        pm = re.match(r'([ns])-(\d+)', plate_id)
        if not pm:
            continue
        ns = pm.group(1)
        plate = int(pm.group(2))
        mat_type = str(row[1]).strip() if row[1] else ''
        prefix = 'Hn' if ns == 'n' else 'Hs'
        a_prefix = 'An' if ns == 'n' else 'As'
        for slot in range(1, 5):
            config = str(row[slot + 1]).strip() if row[slot + 1] else ''
            rows.append({
                'sample_id': '%s-%d-%d' % (prefix, plate, slot),
                'plate_number': plate,
                'slot_number': slot,
                'material': mat_type,
                'sample_type': 'H',
                'environment': 'tunnel',
            })
            # A-samples (2 per slot)
            for pair in [1, 2]:
                rows.append({
                    'sample_id': '%s-%d-%d-%d' % (a_prefix, plate, slot, pair),
                    'plate_number': plate,
                    'slot_number': slot,
                    'material': mat_type,
                    'sample_type': 'A',
                    'environment': 'tunnel',
                })

    # Lab pair arrangements
    for row in wb['Lab - Pair Arrangements'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        plate_id = str(row[0]).strip().lower()
        pm = re.match(r'([ns])-(\d+)', plate_id)
        if not pm:
            continue
        ns = pm.group(1)
        plate = int(pm.group(2))
        mat_type = str(row[1]).strip() if row[1] else ''
        prefix = 'Hn' if ns == 'n' else 'Hs'
        a_prefix = 'An' if ns == 'n' else 'As'
        for slot in range(1, 5):
            config = str(row[slot + 1]).strip() if row[slot + 1] else ''
            rows.append({
                'sample_id': '%s-%d-%d' % (prefix, plate, slot),
                'plate_number': plate,
                'slot_number': slot,
                'material': mat_type,
                'sample_type': 'H',
                'environment': 'lab',
            })
            for pair in [1, 2]:
                rows.append({
                    'sample_id': '%s-%d-%d-%d' % (a_prefix, plate, slot, pair),
                    'plate_number': plate,
                    'slot_number': slot,
                    'material': mat_type,
                    'sample_type': 'A',
                    'environment': 'lab',
                })

    fields = ['sample_id', 'plate_number', 'slot_number', 'material',
              'sample_type', 'environment']
    outpath = os.path.join(PACKAGE_DIR, '01_Sample_Configuration',
                           'materials_arrangements.csv')
    write_csv(outpath, rows, fields)


def export_tunnel_placements():
    """Export tunnel placement table from PLACEMENTS_FULL."""
    rows = []
    for y_plate, h_plate, region, sub_location, line in PLACEMENTS_FULL:
        yn = int(y_plate.replace('Y', ''))
        rows.append({
            'plate_number': yn,
            'y_plate': y_plate,
            'h_plate': h_plate,
            'region': region,
            'sub_location': sub_location,
            'line_position': line,
        })
    fields = ['plate_number', 'y_plate', 'h_plate', 'region',
              'sub_location', 'line_position']
    outpath = os.path.join(PACKAGE_DIR, '01_Sample_Configuration',
                           'tunnel_placements.csv')
    write_csv(outpath, rows, fields)


def export_sample_inventory():
    """Build comprehensive sample inventory CSV."""
    wb = openpyxl.load_workbook(os.path.join(BASE, 'Materials_Arrangements.xlsx'),
                                data_only=True)
    rows = []

    # Build region lookup for tunnel plates
    region_lookup = {}
    config_lookup = {}
    for y_plate, h_plate, region, sub_loc, line in PLACEMENTS_FULL:
        yn = int(y_plate.replace('Y', ''))
        region_lookup[yn] = region
        # Parse h_plate to get plate number for H/A
        hm = re.match(r'([NS])(\d+)', h_plate)
        if hm:
            ns_upper = hm.group(1)
            h_num = int(hm.group(2))
            region_lookup[('H', ns_upper, h_num)] = region

    # Tunnel pair configs
    for row in wb['Tunnel - Pair Arrangements'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        plate_id = str(row[0]).strip().lower()
        pm = re.match(r'([ns])-(\d+)', plate_id)
        if not pm:
            continue
        ns = pm.group(1)
        plate = int(pm.group(2))
        for slot in range(1, 5):
            config = str(row[slot + 1]).strip() if row[slot + 1] else ''
            config_lookup[('H', ns, plate, slot)] = config

    # Lab pair configs
    for row in wb['Lab - Pair Arrangements'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        plate_id = str(row[0]).strip().lower()
        pm = re.match(r'([ns])-(\d+)', plate_id)
        if not pm:
            continue
        ns = pm.group(1)
        plate = int(pm.group(2))
        for slot in range(1, 5):
            config = str(row[slot + 1]).strip() if row[slot + 1] else ''
            config_lookup[('H', ns, plate, slot)] = config

    # Tunnel Y-plates
    for row in wb['Tunnel - Y Materials'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pm = re.match(r'[yY]-?(\d+)', str(row[0]).strip())
        if not pm:
            continue
        plate = int(pm.group(1))
        region = region_lookup.get(plate, '')
        for slot in range(1, 5):
            mat = str(row[slot]).strip() if row[slot] else ''
            if mat:
                rows.append({
                    'sample_id': 'Y-%d-%d' % (plate, slot),
                    'sample_type': 'Y',
                    'material': mat,
                    'plate': plate,
                    'slot': slot,
                    'environment': 'tunnel',
                    'region': region,
                    'config': '',
                })

    # Lab Y-plates
    for row in wb['Lab - Y Materials'].iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        pm = re.match(r'[yY]-?(\d+)', str(row[0]).strip())
        if not pm:
            continue
        plate = int(pm.group(1))
        for slot in range(1, 5):
            mat = str(row[slot]).strip() if row[slot] else ''
            if mat:
                rows.append({
                    'sample_id': 'Y-%d-%d' % (plate, slot),
                    'sample_type': 'Y',
                    'material': mat,
                    'plate': plate,
                    'slot': slot,
                    'environment': 'lab',
                    'region': '',
                    'config': '',
                })

    # H-plates (tunnel and lab via pair arrangements sheets)
    for sheet_name, env in [('Tunnel - Pair Arrangements', 'tunnel'),
                            ('Lab - Pair Arrangements', 'lab')]:
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            plate_id = str(row[0]).strip().lower()
            pm = re.match(r'([ns])-(\d+)', plate_id)
            if not pm:
                continue
            ns = pm.group(1)
            plate = int(pm.group(2))
            mat_type = str(row[1]).strip() if row[1] else ''
            prefix = 'Hn' if ns == 'n' else 'Hs'
            a_prefix = 'An' if ns == 'n' else 'As'
            ns_upper = ns.upper()
            region = region_lookup.get(('H', ns_upper, plate), '') if env == 'tunnel' else ''

            for slot in range(1, 5):
                config = str(row[slot + 1]).strip() if row[slot + 1] else ''
                h_id = '%s-%d-%d' % (prefix, plate, slot)
                rows.append({
                    'sample_id': h_id,
                    'sample_type': 'H',
                    'material': mat_type,
                    'plate': plate,
                    'slot': slot,
                    'environment': env,
                    'region': region,
                    'config': config,
                })
                for pair in [1, 2]:
                    rows.append({
                        'sample_id': '%s-%d-%d-%d' % (a_prefix, plate, slot, pair),
                        'sample_type': 'A',
                        'material': mat_type,
                        'plate': plate,
                        'slot': slot,
                        'environment': env,
                        'region': region,
                        'config': config,
                    })

    fields = ['sample_id', 'sample_type', 'material', 'plate', 'slot',
              'environment', 'region', 'config']
    outpath = os.path.join(PACKAGE_DIR, '01_Sample_Configuration',
                           'sample_inventory.csv')
    write_csv(outpath, rows, fields)


# ═══════════════════════════════════════════════════════════════════════════════
# 02_Magnetic_Measurements
# ═══════════════════════════════════════════════════════════════════════════════

def export_y_plate_degradation(results, y_materials):
    """Export Y-plate degradation results to CSV.

    Includes both tunnel and lab plates.
    """
    rows = []
    for r in results:
        # Determine line position from PLACEMENTS_FULL
        line_pos = 0
        for y_p, h_p, reg, sub, ln in PLACEMENTS_FULL:
            if int(y_p.replace('Y', '')) == r['plate']:
                line_pos = ln
                break

        rows.append({
            'plate': r['plate'],
            'slot': r['slot'],
            'sample_id': r['sample'],
            'material': r['material'],
            'region': r['region'],
            'line_position': line_pos,
            'environment': 'tunnel',
            'baseline_mean_mWC': round(r['bl_mean'], 4),
            'latest_mWC': round(r['bl_mean'] * (1 + r['pct_change'] / 100.0), 4),
            'pct_change': round(r['pct_change'], 4),
            'baseline_sem_pct': round(r['bl_sem_pct'], 4),
            'n_baseline': r['n_baseline'],
            'n_baseline_sessions': r['n_baseline_sessions'],
            'is_outlier': r['is_outlier'],
        })

    # Add lab Y-plates
    try:
        lab_results = load_lab_y_plates(apply_temp_correction=True)
        lab_mats = load_lab_y_materials()
        for plate_num, lr in lab_results.items():
            for mat, pct in lr['slot_pcts'].items():
                # Find slot number for this material
                slot = None
                for s in [1, 2, 3, 4]:
                    sid = 'Y-%d-%d' % (plate_num, s)
                    if lab_mats.get(sid) == mat:
                        slot = s
                        break
                if slot is None:
                    continue
                sid = 'Y-%d-%d' % (plate_num, slot)
                mwc_data = lr.get('slot_mwc', {}).get(mat, {})
                bl_mwc = mwc_data.get('baseline_mwc', 0.0)
                lt_mwc = mwc_data.get('latest_mwc', 0.0)
                rows.append({
                    'plate': plate_num,
                    'slot': slot,
                    'sample_id': sid,
                    'material': mat,
                    'region': '',
                    'line_position': 0,
                    'environment': 'lab',
                    'baseline_mean_mWC': round(bl_mwc, 4),
                    'latest_mWC': round(lt_mwc, 4),
                    'pct_change': round(pct, 4),
                    'baseline_sem_pct': 0.0,
                    'n_baseline': 1,
                    'n_baseline_sessions': 1,
                    'is_outlier': False,
                })
    except Exception as e:
        print(f"  Warning: Could not load lab Y-plates: {e}")

    fields = ['plate', 'slot', 'sample_id', 'material', 'region',
              'line_position', 'environment', 'baseline_mean_mWC',
              'latest_mWC', 'pct_change', 'baseline_sem_pct',
              'n_baseline', 'n_baseline_sessions', 'is_outlier']
    outpath = os.path.join(PACKAGE_DIR, '02_Magnetic_Measurements',
                           'y_plate_degradation.csv')
    write_csv(outpath, rows, fields)


def export_h_plate_degradation(h_results):
    """Export H-plate degradation results to CSV."""
    rows = []
    for r in h_results:
        rows.append({
            'pair_id': r['sample'],
            'plate': r['plate'],
            'slot': r['slot'],
            'material': r['material'],
            'region': r['region'],
            'config': r.get('config', ''),
            'baseline_mean': round(r['baseline_mean'], 4),
            'latest_mean': round(r['latest_corr'], 4),
            'pct_change': round(r['pct_change'], 4),
            'n_baseline': r['n_baseline'],
            'environment': 'tunnel',
            'is_outlier': r.get('is_outlier', False),
        })
    fields = ['pair_id', 'plate', 'slot', 'material', 'region', 'config',
              'baseline_mean', 'latest_mean', 'pct_change', 'n_baseline',
              'environment', 'is_outlier']
    outpath = os.path.join(PACKAGE_DIR, '02_Magnetic_Measurements',
                           'h_plate_degradation.csv')
    write_csv(outpath, rows, fields)


def export_a_sample_degradation(a_results):
    """Export A-sample degradation results to CSV."""
    rows = []
    for r in a_results:
        rows.append({
            'assembly_id': r['sample'],
            'plate': r['plate'],
            'slot': r['pair_slot'],
            'pair': r['magnet_idx'],
            'material': r['material'],
            'region': r.get('region', ''),
            'baseline_mean': round(r['bl_mean'], 4),
            'latest_mean': round(r['bl_mean'] * (1 + r['pct_change'] / 100.0), 4),
            'pct_change': round(r['pct_change'], 4),
            'n_baseline': r['n_baseline'],
            'temp_corrected': r['temp_corrected'],
            'environment': 'tunnel',
            'is_outlier': r.get('is_outlier', False),
        })
    fields = ['assembly_id', 'plate', 'slot', 'pair', 'material', 'region',
              'baseline_mean', 'latest_mean', 'pct_change', 'n_baseline',
              'temp_corrected', 'environment', 'is_outlier']
    outpath = os.path.join(PACKAGE_DIR, '02_Magnetic_Measurements',
                           'a_sample_degradation.csv')
    write_csv(outpath, rows, fields)


def export_y_plate_time_series(results):
    """Export per-plate per-date time series of % change."""
    rows = []
    for r in results:
        for dt, pct in r['date_pcts']:
            rows.append({
                'plate': r['plate'],
                'slot': r['slot'],
                'sample_id': r['sample'],
                'material': r['material'],
                'region': r['region'],
                'date': dt.strftime('%Y-%m-%d'),
                'pct_change': round(pct, 4),
            })
    fields = ['plate', 'slot', 'sample_id', 'material', 'region',
              'date', 'pct_change']
    outpath = os.path.join(PACKAGE_DIR, '02_Magnetic_Measurements',
                           'y_plate_time_series.csv')
    write_csv(outpath, rows, fields)


def export_temperature_corrections(temp_final, y_materials):
    """Export temperature correction details for every (sample, date) pair."""
    rows = []
    for (sample, date_str), (t_mean, t_std) in sorted(temp_final.items()):
        # Determine if this date had a Y_BASELINE_TEMP_LOOKUP override
        if date_str in Y_BASELINE_TEMP_LOOKUP:
            corr_temp, corr_unc = Y_BASELINE_TEMP_LOOKUP[date_str]
            source = 'Y_BASELINE_TEMP_LOOKUP'
        else:
            corr_temp = t_mean
            corr_unc = t_std
            source = 'teslameter_probe'

        # Look up alpha for this sample using actual material assignment
        mat = y_materials.get(sample, '')
        alpha_used = ''
        if mat:
            alpha_val = ALPHA.get(mat)
            if alpha_val is not None:
                alpha_used = str(alpha_val)

        rows.append({
            'sample_id': sample,
            'date': date_str,
            'probe_temp_C': round(t_mean, 2),
            'corrected_temp_C': round(corr_temp, 2),
            'temp_uncertainty_C': round(corr_unc, 2),
            'correction_source': source,
            'alpha_used': alpha_used,
        })
    fields = ['sample_id', 'date', 'probe_temp_C', 'corrected_temp_C',
              'temp_uncertainty_C', 'correction_source', 'alpha_used']
    outpath = os.path.join(PACKAGE_DIR, '02_Magnetic_Measurements',
                           'temperature_corrections.csv')
    write_csv(outpath, rows, fields)


# ═══════════════════════════════════════════════════════════════════════════════
# 04_Results — headline_results.csv
# ═══════════════════════════════════════════════════════════════════════════════

def export_headline_results(results, gain_syst):
    """Export key headline numbers as machine-readable CSV."""
    from scipy import stats

    clean = [r for r in results if not r['is_outlier']]

    # Per-material means
    mat_groups = defaultdict(list)
    for r in clean:
        mat_groups[r['material']].append(r['pct_change'])

    # NdFeB-SmCo differential (tunnel)
    intra_diffs, _ = compute_intra_plate_diffs(clean)

    rows = []

    # Overall differential
    diff_mean = diff_sem = diff_sig = 0
    if intra_diffs:
        diff_mean = np.mean(intra_diffs)
        diff_sem = np.std(intra_diffs, ddof=1) / np.sqrt(len(intra_diffs))
        diff_sig = abs(diff_mean) / diff_sem if diff_sem > 0 else 0
        rows.append({
            'metric': 'NdFeB_minus_SmCo_differential_pct',
            'value': round(diff_mean, 3),
            'uncertainty': round(diff_sem, 3),
            'significance': '%.1f sigma' % diff_sig,
            'notes': 'Gain-immune intra-plate differential, N=%d plates' % len(intra_diffs),
        })

    # Lab control differential
    try:
        lab_results = load_lab_y_plates(apply_temp_correction=True)
        lab_diffs = [lr['diff'] for lr in lab_results.values()
                     if not np.isnan(lr['diff'])]
        if lab_diffs:
            lab_mean = np.mean(lab_diffs)
            lab_sem = np.std(lab_diffs, ddof=1) / np.sqrt(len(lab_diffs))
            lab_sig = abs(lab_mean) / lab_sem if lab_sem > 0 else 0
            rows.append({
                'metric': 'lab_control_differential_pct',
                'value': round(lab_mean, 3),
                'uncertainty': round(lab_sem, 3),
                'significance': '%.1f sigma' % lab_sig,
                'notes': 'Lab Y-plate differential (no radiation), N=%d plates' % len(lab_diffs),
            })

            # Tunnel-Lab excess
            if intra_diffs:
                tl_excess = diff_mean - lab_mean
                tl_sem = np.sqrt(diff_sem**2 + lab_sem**2)
                tl_sig = abs(tl_excess) / tl_sem if tl_sem > 0 else 0
                rows.append({
                    'metric': 'tunnel_minus_lab_excess_pct',
                    'value': round(tl_excess, 3),
                    'uncertainty': round(tl_sem, 3),
                    'significance': '%.1f sigma' % tl_sig,
                    'notes': 'Tunnel differential minus lab differential',
                })
    except Exception as e:
        print(f"  Warning: Could not compute lab differential: {e}")

    # Per-material means
    for mat in ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']:
        vals = mat_groups.get(mat, [])
        if vals:
            mean_val = np.mean(vals)
            sem_val = np.std(vals, ddof=1) / np.sqrt(len(vals)) if len(vals) > 1 else 0
            rows.append({
                'metric': '%s_mean_pct_change' % mat,
                'value': round(mean_val, 3),
                'uncertainty': round(sem_val, 3),
                'significance': '',
                'notes': 'N=%d tunnel Y-plate samples (outliers excluded)' % len(vals),
            })

    # Gain systematic
    rows.append({
        'metric': 'gain_systematic_cleaned_pct',
        'value': round(gain_syst.gain_syst, 3),
        'uncertainty': 0,
        'significance': '',
        'notes': 'Half-range of session offsets, cleaned (bad baselines + >%.0f%% excluded)' % gain_syst.pct_threshold,
    })
    rows.append({
        'metric': 'gain_systematic_uncleaned_pct',
        'value': round(gain_syst.gain_syst_raw, 3),
        'uncertainty': 0,
        'significance': '',
        'notes': 'Half-range of session offsets, all samples',
    })

    # Temperature coefficients
    for mat, alpha in ALPHA.items():
        rows.append({
            'metric': 'temp_coefficient_%s_per_C' % mat,
            'value': alpha * 100,  # convert to %/C
            'uncertainty': 0,
            'significance': '',
            'notes': 'Manufacturer specification, used for temperature correction',
        })

    # Differential temperature sensitivity
    nd_alphas = [ALPHA[m] for m in ['N42EH', 'N52SH'] if m in ALPHA]
    sm_alphas = [ALPHA[m] for m in ['SmCo33H', 'SmCo35'] if m in ALPHA]
    if nd_alphas and sm_alphas:
        diff_alpha = abs(np.mean(nd_alphas) - np.mean(sm_alphas)) * 100
        rows.append({
            'metric': 'differential_temp_sensitivity_pct_per_C',
            'value': round(diff_alpha, 3),
            'uncertainty': 0,
            'significance': '',
            'notes': 'mean(NdFeB alpha) - mean(SmCo alpha), determines temp systematic on differential',
        })

    # Systematic uncertainties
    temp_syst = 0.033  # from temperature_corrected_analysis.py
    alpha_syst = 0.014  # from alpha uncertainty
    total_syst = np.sqrt(temp_syst**2 + alpha_syst**2)
    combined_unc = np.sqrt(diff_sem**2 + total_syst**2) if diff_sem > 0 else total_syst
    combined_sig = abs(diff_mean) / combined_unc if combined_unc > 0 else 0
    rows.append({
        'metric': 'temp_systematic_pct',
        'value': round(temp_syst, 3),
        'uncertainty': 0,
        'significance': '',
        'notes': 'Temperature correction systematic on differential (half-range of v2-v3 spread)',
    })
    rows.append({
        'metric': 'alpha_uncertainty_systematic_pct',
        'value': round(alpha_syst, 3),
        'uncertainty': 0,
        'significance': '',
        'notes': 'Alpha coefficient uncertainty propagated to differential',
    })
    rows.append({
        'metric': 'total_systematic_pct',
        'value': round(total_syst, 3),
        'uncertainty': 0,
        'significance': '',
        'notes': 'sqrt(temp_systematic^2 + alpha_systematic^2)',
    })
    rows.append({
        'metric': 'combined_uncertainty_pct',
        'value': round(combined_unc, 3),
        'uncertainty': 0,
        'significance': '%.1f sigma' % combined_sig,
        'notes': 'sqrt(stat^2 + syst^2), stat=%.3f, syst=%.3f' % (diff_sem, total_syst),
    })

    # Dose-degradation correlations (load from merged dose CSV)
    try:
        dose_csv = os.path.join(PACKAGE_DIR, '03_Dosimetry', 'dose_vs_degradation.csv')
        if os.path.exists(dose_csv):
            import csv as csv_mod
            with open(dose_csv) as f:
                reader = csv_mod.DictReader(f)
                dose_rows = list(reader)

            for dose_type, dose_col, label in [
                ('gamma', 'ai_photon_gy', 'AIdata photon (Gy)'),
                ('neutron', 'ai_neutron_rem', 'AIdata neutron (rem)'),
            ]:
                doses = []
                ndfeb_vals = []
                smco_vals = []
                diff_vals = []
                for dr in dose_rows:
                    d = float(dr[dose_col])
                    if d > 0:
                        doses.append(np.log10(d))
                        ndfeb_vals.append(float(dr['ndfeb_mean_pct']))
                        smco_vals.append(float(dr['smco_mean_pct']))
                        diff_vals.append(float(dr['intra_plate_diff']))

                if len(doses) >= 5:
                    for metric_name, vals, desc in [
                        ('ndfeb_mean', ndfeb_vals, 'NdFeB mean degradation'),
                        ('smco_mean', smco_vals, 'SmCo mean degradation'),
                        ('differential', diff_vals, 'NdFeB-SmCo differential'),
                    ]:
                        rho, p = stats.spearmanr(doses, vals)
                        rows.append({
                            'metric': '%s_vs_%s_spearman_rho' % (dose_type, metric_name),
                            'value': round(rho, 3),
                            'uncertainty': round(p, 4),
                            'significance': 'p=%.4f' % p,
                            'notes': 'Spearman rho of log10(%s) vs %s, N=%d' % (label, desc, len(doses)),
                        })
    except Exception as e:
        print(f"  Warning: Could not compute dose correlations: {e}")

    fields = ['metric', 'value', 'uncertainty', 'significance', 'notes']
    outpath = os.path.join(PACKAGE_DIR, '04_Results', 'headline_results.csv')
    write_csv(outpath, rows, fields)


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("LDRD FFA@CEBAF Data Package Export")
    print("=" * 70)

    # Step 1: Load all Y-plate data
    print("\n[1/8] Loading Y-plate data (v3)...")
    results, helm_raw, temp_final, y_materials = load_all()
    print("  Loaded %d Y-plate results" % len(results))

    # Step 2: Load H-plate data
    print("\n[2/8] Loading H-plate data (v2)...")
    y_mats_v2, pair_arrangements = load_materials()
    temp_lookup = build_temperature_lookup()
    h_results, h_excluded = compute_h_plate_degradation(pair_arrangements, temp_lookup)
    print("  Loaded %d H-plate results (%d excluded)" % (len(h_results), len(h_excluded)))

    # Step 3: Load A-sample data
    print("\n[3/8] Loading A-sample data (v5_polish)...")
    a_results = load_a_sample_helmholtz(temp_lookup)
    print("  Loaded %d A-sample results" % len(a_results))

    # Step 4: Compute gain systematic
    print("\n[4/8] Computing gain systematic...")
    gain_syst = get_gain_syst(helm_raw)
    print("  Gain systematic: +/-%.3f%% (cleaned), +/-%.3f%% (uncleaned)" %
          (gain_syst.gain_syst, gain_syst.gain_syst_raw))

    # Step 5: Export 01_Sample_Configuration
    print("\n[5/8] Exporting sample configuration CSVs...")
    export_materials_arrangements()
    export_tunnel_placements()
    export_sample_inventory()

    # Step 6: Export 02_Magnetic_Measurements
    print("\n[6/8] Exporting magnetic measurement CSVs...")
    export_y_plate_degradation(results, y_materials)
    export_h_plate_degradation(h_results)
    export_a_sample_degradation(a_results)
    export_y_plate_time_series(results)
    export_temperature_corrections(temp_final, y_materials)

    # Step 7: Export 04_Results
    print("\n[7/8] Exporting headline results...")
    export_headline_results(results, gain_syst)

    # Step 8: Summary
    print("\n[8/8] Export complete!")
    print("=" * 70)

    # Verification summary
    clean = [r for r in results if not r['is_outlier']]
    tunnel = [r for r in results if r['region'] != '']
    print("\nVerification:")
    print("  Total Y-plate results: %d" % len(results))
    print("  Tunnel Y-plates: %d" % len(tunnel))
    print("  Clean (non-outlier): %d" % len(clean))
    print("  H-plate results: %d" % len(h_results))
    print("  A-sample results: %d" % len(a_results))

    intra_diffs, _ = compute_intra_plate_diffs(clean)
    if intra_diffs:
        diff_mean = np.mean(intra_diffs)
        diff_sem = np.std(intra_diffs, ddof=1) / np.sqrt(len(intra_diffs))
        print("  NdFeB-SmCo differential: %.3f%% +/- %.3f%% (%.1f sigma)" %
              (diff_mean, diff_sem, abs(diff_mean) / diff_sem))


if __name__ == '__main__':
    main()
