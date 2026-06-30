#!/usr/bin/env python3
"""
Comprehensive validation and overview plots for the Cleanup_Claude merged data.

Produces:
  1. All Y-plate Helmholtz time-series (one subplot per plate, 4 materials overlaid)
  2. Y-plate Helmholtz grouped by material (all samples of same grade overlaid)
  3. Pair assembly (Hn/Hs) Helmholtz time-series
  4. Teslameter time-series for selected Y-plates and pair magnets
  5. Aug 26 vs Aug 27 direct comparison (reproducing prior plot)
  6. Jul17 vs Jul30 group systematic comparison (reproducing prior analysis)
  7. Intra-plate NdFeB-SmCo differential (reproducing prior analysis Table)
"""

import os
import re
import glob
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
from datetime import datetime
from collections import defaultdict
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
PLOT_DIR = os.path.join(BASE, 'Validation_Plots')
os.makedirs(PLOT_DIR, exist_ok=True)

# Colorblind-safe palette
CB = {
    'N42EH':   '#EE6677',  # red
    'N52SH':   '#4477AA',  # blue
    'SmCo33H': '#228833',  # green
    'SmCo35':  '#CCBB44',  # yellow
    'NdFeB':   '#AA3377',  # purple
    'SmCo':    '#66CCEE',  # cyan
}
CB_LIST = ['#4477AA', '#EE6677', '#228833', '#CCBB44', '#66CCEE', '#AA3377', '#BBBBBB', '#000000']

# ─── Material assignments ────────────────────────────────────────────────────

def load_materials():
    """Load material assignments from the Materials spreadsheet."""
    mat_path = os.path.join(BASE, 'Materials_Arrangements.xlsx')
    wb = openpyxl.load_workbook(mat_path)

    # Y-plate materials: plate -> [slot1_material, slot2, slot3, slot4]
    y_materials = {}
    ws = wb['Tunnel - Y Materials']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()  # e.g. 'y-1'
            plate_num = plate_id.replace('y-', '')
            y_materials[plate_num] = [str(row[i]).strip() if row[i] else '' for i in range(1, 5)]

    # Pair arrangements: plate -> (material_type, [slot1_config, ...])
    pair_arrangements = {}
    ws2 = wb['Tunnel - Pair Arrangements']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()  # e.g. 'n-6' or 's-1'
            mat_type = str(row[1]).strip() if row[1] else ''
            configs = [str(row[i]).strip() if row[i] else '' for i in range(2, 6)]
            pair_arrangements[plate_id] = (mat_type, configs)

    return y_materials, pair_arrangements


def get_y_material(y_materials, sample_name):
    """Get material for a Y-plate sample like 'Y-22-4'."""
    m = re.match(r'Y-(\d+)-(\d+)', sample_name)
    if not m:
        return None
    plate = m.group(1)
    slot = int(m.group(2))
    if plate in y_materials and 1 <= slot <= 4:
        return y_materials[plate][slot - 1]
    return None


# Jul 17 plates vs Jul 30 plates (from prior analysis)
JUL17_PLATES = {4, 6, 7, 9, 12, 16, 17, 18, 21, 22, 25, 34, 36, 38, 39}
JUL30_PLATES = {1, 3, 5, 10, 11, 13, 15, 19, 20, 23, 24, 26, 30, 32, 40}

# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_helmholtz_file(filepath):
    """Parse a helmholtz .dat file. Returns list of (datetime, value, unit)."""
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # New format
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                # Old format
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            val_match = re.search(r'([+-]?\d+\.?\d*)\s*(mWC|kT|kBG)', rest)
            if val_match:
                value = float(val_match.group(1))
                unit = val_match.group(2)
                rows.append((dt, value, unit))
    return rows


def parse_teslameter_file(filepath):
    """Parse a teslameter .dat file. Returns list of (datetime, [f1,f2,f3], temp)."""
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            nums = re.findall(r'(-?\d+\.\d+)', rest)
            if len(nums) >= 4:
                rows.append((dt, [float(x) for x in nums[:3]], float(nums[3])))
            elif len(nums) >= 3:
                rows.append((dt, [float(x) for x in nums[:3]], None))
    return rows


def get_mwc_rows(rows):
    """Filter to only mWC unit rows."""
    return [(dt, val) for dt, val, unit in rows if unit == 'mWC']


def get_tunnel_rows(rows):
    """Filter to tunnel campaign period only (Jul 2025+), mWC only.
    Also removes obvious bad re-takes: if two readings are within 5 minutes
    on the same day and one is >50% different from the median, drop the outlier."""
    tunnel = [(dt, val) for dt, val, unit in rows
              if unit == 'mWC' and dt >= datetime(2025, 7, 1)]
    return _clean_retakes(tunnel)


def get_all_mwc_rows(rows):
    """All mWC rows (including pre-deployment baselines), with re-take cleaning."""
    all_mwc = [(dt, val) for dt, val, unit in rows if unit == 'mWC']
    return _clean_retakes(all_mwc)


def _clean_retakes(data):
    """Remove obvious bad re-takes from a list of (dt, val) tuples."""
    if len(data) < 3:
        return data

    median_val = np.median([v for _, v in data])

    cleaned = []
    skip = set()
    for i in range(len(data)):
        if i in skip:
            continue
        if i + 1 < len(data):
            dt1, v1 = data[i]
            dt2, v2 = data[i + 1]
            if abs((dt2 - dt1).total_seconds()) < 300:
                if abs(v1 - median_val) > abs(v2 - median_val):
                    skip.add(i)
                    continue
                else:
                    skip.add(i + 1)
        cleaned.append(data[i])

    for i in range(len(data)):
        if i not in skip and data[i] not in cleaned:
            cleaned.append(data[i])

    return sorted(cleaned, key=lambda x: x[0])


# ─── Plot 1: All Y-plate Helmholtz time-series ──────────────────────────────

def plot_all_y_plates_helmholtz(y_materials):
    """One subplot per Y-plate, 4 material slots overlaid."""
    # Get all Y-plate numbers
    plate_nums = sorted(set(
        int(re.match(r'Y-(\d+)-\d+', f.replace('_helmholtz.dat', '')).group(1))
        for f in os.listdir(os.path.join(BASE, 'Y_Plates', 'Helmholtz'))
        if re.match(r'Y-\d+-\d+_helmholtz\.dat$', f)
    ))

    n_plates = len(plate_nums)
    ncols = 5
    nrows = (n_plates + ncols - 1) // ncols

    fig, axes = plt.subplots(nrows, ncols, figsize=(24, 4 * nrows), sharex=True)
    axes = axes.flatten()

    for idx, plate_num in enumerate(plate_nums):
        ax = axes[idx]
        for slot in range(1, 5):
            sample = f'Y-{plate_num}-{slot}'
            fpath = os.path.join(BASE, 'Y_Plates', 'Helmholtz', f'{sample}_helmholtz.dat')
            if not os.path.exists(fpath):
                continue
            rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
            if not rows:
                continue
            dates, vals = zip(*rows)
            material = get_y_material(y_materials, sample) or f'Slot {slot}'
            color = CB.get(material, CB_LIST[slot - 1])
            ax.plot(dates, vals, 'o-', color=color, markersize=3, linewidth=1,
                    label=f'{slot}: {material}')

        ax.set_title(f'Y-{plate_num}', fontsize=10, fontweight='bold')
        ax.tick_params(labelsize=7)
        ax.grid(True, alpha=0.2)
        if idx == 0:
            ax.legend(fontsize=6, loc='best')

    # Hide unused axes
    for idx in range(n_plates, len(axes)):
        axes[idx].set_visible(False)

    # Format x-axis on bottom row
    for idx in range((nrows - 1) * ncols, min(nrows * ncols, n_plates)):
        axes[idx].xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        axes[idx].xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in axes[idx].get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('All Y-Plate Helmholtz Time-Series (mWC, full timeline)',
                 fontsize=16, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.97])

    outpath = os.path.join(PLOT_DIR, '01_all_y_plates_helmholtz.png')
    plt.savefig(outpath, dpi=150)
    print(f"Saved: {outpath}")
    plt.close()


# ─── Plot 2: Y-plate Helmholtz grouped by material ──────────────────────────

def plot_y_by_material(y_materials):
    """One subplot per material grade, all Y-plate samples of that grade overlaid."""
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 2, figsize=(18, 12), sharex=True)
    axes = axes.flatten()

    for ax, material in zip(axes, materials):
        # Find all samples of this material
        sample_data = []
        for f in sorted(os.listdir(os.path.join(BASE, 'Y_Plates', 'Helmholtz'))):
            if not f.endswith('_helmholtz.dat'):
                continue
            sample = f.replace('_helmholtz.dat', '')
            mat = get_y_material(y_materials, sample)
            if mat and mat.strip() == material:
                fpath = os.path.join(BASE, 'Y_Plates', 'Helmholtz', f)
                rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
                if rows:
                    sample_data.append((sample, rows))

        for i, (sample, rows) in enumerate(sample_data):
            dates, vals = zip(*rows)
            # Normalize to first measurement for comparison
            ax.plot(dates, vals, 'o-', markersize=3, linewidth=0.8, alpha=0.6,
                    label=sample if i < 5 else None)  # Only label first 5 to keep legend sane

        ax.set_title(f'{material} (n={len(sample_data)} samples)', fontsize=13, fontweight='bold')
        ax.set_ylabel('mWC', fontsize=11)
        ax.grid(True, alpha=0.3)
        if len(sample_data) <= 10:
            ax.legend(fontsize=7, ncol=2, loc='best')

    for ax in axes[2:]:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in ax.get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('Y-Plate Helmholtz by Material Grade (mWC, full timeline)',
                 fontsize=15, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.96])

    outpath = os.path.join(PLOT_DIR, '02_y_plates_by_material.png')
    plt.savefig(outpath, dpi=150)
    print(f"Saved: {outpath}")
    plt.close()


# ─── Plot 2b: Normalized % change by material ───────────────────────────────

def plot_y_pct_change_by_material(y_materials):
    """% change from first tunnel measurement, grouped by material."""
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 2, figsize=(18, 12), sharex=True)
    axes = axes.flatten()

    for ax, material in zip(axes, materials):
        sample_data = []
        for f in sorted(os.listdir(os.path.join(BASE, 'Y_Plates', 'Helmholtz'))):
            if not f.endswith('_helmholtz.dat'):
                continue
            sample = f.replace('_helmholtz.dat', '')
            mat = get_y_material(y_materials, sample)
            if mat and mat.strip() == material:
                fpath = os.path.join(BASE, 'Y_Plates', 'Helmholtz', f)
                rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
                if len(rows) >= 2:
                    sample_data.append((sample, rows))

        pct_traces = []
        for sample, rows in sample_data:
            dates, vals = zip(*rows)
            # Use median to find stable baseline, then use first value within 20% of median
            median_val = np.median(vals)
            base = None
            base_idx = 0
            for i, v in enumerate(vals):
                if abs(v - median_val) / abs(median_val) < 0.20:
                    base = v
                    base_idx = i
                    break
            if base is None or base == 0:
                continue
            # Only plot from the first stable reading onward
            pcts = [(v - base) / abs(base) * 100 for v in vals[base_idx:]]
            pct_dates = dates[base_idx:]
            pct_traces.append((sample, pct_dates, pcts))

        for i, (sample, dates, pcts) in enumerate(pct_traces):
            ax.plot(dates, pcts, 'o-', markersize=3, linewidth=0.8, alpha=0.5)

        # Mean trace
        if pct_traces:
            # Collect all unique dates across samples
            all_dates = sorted(set(d for _, dates, _ in pct_traces for d in dates))
            mean_by_date = {}
            for d in all_dates:
                vals_at_d = []
                for _, dates, pcts in pct_traces:
                    for dd, pp in zip(dates, pcts):
                        if dd == d:
                            vals_at_d.append(pp)
                if vals_at_d:
                    mean_by_date[d] = np.mean(vals_at_d)

            if mean_by_date:
                md, mv = zip(*sorted(mean_by_date.items()))
                ax.plot(md, mv, 'k-', linewidth=2.5, label='Mean', zorder=10)

        ax.axhline(0, color='gray', linestyle='--', linewidth=0.8)
        ax.set_title(f'{material} (n={len(pct_traces)})', fontsize=13, fontweight='bold')
        ax.set_ylabel('% change from first tunnel meas.', fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.legend(fontsize=9)

    for ax in axes[2:]:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in ax.get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('Y-Plate Helmholtz % Change from First Measurement (by material)',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.96])

    outpath = os.path.join(PLOT_DIR, '03_y_plates_pct_change_by_material.png')
    plt.savefig(outpath, dpi=150)
    print(f"Saved: {outpath}")
    plt.close()


# ─── Plot 3: Pair Assembly Helmholtz time-series ────────────────────────────

def plot_pair_assemblies_helmholtz(pair_arrangements):
    """Hn and Hs plate Helmholtz time-series."""
    helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')

    # Collect Hn and Hs files
    hn_files = sorted(glob.glob(os.path.join(helm_dir, 'Hn-*_helmholtz.dat')))
    hs_files = sorted(glob.glob(os.path.join(helm_dir, 'Hs-*_helmholtz.dat')))

    for prefix, files, title in [('Hn', hn_files, 'NdFeB Pair Plates (Hn)'),
                                   ('Hs', hs_files, 'SmCo Pair Plates (Hs)')]:
        # Group by plate number
        plates = defaultdict(list)
        for f in files:
            sample = os.path.basename(f).replace('_helmholtz.dat', '')
            m = re.match(r'H[ns]-(\d+)-(\d+)', sample)
            if m:
                plates[int(m.group(1))].append((int(m.group(2)), sample, f))

        plate_nums = sorted(plates.keys())
        n = len(plate_nums)
        if n == 0:
            continue

        ncols = 5
        nrows = max(1, (n + ncols - 1) // ncols)
        fig, axes = plt.subplots(nrows, ncols, figsize=(24, 4 * nrows), sharex=True)
        if nrows == 1 and ncols == 1:
            axes = np.array([axes])
        axes = axes.flatten()

        configs = ['Alpha', 'Beta', 'Gamma', 'Delta']
        config_colors = {'Alpha': CB_LIST[0], 'Beta': CB_LIST[1],
                         'Gamma': CB_LIST[2], 'Delta': CB_LIST[3]}

        for idx, plate_num in enumerate(plate_nums):
            ax = axes[idx]
            slots = sorted(plates[plate_num])
            plate_key = f"{'n' if prefix == 'Hn' else 's'}-{plate_num}"

            for slot_num, sample, fpath in slots:
                rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
                if not rows:
                    continue
                dates, vals = zip(*rows)

                # Get configuration if available
                config = ''
                if plate_key in pair_arrangements:
                    _, cfgs = pair_arrangements[plate_key]
                    if 1 <= slot_num <= 4:
                        config = cfgs[slot_num - 1]
                color = config_colors.get(config, CB_LIST[slot_num - 1])
                label = f'{slot_num}: {config}' if config else f'Slot {slot_num}'

                ax.plot(dates, vals, 'o-', color=color, markersize=3,
                        linewidth=1, label=label)

            ax.set_title(f'{prefix}-{plate_num}', fontsize=10, fontweight='bold')
            ax.tick_params(labelsize=7)
            ax.grid(True, alpha=0.2)
            if idx == 0:
                ax.legend(fontsize=6, loc='best')

        for idx in range(n, len(axes)):
            axes[idx].set_visible(False)

        for idx in range(max(0, (nrows - 1) * ncols), min(nrows * ncols, n)):
            axes[idx].xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
            axes[idx].xaxis.set_major_locator(mdates.MonthLocator(interval=2))

        fig.suptitle(f'{title} — Helmholtz Time-Series (mWC, full timeline)',
                     fontsize=15, fontweight='bold')
        fig.tight_layout(rect=[0, 0, 1, 0.97])

        outpath = os.path.join(PLOT_DIR, f'04_{prefix.lower()}_pair_assemblies.png')
        plt.savefig(outpath, dpi=150)
        print(f"Saved: {outpath}")
        plt.close()


# ─── Plot 4: Teslameter time-series ─────────────────────────────────────────

def plot_teslameter_overview(y_materials):
    """Teslameter front/side/top for a selection of Y-plates."""
    # Pick a spread of samples
    samples = ['Y-1-1', 'Y-7-3', 'Y-15-1', 'Y-22-4', 'Y-30-1', 'Y-40-2']
    faces = ['front', 'side', 'top']
    face_labels = {'front': 'Front', 'side': 'Side', 'top': 'Top'}

    fig, axes = plt.subplots(len(samples), 3, figsize=(22, 3.5 * len(samples)),
                              sharex=True)

    for row_idx, sample in enumerate(samples):
        material = get_y_material(y_materials, sample) or '?'
        for col_idx, face in enumerate(faces):
            ax = axes[row_idx][col_idx]
            if sample.startswith('Y-'):
                fpath = os.path.join(BASE, 'Y_Plates', 'Teslameter', f'{sample}_{face}.dat')
            else:
                fpath = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter', f'{sample}_{face}.dat')

            if not os.path.exists(fpath):
                ax.text(0.5, 0.5, 'No data', transform=ax.transAxes, ha='center')
                continue

            rows = parse_teslameter_file(fpath)
            if not rows:
                ax.text(0.5, 0.5, 'No data', transform=ax.transAxes, ha='center')
                continue

            dates = [r[0] for r in rows]
            f1 = [r[1][0] for r in rows]
            f2 = [r[1][1] for r in rows]
            f3 = [r[1][2] for r in rows]

            ax.plot(dates, f1, 'o-', color=CB_LIST[0], markersize=3, linewidth=1, label='F1')
            ax.plot(dates, f2, 's-', color=CB_LIST[1], markersize=3, linewidth=1, label='F2')
            ax.plot(dates, f3, '^-', color=CB_LIST[2], markersize=3, linewidth=1, label='F3')

            ax.grid(True, alpha=0.2)
            ax.tick_params(labelsize=7)

            if col_idx == 0:
                ax.set_ylabel(f'{sample}\n({material})', fontsize=9, fontweight='bold')
            if row_idx == 0:
                ax.set_title(f'{face_labels[face]}', fontsize=11, fontweight='bold')
            if row_idx == 0 and col_idx == 0:
                ax.legend(fontsize=7, ncol=3)

    for ax in axes[-1]:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%y'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        for label in ax.get_xticklabels():
            label.set_rotation(45)

    fig.suptitle('Teslameter Overview — Selected Y-Plates (mT, full timeline)',
                 fontsize=15, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.97])

    outpath = os.path.join(PLOT_DIR, '05_teslameter_overview.png')
    plt.savefig(outpath, dpi=150)
    print(f"Saved: {outpath}")
    plt.close()


# ─── Plot 5: Aug 26 vs Aug 27 comparison ────────────────────────────────────

def plot_aug26_vs_aug27():
    """Reproduce the prior Compare_Lab_Tunnel Helmholtz comparison."""
    samples = [
        'Y-8-1', 'Y-8-2', 'Y-8-3', 'Y-8-4',
        'Hs-29-1', 'Hs-29-2', 'Hs-29-3', 'Hs-29-4',
        'As-29-1-1', 'As-29-1-2', 'As-29-2-1', 'As-29-2-2',
        'As-29-3-1', 'As-29-3-2', 'As-29-4-1', 'As-29-4-2',
    ]

    initial_vals = []
    final_vals = []
    valid_samples = []

    for sample in samples:
        if sample.startswith('Y-'):
            fpath = os.path.join(BASE, 'Y_Plates', 'Helmholtz', f'{sample}_helmholtz.dat')
        else:
            fpath = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz', f'{sample}_helmholtz.dat')

        if not os.path.exists(fpath):
            continue
        rows = parse_helmholtz_file(fpath)
        init = final = None
        for dt, val, unit in rows:
            ds = dt.strftime('%Y-%m-%d')
            if ds == '2025-08-26':
                init = val
            elif ds == '2025-08-27':
                final = val
        if init is not None and final is not None:
            initial_vals.append(init)
            final_vals.append(final)
            valid_samples.append(sample)

    if not valid_samples:
        print("No data for Aug 26 vs 27 plot")
        return

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(16, 11))
    x = np.arange(len(valid_samples))

    # Raw
    ax1.plot(x, initial_vals, 'o-', color=CB_LIST[0], label='Initial (2025-08-26)', markersize=7)
    ax1.plot(x, final_vals, 's-', color=CB_LIST[1], label='Final (2025-08-27)', markersize=7)
    ax1.set_xticks(x)
    ax1.set_xticklabels(valid_samples, rotation=45, ha='right', fontsize=10)
    ax1.set_ylabel('Measurement (mWC)', fontsize=12)
    ax1.set_title('Raw Helmholtz Measurements', fontsize=14, fontweight='bold')
    ax1.legend(fontsize=11)
    ax1.grid(True, alpha=0.3)

    # Percent difference
    pct = [((f - i) / abs(i)) * 100 if i != 0 else 0
           for i, f in zip(initial_vals, final_vals)]

    bars = ax2.bar(x, pct, width=0.6, color=CB_LIST[0])
    for bar, d in zip(bars, pct):
        ax2.annotate(f'{d:.1f}%',
                     xy=(bar.get_x() + bar.get_width() / 2, bar.get_height()),
                     xytext=(0, 5 if d >= 0 else -14),
                     textcoords='offset points', ha='center',
                     va='bottom' if d >= 0 else 'top',
                     fontsize=9, fontweight='bold')
    ax2.set_xticks(x)
    ax2.set_xticklabels(valid_samples, rotation=45, ha='right', fontsize=10)
    ax2.set_ylabel('Percent Difference (%)', fontsize=12)
    ax2.set_title('Percent Difference in Helmholtz (Aug 26 → Aug 27)', fontsize=14, fontweight='bold')
    ax2.axhline(0, color='black', linewidth=0.8, linestyle='--')
    ax2.grid(axis='y', alpha=0.3)

    fig.tight_layout()
    outpath = os.path.join(PLOT_DIR, '06_aug26_vs_aug27_comparison.png')
    plt.savefig(outpath, dpi=200)
    print(f"Saved: {outpath}")
    plt.close()


# ─── Plot 6: Jul 17 vs Jul 30 group systematic ─────────────────────────────

def plot_jul17_vs_jul30_systematic(y_materials):
    """
    Reproduce the Jul17 vs Jul30 group systematic comparison from the prior analysis.
    Shows mean raw % change by material for each group.
    """
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    # Collect per-sample first-to-last % change
    group_data = {
        'Jul 17': defaultdict(list),
        'Jul 30': defaultdict(list),
    }

    for f in os.listdir(os.path.join(BASE, 'Y_Plates', 'Helmholtz')):
        if not f.endswith('_helmholtz.dat'):
            continue
        sample = f.replace('_helmholtz.dat', '')
        m = re.match(r'Y-(\d+)-(\d+)', sample)
        if not m:
            continue
        plate_num = int(m.group(1))
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()
        if material not in materials:
            continue

        if plate_num in JUL17_PLATES:
            group = 'Jul 17'
        elif plate_num in JUL30_PLATES:
            group = 'Jul 30'
        else:
            continue

        fpath = os.path.join(BASE, 'Y_Plates', 'Helmholtz', f)
        rows = get_tunnel_rows(parse_helmholtz_file(fpath))
        if len(rows) < 2:
            continue

        first_val = rows[0][1]
        last_val = rows[-1][1]
        if first_val == 0:
            continue
        pct = (last_val - first_val) / abs(first_val) * 100
        group_data[group][material].append(pct)

    # Plot grouped bar chart
    fig, ax = plt.subplots(figsize=(12, 7))
    x = np.arange(len(materials))
    width = 0.35

    jul17_means = [np.mean(group_data['Jul 17'].get(m, [0])) for m in materials]
    jul30_means = [np.mean(group_data['Jul 30'].get(m, [0])) for m in materials]
    jul17_stds = [np.std(group_data['Jul 17'].get(m, [0])) for m in materials]
    jul30_stds = [np.std(group_data['Jul 30'].get(m, [0])) for m in materials]
    jul17_n = [len(group_data['Jul 17'].get(m, [])) for m in materials]
    jul30_n = [len(group_data['Jul 30'].get(m, [])) for m in materials]

    bars1 = ax.bar(x - width/2, jul17_means, width, yerr=jul17_stds, capsize=4,
                   label='Jul 17 group', color=CB_LIST[0], alpha=0.8)
    bars2 = ax.bar(x + width/2, jul30_means, width, yerr=jul30_stds, capsize=4,
                   label='Jul 30 group', color=CB_LIST[1], alpha=0.8)

    # Annotate with values
    for bar, val, n in zip(bars1, jul17_means, jul17_n):
        ax.annotate(f'{val:+.3f}%\n(n={n})',
                    xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                    xytext=(0, 8), textcoords='offset points',
                    ha='center', fontsize=9, fontweight='bold')
    for bar, val, n in zip(bars2, jul30_means, jul30_n):
        ax.annotate(f'{val:+.3f}%\n(n={n})',
                    xy=(bar.get_x() + bar.get_width()/2, bar.get_height()),
                    xytext=(0, 8), textcoords='offset points',
                    ha='center', fontsize=9, fontweight='bold')

    ax.set_xticks(x)
    ax.set_xticklabels(materials, fontsize=12)
    ax.set_ylabel('Mean raw % change (first → last tunnel meas.)', fontsize=12)
    ax.set_title('Jul 17 vs Jul 30 Group Systematic\n(raw % change, first to last tunnel measurement)',
                 fontsize=14, fontweight='bold')
    ax.legend(fontsize=12)
    ax.grid(axis='y', alpha=0.3)
    ax.axhline(0, color='black', linewidth=0.8, linestyle='--')

    fig.tight_layout()
    outpath = os.path.join(PLOT_DIR, '07_jul17_vs_jul30_systematic.png')
    plt.savefig(outpath, dpi=200)
    print(f"Saved: {outpath}")
    plt.close()

    # Print comparison table
    print("\n  Prior analysis values for comparison:")
    print("  Material   | Jul17 (prior) | Jul17 (now) | Jul30 (prior) | Jul30 (now)")
    prior = {'N42EH': (1.280, 1.200), 'N52SH': (1.379, 0.456),
             'SmCo33H': (1.011, 0.229), 'SmCo35': (0.976, 0.149)}
    for m in materials:
        p17, p30 = prior[m]
        n17 = np.mean(group_data['Jul 17'].get(m, [0]))
        n30 = np.mean(group_data['Jul 30'].get(m, [0]))
        print(f"  {m:10s} | {p17:+.3f}%       | {n17:+.3f}%     | {p30:+.3f}%       | {n30:+.3f}%")


# ─── Plot 7: Intra-plate NdFeB-SmCo differential ────────────────────────────

def plot_intraplate_differential(y_materials):
    """
    Reproduce the intra-plate NdFeB-SmCo differential analysis.
    For each plate, compute (NdFeB avg % change) - (SmCo avg % change).
    """
    # For each plate, get first-to-last % change for each material
    plate_diffs = {}  # plate_num -> {material: pct_change}

    for f in os.listdir(os.path.join(BASE, 'Y_Plates', 'Helmholtz')):
        if not f.endswith('_helmholtz.dat'):
            continue
        sample = f.replace('_helmholtz.dat', '')
        m = re.match(r'Y-(\d+)-(\d+)', sample)
        if not m:
            continue
        plate_num = int(m.group(1))
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()

        fpath = os.path.join(BASE, 'Y_Plates', 'Helmholtz', f)
        rows = get_tunnel_rows(parse_helmholtz_file(fpath))
        if len(rows) < 2:
            continue
        first_val = rows[0][1]
        last_val = rows[-1][1]
        if first_val == 0:
            continue
        pct = (last_val - first_val) / abs(first_val) * 100

        if plate_num not in plate_diffs:
            plate_diffs[plate_num] = {}
        plate_diffs[plate_num][material] = pct

    # Compute NdFeB-SmCo differential per plate
    differentials = []
    plate_labels = []
    for plate_num in sorted(plate_diffs.keys()):
        d = plate_diffs[plate_num]
        ndfeb_vals = [d.get('N42EH'), d.get('N52SH')]
        smco_vals = [d.get('SmCo33H'), d.get('SmCo35')]
        ndfeb_vals = [v for v in ndfeb_vals if v is not None]
        smco_vals = [v for v in smco_vals if v is not None]
        if ndfeb_vals and smco_vals:
            diff = np.mean(ndfeb_vals) - np.mean(smco_vals)
            differentials.append(diff)
            plate_labels.append(f'Y-{plate_num}')

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))

    # Bar chart of per-plate differential
    x = np.arange(len(plate_labels))
    colors = [CB_LIST[0] if d >= 0 else CB_LIST[1] for d in differentials]
    ax1.bar(x, differentials, color=colors, alpha=0.8)
    ax1.set_xticks(x)
    ax1.set_xticklabels(plate_labels, rotation=45, ha='right', fontsize=9)
    ax1.axhline(0, color='black', linewidth=0.8, linestyle='--')
    ax1.set_ylabel('NdFeB - SmCo differential (%)', fontsize=11)
    ax1.set_title('Per-Plate NdFeB - SmCo Differential', fontsize=13, fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)

    # Mark outliers
    mean_diff = np.mean(differentials)
    std_diff = np.std(differentials)
    ax1.axhline(mean_diff, color=CB_LIST[2], linewidth=2, linestyle='-',
                label=f'Mean: {mean_diff:+.3f}%')
    ax1.legend(fontsize=10)

    # Histogram
    ax2.hist(differentials, bins=15, color=CB_LIST[0], alpha=0.7, edgecolor='black')
    ax2.axvline(mean_diff, color=CB_LIST[1], linewidth=2, linestyle='--',
                label=f'Mean: {mean_diff:+.3f}% ± {std_diff:.3f}%')
    ax2.axvline(0, color='black', linewidth=0.8, linestyle='--')
    ax2.set_xlabel('NdFeB - SmCo differential (%)', fontsize=11)
    ax2.set_ylabel('Count', fontsize=11)
    ax2.set_title('Distribution of Intra-Plate Differentials', fontsize=13, fontweight='bold')
    ax2.legend(fontsize=10)

    fig.suptitle(f'Intra-Plate NdFeB - SmCo Differential (n={len(differentials)} plates)\n'
                 f'Prior analysis: mean = +0.331%, this data: mean = {mean_diff:+.3f}%',
                 fontsize=14, fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.93])

    outpath = os.path.join(PLOT_DIR, '08_intraplate_differential.png')
    plt.savefig(outpath, dpi=200)
    print(f"Saved: {outpath}")
    print(f"  Intra-plate differential: mean={mean_diff:+.4f}%, std={std_diff:.4f}%, "
          f"n={len(differentials)} plates")
    print(f"  Prior analysis:           mean=+0.3310%, std=0.1460%, n=28 plates")
    plt.close()


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Generating comprehensive validation plots...")
    print("=" * 60)

    y_materials, pair_arrangements = load_materials()
    print(f"Loaded {len(y_materials)} Y-plate and {len(pair_arrangements)} pair plate material assignments\n")

    print("Plot 1: All Y-plate Helmholtz time-series...")
    plot_all_y_plates_helmholtz(y_materials)

    print("\nPlot 2: Y-plates grouped by material...")
    plot_y_by_material(y_materials)

    print("\nPlot 3: Y-plates % change by material...")
    plot_y_pct_change_by_material(y_materials)

    print("\nPlot 4: Pair assembly Helmholtz time-series...")
    plot_pair_assemblies_helmholtz(pair_arrangements)

    print("\nPlot 5: Teslameter overview...")
    plot_teslameter_overview(y_materials)

    print("\nPlot 6: Aug 26 vs Aug 27 comparison...")
    plot_aug26_vs_aug27()

    print("\nPlot 7: Jul 17 vs Jul 30 group systematic...")
    plot_jul17_vs_jul30_systematic(y_materials)

    print("\nPlot 8: Intra-plate NdFeB-SmCo differential...")
    plot_intraplate_differential(y_materials)

    print("\n" + "=" * 60)
    print(f"All plots saved to: {PLOT_DIR}/")
    print("=" * 60)


if __name__ == '__main__':
    main()
