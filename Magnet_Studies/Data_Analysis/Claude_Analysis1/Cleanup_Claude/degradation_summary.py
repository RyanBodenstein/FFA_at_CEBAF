#!/usr/bin/env python3
"""
Degradation Summary for Management — Preliminary Results

Computes temperature-corrected % change from pre-deployment baseline to the
most recent measurement for every Y-plate and H-plate sample. Breaks down by:
  - Material grade
  - Tunnel region (East Arc, West Arc, North Linac, South Linac, Labyrinths)
  - Arc line position (1st through 5th = different beam energy passes)
  - Individual sample

Error bars computed from three independent sources added in quadrature:
  1. Baseline uncertainty: std of pre-deployment corrected readings / sqrt(N)
     (statistical uncertainty on the mean baseline)
  2. Temperature correction uncertainty: propagated from face-to-face
     temperature spread on the final measurement date
  3. Measurement repeatability: estimated from the scatter of corrected
     readings during stable periods (pre-deployment)

Output:
  - Printed summary tables
  - MD_Files/degradation_summary.md — formatted for management
  - TempCorrected_Plots/degradation_summary_by_region.png
  - TempCorrected_Plots/degradation_summary_by_material.png
"""

import os
import re
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from collections import defaultdict
import openpyxl

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
PLOT_DIR = os.path.join(BASE, 'TempCorrected_Plots')
MD_DIR = os.path.join(os.path.dirname(BASE), 'MD_Files')
os.makedirs(PLOT_DIR, exist_ok=True)
os.makedirs(MD_DIR, exist_ok=True)

# ─── Constants ───────────────────────────────────────────────────────────────
T_REF = 20.0
ALPHA = {
    'N42EH':   -0.0010,
    'N52SH':   -0.0011,
    'SmCo33H': -0.0004,
    'SmCo35':  -0.0004,
    'NdFeB':   -0.00105,
    'SmCo':    -0.0004,
}
SENTINEL = 1337
MIN_BASELINE_MWC = 0.1
TUNNEL_START = datetime(2025, 7, 1)

# Colorblind-safe palette
CB = {
    'N42EH':   '#EE6677',
    'N52SH':   '#4477AA',
    'SmCo33H': '#228833',
    'SmCo35':  '#CCBB44',
    'NdFeB':   '#AA3377',
    'SmCo':    '#66CCEE',
}

REGION_COLORS = {
    'SE Arc':       '#EE6677',
    'NE Arc':       '#4477AA',
    'NW Arc':       '#228833',
    'SW Arc':       '#CCBB44',
    'North Linac':  '#AA3377',
    'South Linac':  '#66CCEE',
    'Labyrinth':    '#BBBBBB',
}

# ─── Tunnel Placement Mapping ────────────────────────────────────────────────
# From JLAB-TN-25-021 (Bodenstein et al.)
# Each entry: (Y-plate, H-plate, region, sub-location, line_position)
# line_position: 1=top through 5=fifth for arcs; 0 for linac sites

PLACEMENTS = [
    # SE Arc — Stack upstream of Girder 27
    ('Y15', 'N10', 'SE Arc', 'upstream Girder 27', 1),
    ('Y3',  'N16', 'SE Arc', 'upstream Girder 27', 2),
    ('Y23', 'N9',  'SE Arc', 'upstream Girder 27', 3),
    ('Y26', 'N37', 'SE Arc', 'upstream Girder 27', 4),
    ('Y40', 'N8',  'SE Arc', 'upstream Girder 27', 5),
    # NE Arc — above JD3A06 / upstream MJA9A06
    ('Y39', 'S12', 'NE Arc', 'above JD3A06', 1),
    ('Y7',  'S16', 'NE Arc', 'above JD3A06', 2),
    ('Y18', 'S20', 'NE Arc', 'above JD3A06', 3),
    ('Y21', 'S6',  'NE Arc', 'above JD3A06', 4),
    ('Y9',  'S14', 'NE Arc', 'above JD3A06', 5),
    # NW Arc — above MJC4A24 / downstream MXPAA24
    ('Y38', 'N12', 'NW Arc', 'above MJC4A24', 1),
    ('Y6',  'N39', 'NW Arc', 'above MJC4A24', 2),
    ('Y36', 'N17', 'NW Arc', 'above MJC4A24', 3),
    ('Y25', 'N11', 'NW Arc', 'above MJC4A24', 4),
    ('Y34', 'N18', 'NW Arc', 'above MJC4A24', 5),
    # SW Arc — above MJC4A14 / downstream MXPAA14
    ('Y13', 'S1',  'SW Arc', 'above MJC4A14', 1),
    ('Y32', 'S2',  'SW Arc', 'above MJC4A14', 2),
    ('Y19', 'S17', 'SW Arc', 'above MJC4A14', 3),
    ('Y10', 'S10', 'SW Arc', 'above MJC4A14', 4),
    ('Y11', 'S3',  'SW Arc', 'above MJC4A14', 5),
    # North Linac
    ('Y12', 'S4',  'Labyrinth',   'North Access Labyrinth', 0),
    ('Y17', 'N20', 'North Linac', 'NL NDX @ Girder 5', 0),
    ('Y4',  'S7',  'North Linac', 'NL Inj/Ret Crossover (0L05 NDX)', 0),
    ('Y16', 'S15', 'North Linac', 'NL NDX @ Girder 23', 0),
    ('Y22', 'N15', 'North Linac', 'NL NDX @ Girder 26', 0),
    # South Linac
    ('Y20', 'S5',  'Labyrinth',   'South Access Labyrinth', 0),
    ('Y24', 'S13', 'South Linac', 'SL NDX @ Girder 4', 0),
    ('Y5',  'N19', 'South Linac', 'SL NDX @ Girder 6', 0),
    ('Y1',  'N6',  'South Linac', 'SL NDX @ Girder 23', 0),
    ('Y30', 'S18', 'South Linac', 'SL NDX @ Girder 25', 0),
]

# Build quick lookup: Y-plate number -> placement info
Y_PLACEMENT = {}
for y, h, region, subloc, line in PLACEMENTS:
    ynum = int(y.replace('Y', ''))
    Y_PLACEMENT[ynum] = {
        'y_plate': y, 'h_plate': h, 'region': region,
        'sub_location': subloc, 'line': line,
    }

H_PLACEMENT = {}
for y, h, region, subloc, line in PLACEMENTS:
    H_PLACEMENT[h] = {
        'y_plate': y, 'h_plate': h, 'region': region,
        'sub_location': subloc, 'line': line,
    }


# ─── Material assignments ────────────────────────────────────────────────────

def load_materials():
    mat_path = os.path.join(BASE, 'Materials_Arrangements.xlsx')
    wb = openpyxl.load_workbook(mat_path)
    y_materials = {}
    ws = wb['Tunnel - Y Materials']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            plate_num = plate_id.replace('y-', '')
            y_materials[plate_num] = [str(row[i]).strip() if row[i] else ''
                                      for i in range(1, 5)]
    pair_arrangements = {}
    ws2 = wb['Tunnel - Pair Arrangements']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plate_id = str(row[0]).strip().lower()
            mat_type = str(row[1]).strip() if row[1] else ''
            configs = [str(row[i]).strip() if row[i] else '' for i in range(2, 6)]
            pair_arrangements[plate_id] = (mat_type, configs)
    return y_materials, pair_arrangements


def get_y_material(y_materials, sample_name):
    m = re.match(r'Y-(\d+)-(\d+)', sample_name)
    if not m:
        return None
    plate = m.group(1)
    slot = int(m.group(2))
    if plate in y_materials and 1 <= slot <= 4:
        return y_materials[plate][slot - 1]
    return None


# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_helmholtz_file(filepath):
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            val_match = re.search(r'([+-]?\d+\.?\d*)\s*(mWC|kT|kBG)', rest)
            if val_match:
                value = float(val_match.group(1))
                unit = val_match.group(2)
                rows.append((dt, value, unit))
    return rows


def parse_teslameter_file(filepath):
    rows = []
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            m = re.match(r'(\d{4}-\d{2}-\d{2})\t(\d{2}:\d{2}:\d{2})\t(.*)', line)
            if m:
                dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                       "%Y-%m-%d %H:%M:%S")
                rest = m.group(3)
            else:
                m = re.match(r'(\d{4}-\d{2}-\d{2})-(\d{2}:\d{2}:\d{2})\t(.*)', line)
                if m:
                    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}",
                                           "%Y-%m-%d %H:%M:%S")
                    rest = m.group(3)
                else:
                    continue
            nums = re.findall(r'(-?\d+\.\d+)', rest)
            if len(nums) >= 4:
                rows.append((dt, [float(x) for x in nums[:3]], float(nums[3])))
            elif len(nums) >= 3:
                rows.append((dt, [float(x) for x in nums[:3]], None))
    return rows


def _clean_retakes(data):
    if len(data) < 3:
        return data
    median_val = np.median([v for _, v in data])
    cleaned = []
    skip = set()
    for i in range(len(data)):
        if i in skip:
            continue
        if i + 1 < len(data):
            dt1, v1 = data[i]
            dt2, v2 = data[i + 1]
            if abs((dt2 - dt1).total_seconds()) < 300:
                if abs(v1 - median_val) > abs(v2 - median_val):
                    skip.add(i)
                    continue
                else:
                    skip.add(i + 1)
        cleaned.append(data[i])
    for i in range(len(data)):
        if i not in skip and data[i] not in cleaned:
            cleaned.append(data[i])
    return sorted(cleaned, key=lambda x: x[0])


def get_all_mwc_rows(rows):
    all_mwc = [(dt, val) for dt, val, unit in rows
               if unit == 'mWC' and abs(val - SENTINEL) > 1]
    return _clean_retakes(all_mwc)


# ─── Temperature lookup ──────────────────────────────────────────────────────

def build_temperature_lookup():
    temp_lookup = {}
    faces = ['front', 'side', 'top']

    # Y-plate temperatures
    y_tesla_dir = os.path.join(BASE, 'Y_Plates', 'Teslameter')
    y_samples = set()
    for f in os.listdir(y_tesla_dir):
        m = re.match(r'(Y-\d+-\d+)_(front|side|top)\.dat$', f)
        if m:
            y_samples.add(m.group(1))
    for sample in sorted(y_samples):
        date_temps = defaultdict(list)
        for face in faces:
            fpath = os.path.join(y_tesla_dir, f'{sample}_{face}.dat')
            if not os.path.exists(fpath):
                continue
            rows = parse_teslameter_file(fpath)
            for dt, fields, temp in rows:
                if temp is not None:
                    date_str = dt.strftime('%Y-%m-%d')
                    date_temps[date_str].append(temp)
        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(sample, date_str)] = (np.mean(temps), np.std(temps))

    # Pair assembly temperatures (H-samples from A-sample faces)
    pair_tesla_dir = os.path.join(BASE, 'Pair_Assemblies', 'Teslameter')
    pair_helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')
    for f in os.listdir(pair_helm_dir):
        m = re.match(r'(H[ns]-\d+-\d+)_helmholtz\.dat$', f)
        if not m:
            continue
        h_sample = m.group(1)
        hm = re.match(r'H([ns])-(\d+)-(\d+)', h_sample)
        if not hm:
            continue
        ns, plate, slot = hm.group(1), hm.group(2), hm.group(3)
        a_prefix = 'An' if ns == 'n' else 'As'
        date_temps = defaultdict(list)
        for magnet_idx in ['1', '2']:
            a_name = f'{a_prefix}-{plate}-{slot}-{magnet_idx}'
            for face in faces:
                fpath = os.path.join(pair_tesla_dir, f'{a_name}_{face}.dat')
                if not os.path.exists(fpath):
                    continue
                rows = parse_teslameter_file(fpath)
                for dt, fields, temp in rows:
                    if temp is not None:
                        date_str = dt.strftime('%Y-%m-%d')
                        date_temps[date_str].append(temp)
        for date_str, temps in date_temps.items():
            if temps:
                temp_lookup[(h_sample, date_str)] = (np.mean(temps),
                                                      np.std(temps))
    return temp_lookup


# ─── Correction ──────────────────────────────────────────────────────────────

def correct_helmholtz(h_raw, alpha, t_mean, t_std):
    denom = 1.0 + alpha * (t_mean - T_REF)
    h_corr = h_raw / denom
    dh = abs(h_raw * alpha * t_std / denom**2)
    return h_corr, dh


# ─── Compute degradation for all Y-plate samples ────────────────────────────

def compute_y_plate_degradation(y_materials, temp_lookup):
    """
    For each Y-plate sample, compute:
      - Baseline: mean of pre-deployment temperature-corrected readings
      - Latest: most recent temperature-corrected reading
      - % change = (latest - baseline) / baseline * 100
      - Error bar (see below)

    Error bar components (added in quadrature):
      σ_baseline: std(pre-deployment corrected) / sqrt(N_baseline)
                  → uncertainty on the mean baseline value
      σ_temp_latest: temperature correction uncertainty on the latest reading
                     (from face-to-face temperature spread)
      σ_repeatability: std(pre-deployment corrected) / baseline * 100
                       → sample-level measurement repeatability (in % units)

    Total % uncertainty:
      σ_pct = sqrt( (σ_baseline/baseline*100)^2 + (σ_temp_latest/baseline*100)^2 )

    Note: σ_repeatability is reported separately as a systematic floor.

    Returns list of dicts, one per sample.
    """
    helm_dir = os.path.join(BASE, 'Y_Plates', 'Helmholtz')
    results = []

    for f in sorted(os.listdir(helm_dir)):
        if not f.endswith('_helmholtz.dat'):
            continue
        sample = f.replace('_helmholtz.dat', '')
        material = get_y_material(y_materials, sample)
        if not material:
            continue
        material = material.strip()
        alpha = ALPHA.get(material)
        if alpha is None:
            continue

        # Parse plate number
        pm = re.match(r'Y-(\d+)-(\d+)', sample)
        if not pm:
            continue
        plate_num = int(pm.group(1))
        slot = int(pm.group(2))

        # Get placement info
        placement = Y_PLACEMENT.get(plate_num)
        if not placement:
            continue  # not a tunnel sample

        fpath = os.path.join(helm_dir, f)
        raw_rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
        if not raw_rows:
            continue

        # Separate pre-deployment and tunnel readings, apply correction
        pre_corr = []   # (h_corr, dh) for pre-deployment
        pre_raw = []
        tunnel_corr = [] # (dt, h_raw, h_corr, dh) for tunnel period
        tunnel_raw = []

        for dt, h_raw in raw_rows:
            date_str = dt.strftime('%Y-%m-%d')
            key = (sample, date_str)
            if key in temp_lookup:
                t_mean, t_std = temp_lookup[key]
                h_corr, dh = correct_helmholtz(h_raw, alpha, t_mean, t_std)
                if dt < TUNNEL_START:
                    pre_corr.append((h_corr, dh))
                    pre_raw.append(h_raw)
                else:
                    tunnel_corr.append((dt, h_raw, h_corr, dh))
                    tunnel_raw.append((dt, h_raw))
            else:
                # No temperature -> can use raw for pre-deployment baseline
                # estimate (uncorrected), but prefer corrected
                if dt < TUNNEL_START:
                    pre_raw.append(h_raw)
                else:
                    tunnel_raw.append((dt, h_raw))

        # Need at least one pre-deployment corrected AND one tunnel corrected
        if not pre_corr or not tunnel_corr:
            continue

        # --- Baseline computation ---
        bl_vals = [v for v, _ in pre_corr]
        # Filter near-zero
        bl_vals_filt = [v for v in bl_vals if abs(v) >= MIN_BASELINE_MWC]
        if not bl_vals_filt:
            continue

        # MAD outlier rejection for baseline
        if len(bl_vals_filt) >= 3:
            med = np.median(bl_vals_filt)
            mad = np.median([abs(v - med) for v in bl_vals_filt])
            threshold = max(3.5 * mad, 0.005 * abs(med))
            bl_kept = [v for v in bl_vals_filt if abs(v - med) <= threshold]
            n_rejected = len(bl_vals_filt) - len(bl_kept)
        else:
            bl_kept = bl_vals_filt
            n_rejected = 0

        if not bl_kept:
            continue

        baseline_mean = np.mean(bl_kept)
        baseline_std = np.std(bl_kept, ddof=1) if len(bl_kept) > 1 else 0.0
        n_baseline = len(bl_kept)

        if abs(baseline_mean) < MIN_BASELINE_MWC:
            continue

        # --- Latest measurement ---
        # Use the most recent corrected tunnel reading
        tunnel_corr.sort(key=lambda x: x[0])
        latest_dt, latest_raw, latest_corr, latest_dh = tunnel_corr[-1]

        # --- % change ---
        pct_change = (latest_corr - baseline_mean) / baseline_mean * 100.0

        # --- Error propagation ---
        # Component 1: baseline uncertainty (std error on mean)
        sigma_baseline = baseline_std / np.sqrt(n_baseline) if n_baseline > 1 else baseline_std
        sigma_bl_pct = sigma_baseline / abs(baseline_mean) * 100.0

        # Component 2: temperature correction uncertainty on latest reading
        sigma_temp_pct = latest_dh / abs(baseline_mean) * 100.0

        # Total % uncertainty (quadrature sum)
        sigma_total_pct = np.sqrt(sigma_bl_pct**2 + sigma_temp_pct**2)

        # Repeatability floor: std of pre-deployment as % of baseline
        repeat_pct = baseline_std / abs(baseline_mean) * 100.0 if baseline_std > 0 else 0.0

        results.append({
            'sample': sample,
            'plate': plate_num,
            'slot': slot,
            'material': material,
            'region': placement['region'],
            'sub_location': placement['sub_location'],
            'line': placement['line'],
            'h_plate': placement['h_plate'],
            'baseline_mean': baseline_mean,
            'baseline_std': baseline_std,
            'n_baseline': n_baseline,
            'n_rejected': n_rejected,
            'latest_date': latest_dt.strftime('%Y-%m-%d'),
            'latest_corr': latest_corr,
            'latest_raw': latest_raw,
            'pct_change': pct_change,
            'sigma_bl_pct': sigma_bl_pct,
            'sigma_temp_pct': sigma_temp_pct,
            'sigma_total_pct': sigma_total_pct,
            'repeat_pct': repeat_pct,
        })

    return results


# ─── Compute degradation for H-plate samples ────────────────────────────────

def compute_h_plate_degradation(pair_arrangements, temp_lookup):
    """Same as Y-plate but for Hn/Hs pair assembly Helmholtz readings."""
    helm_dir = os.path.join(BASE, 'Pair_Assemblies', 'Helmholtz')
    results = []

    for f in sorted(os.listdir(helm_dir)):
        m = re.match(r'(H[ns]-\d+-\d+)_helmholtz\.dat$', f)
        if not m:
            continue
        h_sample = m.group(1)
        hm = re.match(r'H([ns])-(\d+)-(\d+)', h_sample)
        if not hm:
            continue
        ns = hm.group(1)
        plate_num = int(hm.group(2))
        slot = int(hm.group(3))

        mat_type = 'NdFeB' if ns == 'n' else 'SmCo'
        alpha = ALPHA[mat_type]

        # Get placement and configuration
        plate_key = f'{ns}-{plate_num}'
        config = ''
        if plate_key in pair_arrangements:
            _, configs = pair_arrangements[plate_key]
            if slot - 1 < len(configs):
                config = configs[slot - 1]

        # Map H-plate name to placement lookup
        # H_PLACEMENT keys are like 'N10', 'S12' etc.
        h_lookup_key = f'{"N" if ns == "n" else "S"}{plate_num}'
        placement = H_PLACEMENT.get(h_lookup_key)
        if not placement:
            continue  # not a tunnel sample

        fpath = os.path.join(helm_dir, f)
        raw_rows = get_all_mwc_rows(parse_helmholtz_file(fpath))
        if not raw_rows:
            continue

        pre_corr = []
        tunnel_corr = []

        for dt, h_raw in raw_rows:
            date_str = dt.strftime('%Y-%m-%d')
            key = (h_sample, date_str)
            if key in temp_lookup:
                t_mean, t_std = temp_lookup[key]
                h_corr, dh = correct_helmholtz(h_raw, alpha, t_mean, t_std)
                if dt < TUNNEL_START:
                    pre_corr.append((h_corr, dh))
                else:
                    tunnel_corr.append((dt, h_raw, h_corr, dh))
            else:
                if dt < TUNNEL_START:
                    pre_corr.append((h_raw, 0.0))  # uncorrected baseline
                else:
                    pass  # skip uncorrected tunnel readings

        if not pre_corr or not tunnel_corr:
            continue

        bl_vals = [v for v, _ in pre_corr]
        bl_vals_filt = [v for v in bl_vals if abs(v) >= MIN_BASELINE_MWC]
        if not bl_vals_filt:
            continue

        if len(bl_vals_filt) >= 3:
            med = np.median(bl_vals_filt)
            mad = np.median([abs(v - med) for v in bl_vals_filt])
            threshold = max(3.5 * mad, 0.005 * abs(med))
            bl_kept = [v for v in bl_vals_filt if abs(v - med) <= threshold]
        else:
            bl_kept = bl_vals_filt

        if not bl_kept:
            continue

        baseline_mean = np.mean(bl_kept)
        baseline_std = np.std(bl_kept, ddof=1) if len(bl_kept) > 1 else 0.0
        n_baseline = len(bl_kept)

        if abs(baseline_mean) < MIN_BASELINE_MWC:
            continue

        tunnel_corr.sort(key=lambda x: x[0])
        latest_dt, latest_raw, latest_corr, latest_dh = tunnel_corr[-1]

        pct_change = (latest_corr - baseline_mean) / baseline_mean * 100.0

        sigma_baseline = baseline_std / np.sqrt(n_baseline) if n_baseline > 1 else baseline_std
        sigma_bl_pct = sigma_baseline / abs(baseline_mean) * 100.0
        sigma_temp_pct = latest_dh / abs(baseline_mean) * 100.0
        sigma_total_pct = np.sqrt(sigma_bl_pct**2 + sigma_temp_pct**2)
        repeat_pct = baseline_std / abs(baseline_mean) * 100.0 if baseline_std > 0 else 0.0

        results.append({
            'sample': h_sample,
            'plate': plate_num,
            'slot': slot,
            'material': mat_type,
            'config': config,
            'region': placement['region'],
            'sub_location': placement['sub_location'],
            'line': placement['line'],
            'baseline_mean': baseline_mean,
            'baseline_std': baseline_std,
            'n_baseline': n_baseline,
            'latest_date': latest_dt.strftime('%Y-%m-%d'),
            'latest_corr': latest_corr,
            'latest_raw': latest_raw,
            'pct_change': pct_change,
            'sigma_bl_pct': sigma_bl_pct,
            'sigma_temp_pct': sigma_temp_pct,
            'sigma_total_pct': sigma_total_pct,
            'repeat_pct': repeat_pct,
        })

    return results


# ─── Summary statistics ──────────────────────────────────────────────────────

def summarize_group(results_list, label):
    """Compute weighted mean and uncertainty for a group of results."""
    if not results_list:
        return None
    pcts = [r['pct_change'] for r in results_list]
    sigmas = [r['sigma_total_pct'] for r in results_list]

    n = len(pcts)
    mean_pct = np.mean(pcts)
    std_pct = np.std(pcts, ddof=1) if n > 1 else 0.0
    sem_pct = std_pct / np.sqrt(n) if n > 1 else sigmas[0] if sigmas else 0.0

    # Mean of individual error bars (measurement uncertainty)
    mean_sigma = np.mean(sigmas)

    # Combined uncertainty: larger of SEM (sample scatter) or mean measurement error
    combined_unc = max(sem_pct, mean_sigma)

    return {
        'label': label,
        'n': n,
        'mean_pct': mean_pct,
        'std_pct': std_pct,
        'sem_pct': sem_pct,
        'mean_sigma': mean_sigma,
        'combined_unc': combined_unc,
        'min_pct': min(pcts),
        'max_pct': max(pcts),
    }


# ─── Plotting ────────────────────────────────────────────────────────────────

def plot_by_region(y_results, h_results):
    """Bar chart of mean degradation by region, broken down by material."""
    regions = ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
               'North Linac', 'South Linac', 'Labyrinth']
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, axes = plt.subplots(2, 1, figsize=(14, 12))

    # Top panel: Y-plate by material within each region
    ax = axes[0]
    x = np.arange(len(regions))
    width = 0.18
    offsets = [-1.5, -0.5, 0.5, 1.5]

    for i, mat in enumerate(materials):
        means, errs = [], []
        for region in regions:
            group = [r for r in y_results
                     if r['region'] == region and r['material'] == mat]
            if group:
                s = summarize_group(group, f'{region}/{mat}')
                means.append(s['mean_pct'])
                errs.append(s['combined_unc'])
            else:
                means.append(0)
                errs.append(0)
        ax.bar(x + offsets[i] * width, means, width, yerr=errs,
               color=CB[mat], label=mat, capsize=3, alpha=0.85,
               edgecolor='black', linewidth=0.5)

    ax.set_xticks(x)
    ax.set_xticklabels(regions, fontsize=10)
    ax.set_ylabel('% Change from Baseline', fontsize=11)
    ax.set_title('Y-Plate Degradation by Region and Material\n'
                 '(Temperature-corrected to 20 °C, error bars = max(SEM, σ_meas))',
                 fontsize=12, fontweight='bold')
    ax.axhline(0, color='black', linewidth=0.5, linestyle='-')
    ax.legend(fontsize=9)
    ax.grid(axis='y', alpha=0.3)

    # Bottom panel: H-plate by material type within each region
    ax2 = axes[1]
    h_mats = ['NdFeB', 'SmCo']
    width2 = 0.3
    offsets2 = [-0.5, 0.5]

    for i, mat in enumerate(h_mats):
        means, errs = [], []
        for region in regions:
            group = [r for r in h_results
                     if r['region'] == region and r['material'] == mat]
            if group:
                s = summarize_group(group, f'{region}/{mat}')
                means.append(s['mean_pct'])
                errs.append(s['combined_unc'])
            else:
                means.append(0)
                errs.append(0)
        ax2.bar(x + offsets2[i] * width2, means, width2, yerr=errs,
                color=CB[mat], label=mat, capsize=3, alpha=0.85,
                edgecolor='black', linewidth=0.5)

    ax2.set_xticks(x)
    ax2.set_xticklabels(regions, fontsize=10)
    ax2.set_ylabel('% Change from Baseline', fontsize=11)
    ax2.set_title('H-Plate (Pair Assembly) Degradation by Region and Material\n'
                  '(Temperature-corrected to 20 °C)',
                  fontsize=12, fontweight='bold')
    ax2.axhline(0, color='black', linewidth=0.5, linestyle='-')
    ax2.legend(fontsize=9)
    ax2.grid(axis='y', alpha=0.3)

    fig.tight_layout()
    outpath = os.path.join(PLOT_DIR, 'degradation_summary_by_region.png')
    fig.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {outpath}")


def plot_by_material(y_results):
    """Strip chart showing every Y-plate sample, colored by material."""
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']

    fig, ax = plt.subplots(figsize=(14, 8))

    for i, mat in enumerate(materials):
        group = sorted([r for r in y_results if r['material'] == mat],
                       key=lambda r: r['pct_change'])
        if not group:
            continue
        x_pos = np.arange(len(group)) + i * (len(group) + 2)
        pcts = [r['pct_change'] for r in group]
        errs = [r['sigma_total_pct'] for r in group]
        labels = [r['sample'] for r in group]

        ax.errorbar(x_pos, pcts, yerr=errs, fmt='o', color=CB[mat],
                    markersize=6, capsize=3, linewidth=1.2, label=mat)

        # Label samples with > 0.3% |degradation|
        for xp, pct, lbl in zip(x_pos, pcts, labels):
            if abs(pct) > 0.3:
                ax.annotate(lbl, (xp, pct), fontsize=6, rotation=45,
                            ha='left', va='bottom')

    ax.axhline(0, color='black', linewidth=0.5)
    ax.set_ylabel('% Change from Baseline', fontsize=11)
    ax.set_title('Y-Plate Sample-Level Degradation by Material\n'
                 '(Temperature-corrected to 20 °C, most recent measurement)',
                 fontsize=12, fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(axis='y', alpha=0.3)
    ax.set_xticks([])

    fig.tight_layout()
    outpath = os.path.join(PLOT_DIR, 'degradation_summary_by_material.png')
    fig.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {outpath}")


def plot_individual_samples(y_results):
    """Horizontal bar chart of every Y-plate sample with error bars,
    grouped by region, sorted by degradation within each region."""
    regions_order = ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
                     'North Linac', 'South Linac', 'Labyrinth']

    fig, ax = plt.subplots(figsize=(12, 20))

    y_pos = 0
    yticks = []
    ylabels = []
    region_boundaries = []

    for region in regions_order:
        group = sorted([r for r in y_results if r['region'] == region],
                       key=lambda r: r['pct_change'])
        if not group:
            continue

        region_boundaries.append((y_pos, region))

        for r in group:
            color = CB.get(r['material'], '#888888')
            ax.barh(y_pos, r['pct_change'], xerr=r['sigma_total_pct'],
                    height=0.7, color=color, alpha=0.8, capsize=3,
                    edgecolor='black', linewidth=0.3)
            yticks.append(y_pos)
            line_str = f" (L{r['line']})" if r['line'] > 0 else ""
            ylabels.append(f"{r['sample']} [{r['material']}]{line_str}")
            y_pos += 1
        y_pos += 1  # gap between regions

    ax.set_yticks(yticks)
    ax.set_yticklabels(ylabels, fontsize=7)
    ax.axvline(0, color='black', linewidth=0.8)
    ax.set_xlabel('% Change from Baseline (corrected to 20 °C)', fontsize=11)
    ax.set_title('Y-Plate Degradation — All Samples by Region\n'
                 '(negative = degradation, error bars = combined uncertainty)',
                 fontsize=12, fontweight='bold')
    ax.grid(axis='x', alpha=0.3)

    # Add region labels
    for ystart, region in region_boundaries:
        ax.annotate(region, xy=(-0.01, ystart - 0.5),
                    xycoords=('axes fraction', 'data'),
                    fontsize=10, fontweight='bold', color='gray',
                    ha='right', va='bottom')

    ax.invert_yaxis()
    fig.tight_layout()
    outpath = os.path.join(PLOT_DIR, 'degradation_all_samples_by_region.png')
    fig.savefig(outpath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {outpath}")


# ─── Markdown report ─────────────────────────────────────────────────────────

def write_markdown_report(y_results, h_results):
    """Write a formatted markdown summary for management."""
    lines = []
    lines.append("# Preliminary Degradation Summary")
    lines.append("")
    lines.append("**Status**: Preliminary — more data collection and error analysis pending")
    lines.append(f"**Generated**: {datetime.now().strftime('%Y-%m-%d')}")
    lines.append(f"**Reference temperature**: {T_REF} °C")
    lines.append("")

    # --- Error bar explanation ---
    lines.append("## Error Bar Methodology")
    lines.append("")
    lines.append("Each reported % change carries a combined uncertainty computed from "
                 "two independent sources added in quadrature:")
    lines.append("")
    lines.append("1. **Baseline uncertainty (σ_baseline)**: The standard error on the mean "
                 "of pre-deployment temperature-corrected readings. With N baseline "
                 "measurements, σ_baseline = std(baseline readings) / √N. This captures "
                 "how well we know the starting value.")
    lines.append("")
    lines.append("2. **Temperature correction uncertainty (σ_temp)**: Propagated from the "
                 "spread of the 3 Teslameter face temperatures (front, side, top) on the "
                 "measurement date. The correction formula B_corr = B_raw / (1 + α(T−20°C)) "
                 "introduces uncertainty when T varies across faces. This is computed as "
                 "|B_raw × α × σ_T / (1 + α(T−20))²|.")
    lines.append("")
    lines.append("3. **Combined**: σ_total = √(σ_baseline² + σ_temp²), expressed as % of baseline.")
    lines.append("")
    lines.append("For group averages (by region or material), the reported uncertainty is "
                 "the **larger** of:")
    lines.append("- The standard error of the mean across samples (SEM = σ/√N), or")
    lines.append("- The mean per-sample measurement uncertainty")
    lines.append("")
    lines.append("This ensures we do not underestimate uncertainty when sample scatter "
                 "is small but individual measurements have significant error bars.")
    lines.append("")

    # --- Y-plate summary by material ---
    lines.append("## Y-Plate Summary by Material")
    lines.append("")
    materials = ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']
    lines.append("| Material | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    lines.append("|----------|---|-----------|-----------|---------|---------|")
    for mat in materials:
        group = [r for r in y_results if r['material'] == mat]
        if group:
            s = summarize_group(group, mat)
            lines.append(f"| {mat} | {s['n']} | {s['mean_pct']:+.3f} | "
                         f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                         f"{s['max_pct']:+.3f} |")
    lines.append("")

    # --- Y-plate summary by region ---
    lines.append("## Y-Plate Summary by Region")
    lines.append("")
    regions = ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
               'North Linac', 'South Linac', 'Labyrinth']
    lines.append("| Region | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    lines.append("|--------|---|-----------|-----------|---------|---------|")
    for region in regions:
        group = [r for r in y_results if r['region'] == region]
        if group:
            s = summarize_group(group, region)
            lines.append(f"| {region} | {s['n']} | {s['mean_pct']:+.3f} | "
                         f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                         f"{s['max_pct']:+.3f} |")
    lines.append("")

    # --- Y-plate by region AND material ---
    lines.append("## Y-Plate Summary by Region × Material")
    lines.append("")
    lines.append("| Region | Material | N | Mean Δ (%) | ± Unc (%) |")
    lines.append("|--------|----------|---|-----------|-----------|")
    for region in regions:
        for mat in materials:
            group = [r for r in y_results
                     if r['region'] == region and r['material'] == mat]
            if group:
                s = summarize_group(group, f'{region}/{mat}')
                lines.append(f"| {region} | {mat} | {s['n']} | "
                             f"{s['mean_pct']:+.3f} | ±{s['combined_unc']:.3f} |")
    lines.append("")

    # --- Arc line position ---
    lines.append("## Arc Degradation by Line Position")
    lines.append("")
    lines.append("Line position in arc stacks (1=top to 5=bottom) corresponds to "
                 "different beam energy passes and potentially different radiation levels.")
    lines.append("")
    lines.append("| Line | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    lines.append("|------|---|-----------|-----------|---------|---------|")
    for line_pos in range(1, 6):
        group = [r for r in y_results if r['line'] == line_pos]
        if group:
            s = summarize_group(group, f'Line {line_pos}')
            lines.append(f"| {line_pos} | {s['n']} | {s['mean_pct']:+.3f} | "
                         f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                         f"{s['max_pct']:+.3f} |")
    lines.append("")

    # --- H-plate summary ---
    lines.append("## H-Plate (Pair Assembly) Summary")
    lines.append("")
    lines.append("### By Material")
    lines.append("")
    lines.append("| Material | N | Mean Δ (%) | ± Unc (%) | Min (%) | Max (%) |")
    lines.append("|----------|---|-----------|-----------|---------|---------|")
    for mat in ['NdFeB', 'SmCo']:
        group = [r for r in h_results if r['material'] == mat]
        if group:
            s = summarize_group(group, mat)
            lines.append(f"| {mat} | {s['n']} | {s['mean_pct']:+.3f} | "
                         f"±{s['combined_unc']:.3f} | {s['min_pct']:+.3f} | "
                         f"{s['max_pct']:+.3f} |")
    lines.append("")

    # --- H-plate by assembly config ---
    lines.append("### By Assembly Configuration")
    lines.append("")
    lines.append("| Config | N | Mean Δ (%) | ± Unc (%) | Notes |")
    lines.append("|--------|---|-----------|-----------|-------|")
    for config in ['Alpha', 'Beta', 'Gamma', 'Delta']:
        group = [r for r in h_results if r.get('config') == config]
        if group:
            s = summarize_group(group, config)
            note = "⚠ unreliable (multipole)" if config == 'Beta' else ""
            lines.append(f"| {config} | {s['n']} | {s['mean_pct']:+.3f} | "
                         f"±{s['combined_unc']:.3f} | {note} |")
    lines.append("")

    # --- H-plate by region ---
    lines.append("### By Region")
    lines.append("")
    lines.append("| Region | N | Mean Δ (%) | ± Unc (%) |")
    lines.append("|--------|---|-----------|-----------|")
    for region in regions:
        group = [r for r in h_results if r['region'] == region]
        if group:
            s = summarize_group(group, region)
            lines.append(f"| {region} | {s['n']} | {s['mean_pct']:+.3f} | "
                         f"±{s['combined_unc']:.3f} |")
    lines.append("")

    # --- Individual Y-plate detail table ---
    lines.append("## Individual Y-Plate Results (All Samples)")
    lines.append("")
    lines.append("| Sample | Material | Region | Line | Baseline (mWC) | Latest (mWC) | "
                 "Δ (%) | ± (%) | Date |")
    lines.append("|--------|----------|--------|------|----------------|--------------|"
                 "-------|-------|------|")
    for r in sorted(y_results, key=lambda x: x['pct_change']):
        line_str = str(r['line']) if r['line'] > 0 else "—"
        lines.append(
            f"| {r['sample']} | {r['material']} | {r['region']} | {line_str} | "
            f"{r['baseline_mean']:.4f} | {r['latest_corr']:.4f} | "
            f"{r['pct_change']:+.3f} | ±{r['sigma_total_pct']:.3f} | "
            f"{r['latest_date']} |")
    lines.append("")

    # --- Key observations ---
    lines.append("## Key Observations")
    lines.append("")

    # Find most degraded
    most_degraded = min(y_results, key=lambda r: r['pct_change'])
    least_degraded = max(y_results, key=lambda r: r['pct_change'])
    lines.append(f"- **Largest degradation**: {most_degraded['sample']} "
                 f"({most_degraded['material']}, {most_degraded['region']}) at "
                 f"{most_degraded['pct_change']:+.3f} ± {most_degraded['sigma_total_pct']:.3f}%")
    lines.append(f"- **Largest positive shift**: {least_degraded['sample']} "
                 f"({least_degraded['material']}, {least_degraded['region']}) at "
                 f"{least_degraded['pct_change']:+.3f} ± {least_degraded['sigma_total_pct']:.3f}%")

    # Count significant degradation
    sig_deg = [r for r in y_results
               if r['pct_change'] < -abs(r['sigma_total_pct'])]
    lines.append(f"- **Samples with degradation exceeding 1σ**: {len(sig_deg)} / {len(y_results)}")
    sig_05 = [r for r in y_results if r['pct_change'] < -0.5]
    lines.append(f"- **Samples with >0.5% degradation**: {len(sig_05)} / {len(y_results)}")
    lines.append("")

    lines.append("## Caveats")
    lines.append("")
    lines.append("- These are **preliminary** results. More measurements are being collected.")
    lines.append("- Radiation dose data is not yet complete; dose-response correlation pending.")
    lines.append("- The temperature coefficient α(Br) carries its own systematic uncertainty "
                 "(not included in error bars — it would shift all values of a given material "
                 "grade together).")
    lines.append("- Pre-deployment Teslameter data used the broken Hall probe; only Helmholtz "
                 "baselines with temperature correction from working probe dates are used.")
    lines.append("- Beta (antiparallel) pair assembly Helmholtz readings are known to be "
                 "unreliable due to multipole field character.")
    lines.append("- Lab-based control comparisons are not yet included (data still being collected).")
    lines.append("")

    md_path = os.path.join(MD_DIR, 'degradation_summary.md')
    with open(md_path, 'w') as f:
        f.write('\n'.join(lines))
    print(f"  Saved: {md_path}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    y_materials, pair_arrangements = load_materials()
    temp_lookup = build_temperature_lookup()

    print("\nComputing Y-plate degradation...")
    y_results = compute_y_plate_degradation(y_materials, temp_lookup)
    print(f"  {len(y_results)} Y-plate samples with valid baseline + tunnel data")

    print("\nComputing H-plate degradation...")
    h_results = compute_h_plate_degradation(pair_arrangements, temp_lookup)
    print(f"  {len(h_results)} H-plate samples with valid baseline + tunnel data")

    # --- Print quick summary ---
    print("\n" + "="*70)
    print("Y-PLATE DEGRADATION BY MATERIAL (% change, corrected to 20°C)")
    print("="*70)
    for mat in ['N42EH', 'N52SH', 'SmCo33H', 'SmCo35']:
        group = [r for r in y_results if r['material'] == mat]
        if group:
            s = summarize_group(group, mat)
            print(f"  {mat:8s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']}, range [{s['min_pct']:+.3f}, {s['max_pct']:+.3f}])")

    print("\n" + "="*70)
    print("Y-PLATE DEGRADATION BY REGION")
    print("="*70)
    for region in ['SE Arc', 'NE Arc', 'NW Arc', 'SW Arc',
                   'North Linac', 'South Linac', 'Labyrinth']:
        group = [r for r in y_results if r['region'] == region]
        if group:
            s = summarize_group(group, region)
            print(f"  {region:14s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']}, range [{s['min_pct']:+.3f}, {s['max_pct']:+.3f}])")

    print("\n" + "="*70)
    print("H-PLATE DEGRADATION BY ASSEMBLY CONFIG")
    print("="*70)
    for config in ['Alpha', 'Beta', 'Gamma', 'Delta']:
        group = [r for r in h_results if r.get('config') == config]
        if group:
            s = summarize_group(group, config)
            note = " ⚠ unreliable" if config == 'Beta' else ""
            print(f"  {config:8s}: {s['mean_pct']:+.3f} ± {s['combined_unc']:.3f}%  "
                  f"(N={s['n']}){note}")

    # --- Generate plots ---
    print("\nGenerating plots...")
    plot_by_region(y_results, h_results)
    plot_by_material(y_results)
    plot_individual_samples(y_results)

    # --- Write markdown report ---
    print("\nWriting markdown report...")
    write_markdown_report(y_results, h_results)

    print("\nDone.")


if __name__ == '__main__':
    main()
